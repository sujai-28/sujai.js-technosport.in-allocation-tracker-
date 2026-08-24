"""
process_inwards.py
==================
Reads all CSV / XLSX files from:
    D:\INCREFF ORDER PUNCH\inward details   (all sub-folders)

Aggregates the data (groups by GRN Date, Style, Category, Color, Size, MRP)
and writes a single clean Excel file to:
    D:\INCREFF ORDER PUNCH\inward details output\Inwards_Processed.xlsx

Run this script whenever new inward files are added to the source folder.
The portal (Inwards Details tab) reads from the output file directly.
"""

import os
import glob
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# ── Paths ──────────────────────────────────────────────────────────────────
SOURCE_DIR = r"D:\INCREFF ORDER PUNCH\inward details"
OUTPUT_DIR = r"D:\INCREFF ORDER PUNCH\inward details output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Inwards_Processed.xlsx")

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Column name aliases ─────────────────────────────────────────────────────
COL_ALIASES = {
    "GRN Date":          ["grn date"],
    "SKU":               ["client sku id / ean", "client sku id/ean", "sku", "ean",
                          "barcode", "client_sku_id"],
    "Category":          ["category", "cat"],
    "Style":             ["style", "style id", "style_id"],
    "Size":              ["size"],
    "Color":             ["color", "colour"],
    "MRP":               ["mrp", "price", "selling price"],
    "Received Qty":      ["recieved qty", "received qty", "received_qty",
                          "qty", "quantity", "inward qty"],
}

def find_col(columns, aliases):
    """Return the first column name that matches any alias (case-insensitive)."""
    col_lower = {c.strip().lower(): c for c in columns}
    for alias in aliases:
        if alias.lower() in col_lower:
            return col_lower[alias.lower()]
    return None

# ── Read all source files ───────────────────────────────────────────────────
print("=" * 60)
print("  Inwards Processor  –  Increff Order Punch")
print("=" * 60)
print(f"\nSource folder : {SOURCE_DIR}")
print(f"Output file   : {OUTPUT_FILE}\n")

all_frames = []
files_found = 0

for root, dirs, files in os.walk(SOURCE_DIR):
    for fname in sorted(files):
        if fname.startswith("~"):
            continue
        fpath = os.path.join(root, fname)
        try:
            if fname.lower().endswith(".csv"):
                df_tmp = pd.read_csv(fpath, encoding="utf-8-sig")
            elif fname.lower().endswith((".xlsx", ".xls")):
                df_tmp = pd.read_excel(fpath)
            else:
                continue

            df_tmp.columns = [str(c).strip() for c in df_tmp.columns]
            subfolder = os.path.relpath(root, SOURCE_DIR)
            df_tmp["_source_file"]   = fname
            df_tmp["_source_folder"] = subfolder if subfolder != "." else "root"
            all_frames.append(df_tmp)
            files_found += 1
            print(f"  OK  {subfolder}/{fname}  ({len(df_tmp):,} rows)")
        except Exception as e:
            print(f"  ERR {fname} - ERROR: {e}")

if not all_frames:
    print("\nNo files found. Exiting.")
    exit(1)

print(f"\nTotal files loaded: {files_found}")

# ── Combine ─────────────────────────────────────────────────────────────────
df = pd.concat(all_frames, ignore_index=True)
print(f"Total raw rows : {len(df):,}")

# ── Map columns ─────────────────────────────────────────────────────────────
rename_map = {}
for standard, aliases in COL_ALIASES.items():
    found = find_col(df.columns, aliases)
    if found and found not in rename_map.values():
        rename_map[found] = standard

df.rename(columns=rename_map, inplace=True)

for col in list(COL_ALIASES.keys()) + ["_source_file", "_source_folder"]:
    if col not in df.columns:
        df[col] = ""

# ── Parse & clean ───────────────────────────────────────────────────────────
# GRN Date
df["GRN Date"] = pd.to_datetime(df["GRN Date"], dayfirst=True, errors="coerce")
df["GRN Date Display"] = df["GRN Date"].apply(
    lambda x: x.strftime("%d-%b-%Y") if pd.notna(x) else ""
)
df["GRN Date ISO"] = df["GRN Date"].apply(
    lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else ""
)

# Numerics
df["Received Qty"] = pd.to_numeric(df["Received Qty"], errors="coerce").fillna(0).astype(int)
df["MRP"]         = pd.to_numeric(df["MRP"],          errors="coerce").fillna(0).round(0).astype(int)

# Strings
for col in ["SKU", "Category", "Style", "Size", "Color"]:
    df[col] = df[col].astype(str).str.strip()
    df[col] = df[col].replace({"nan": "", "NaN": "", "None": ""})

# ── Aggregate (pivot-style) ──────────────────────────────────────────────────
# Raw sheet: all rows cleaned up
raw_cols = ["GRN Date Display", "GRN Date ISO", "SKU", "Category",
            "Style", "Size", "Color", "MRP", "Received Qty",
            "_source_folder", "_source_file"]
df_raw = df[raw_cols].copy()
df_raw.rename(columns={
    "GRN Date Display": "GRN Date",
    "_source_folder":   "Source Folder",
    "_source_file":     "Source File",
}, inplace=True)
df_raw.sort_values(["GRN Date ISO", "Style"], inplace=True)

# Summary sheet: grouped by Date + Style + Category + Color + Size + MRP
df_agg = (
    df.groupby(
        ["GRN Date ISO", "GRN Date Display", "Category", "Style", "Size", "Color", "MRP"],
        dropna=False
    )["Received Qty"]
    .sum()
    .reset_index()
)
df_agg.rename(columns={"GRN Date Display": "GRN Date"}, inplace=True)
df_agg.sort_values(["GRN Date ISO", "Style"], inplace=True)
df_agg = df_agg[["GRN Date", "GRN Date ISO", "Category", "Style", "Size", "Color", "MRP", "Received Qty"]]

# Category pivot
df_cat = (
    df.groupby("Category")["Received Qty"]
    .agg(Total_Qty="sum", SKU_Count="count")
    .reset_index()
    .sort_values("Total_Qty", ascending=False)
)
df_cat.columns = ["Category", "Total Received Qty", "Row Count"]

# Style pivot
df_style = (
    df.groupby("Style")["Received Qty"]
    .sum()
    .reset_index()
    .sort_values("Received Qty", ascending=False)
    .head(200)
)
df_style.columns = ["Style", "Total Received Qty"]

print(f"Aggregated rows: {len(df_agg):,}")
print(f"Unique styles  : {df['Style'].nunique():,}")
print(f"Unique SKUs    : {df['SKU'].nunique():,}")
print(f"Unique cats    : {df['Category'].nunique():,}")
total_qty = int(df["Received Qty"].sum())
print(f"Total qty      : {total_qty:,}")

# ── Write Excel ──────────────────────────────────────────────────────────────
print(f"\nWriting to: {OUTPUT_FILE}")

DARK  = "1F2937"
BLUE  = "3B82F6"
GREEN = "10B981"
PURP  = "7C3AED"
AMBER = "F59E0B"
WHITE = "FFFFFF"
LIGHT = "F3F4F6"

def style_header(ws, header_color=DARK):
    """Apply dark header styling to row 1."""
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border

def auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

def zebra_rows(ws, start_row=2, even_color="EEF2FF"):
    """Alternating row colors."""
    fill_even = PatternFill(start_color=even_color, end_color=even_color, fill_type="solid")
    for i, row in enumerate(ws.iter_rows(min_row=start_row), start=0):
        for cell in row:
            if i % 2 == 0:
                cell.fill = fill_even
            cell.alignment = Alignment(vertical="center")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    # Sheet 1: Summary (aggregated)
    df_agg.to_excel(writer, sheet_name="Summary", index=False)
    ws_sum = writer.sheets["Summary"]
    style_header(ws_sum, DARK)
    zebra_rows(ws_sum)
    auto_width(ws_sum)
    ws_sum.freeze_panes = "A2"

    # Sheet 2: Raw Data
    df_raw.to_excel(writer, sheet_name="Raw Data", index=False)
    ws_raw = writer.sheets["Raw Data"]
    style_header(ws_raw, DARK)
    zebra_rows(ws_raw)
    auto_width(ws_raw)
    ws_raw.freeze_panes = "A2"

    # Sheet 3: Category Pivot
    df_cat.to_excel(writer, sheet_name="Category Pivot", index=False)
    ws_cat = writer.sheets["Category Pivot"]
    style_header(ws_cat, GREEN)
    zebra_rows(ws_cat)
    auto_width(ws_cat)

    # Sheet 4: Style Pivot (top 200)
    df_style.to_excel(writer, sheet_name="Style Pivot (Top 200)", index=False)
    ws_sty = writer.sheets["Style Pivot (Top 200)"]
    style_header(ws_sty, PURP)
    zebra_rows(ws_sty)
    auto_width(ws_sty)

print(f"\nDone! Output saved to:\n    {OUTPUT_FILE}")
print(f"\nSummary:")
print(f"  Rows in Summary sheet : {len(df_agg):,}")
print(f"  Rows in Raw Data sheet: {len(df_raw):,}")
print(f"  Total Received Qty    : {total_qty:,}")
print(f"  Generated at          : {datetime.datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
print("=" * 60)
