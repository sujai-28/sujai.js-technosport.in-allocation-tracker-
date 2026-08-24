#!/usr/bin/env python
"""
top_100_sales.py
----------------
Generates a Top 100 Sales Styles report, detailing:
  - Top 100 styles by total sales quantity.
  - Overall sales for each style in the last 10 days and last 30 days.
  - Store-wise sales breakdown for each style (if transaction-level data is provided).
  - Valid Options and No. of Stores Live (combining Stock, Transit, and Allocation).
  - OMS reservation pools.

Inputs:
  - Sales Excel or CSV file (either transactional or style-level aggregated).
  - EBO Current Stock, Transit Stock, and Allocation Stock sheets.
  - Validation Output LATEST.

Outputs:
  - A styled multi-sheet Excel file saved under "ebo stock track data/output of top 100".
"""

import os
import sys
import glob
import argparse
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────
#  NORMALIZATION HELPERS
# ──────────────────────────────────────────────────────────
def clean_store_code(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        return s[:-2]
    return s

def find_column(columns, candidates):
    for cand in candidates:
        cand_clean = cand.strip().lower().replace("_", " ").replace("  ", " ")
        for col in columns:
            col_clean = str(col).strip().lower().replace("_", " ").replace("  ", " ")
            if col_clean == cand_clean:
                return col
    return None

def clean_currency(val):
    if pd.isna(val):
        return 0
    if isinstance(val, (int, float)):
        return val
    # Remove currency symbol, commas, and strip whitespace
    s = str(val).replace('₹', '').replace(',', '').strip()
    if not s or s.lower() in ('#n/a', 'n/a', 'nan', '-', '₹ -'):
        return 0
    try:
        return float(s)
    except ValueError:
        import re
        match = re.search(r'[-+]?\d*\.?\d+', s)
        if match:
            return float(match.group())
        return 0

def get_latest_file(directory, patterns=['*.xlsx', '*.csv']):
    files = []
    for pat in patterns:
        files.extend(glob.glob(os.path.join(directory, pat)))
    # Exclude temp files
    files = [f for f in files if not os.path.basename(f).startswith('~')]
    if not files:
        return None
    return max(files, key=os.path.getmtime)

# ──────────────────────────────────────────────────────────
#  PRIORITY LOADER
# ──────────────────────────────────────────────────────────
def load_store_priorities():
    priority_dict = {}
    
    # 1. Try loading from ag validation priority/Store wise May sales qty.xlsx
    file1 = r"d:\INCREFF ORDER PUNCH\ag validation priority\Store wise May sales qty.xlsx"
    if os.path.exists(file1):
        try:
            df = pd.read_excel(file1)
            for _, row in df.iterrows():
                code = clean_store_code(row.get('Store Code'))
                name = str(row.get('EBO NAME', '')).strip().lower()
                pri = row.get('priority')
                try:
                    pri_val = int(pri)
                except (ValueError, TypeError):
                    pri_val = 999999
                
                if code:
                    priority_dict[code] = pri_val
                if name:
                    priority_dict[name] = pri_val
            print(f"[INFO] Loaded priorities from {os.path.basename(file1)}")
        except Exception as e:
            print(f"[WARN] Error loading {file1}: {e}")
            
    # 2. Try loading from priority list/Priority list.xlsx
    file2 = r"d:\INCREFF ORDER PUNCH\priority list\Priority list.xlsx"
    if os.path.exists(file2):
        try:
            df = pd.read_excel(file2)
            p_store_col = next((c for c in df.columns if str(c).strip().lower() in ["store name", "store_name", "storename"]), None)
            p_num_col = next((c for c in df.columns if str(c).strip().lower() in ["priority number", "priority", "priority list", "priority_number", "priority_list"]), None)
            if p_store_col and p_num_col:
                for _, row in df.iterrows():
                    name = str(row[p_store_col]).strip().lower()
                    pri = row[p_num_col]
                    try:
                        pri_val = int(pri)
                    except (ValueError, TypeError):
                        pri_val = 999999
                    
                    if name and name not in priority_dict:
                        priority_dict[name] = pri_val
                print(f"[INFO] Loaded priorities from {os.path.basename(file2)}")
        except Exception as e:
            print(f"[WARN] Error loading {file2}: {e}")
            
    # 3. Add PRIORITY_TO_EBO manual mappings from check_allocation if available
    try:
        from check_allocation import PRIORITY_TO_EBO
        for k, v in PRIORITY_TO_EBO.items():
            k_low = str(k).strip().lower()
            v_low = str(v).strip().lower() if v else None
            if k_low in priority_dict:
                p_val = priority_dict[k_low]
                if v_low and v_low not in priority_dict:
                    priority_dict[v_low] = p_val
    except Exception as e:
        pass
        
    return priority_dict

def get_label_priority(label, priority_dict):
    if " - " in label:
        parts = label.split(" - ")
        code = parts[0].strip()
        name = parts[1].strip().lower()
        if code in priority_dict:
            return priority_dict[code]
        if name in priority_dict:
            return priority_dict[name]
    else:
        val = label.strip().lower()
        if val in priority_dict:
            return priority_dict[val]
        code_clean = clean_store_code(label)
        if code_clean in priority_dict:
            return priority_dict[code_clean]
            
    # Substring search fallback
    lbl_clean = label.strip().lower()
    for k, pri in priority_dict.items():
        if k in lbl_clean or lbl_clean in k:
            return pri
            
    return 999999

# ──────────────────────────────────────────────────────────
#  EXCEL STYLING HELPERS
# ──────────────────────────────────────────────────────────
def style_sheet(ws, has_index=False, index_name='STYLE'):
    fill_header = PatternFill("solid", fgColor="1F3864")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    fill_zebra = PatternFill("solid", fgColor="F2F5F8")
    font_data = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    if has_index:
        if ws.cell(row=1, column=1).value is None or ws.cell(row=1, column=1).value == "":
            ws.cell(row=1, column=1).value = index_name

    # Header row formatting
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = thin_border
        
    # Data row formatting
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        use_zebra = (row_idx % 2 == 0)
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if use_zebra:
                cell.fill = fill_zebra
            cell.font = font_data
            cell.border = thin_border
            
            # Alignments
            if col_idx <= (2 if not has_index else 1):
                cell.alignment = align_left
            else:
                cell.alignment = align_center

    # Freeze panes
    if ws.title == 'Summary':
        ws.freeze_panes = "C2"
    elif ws.title.startswith('Store Sales'):
        ws.freeze_panes = "B2"
    else:
        ws.freeze_panes = "A2"

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                val_len = max(len(s) for s in val_str.split('\n'))
            else:
                val_len = len(val_str)
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# ──────────────────────────────────────────────────────────
#  MAIN PROCESSING
# ──────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate Top 100 Sales Styles Report.")
    parser.add_argument('input_file', nargs='?', default=None, help="Path to sales Excel or CSV file")
    parser.add_argument('--days1', type=int, default=10, help="First date range window size in days (default: 10)")
    parser.add_argument('--days2', type=int, default=30, help="Second date range window size in days (default: 30)")
    parser.add_argument('--top', type=int, default=100, help="Number of top styles to retrieve (default: 100)")
    
    args = parser.parse_args()
    
    # Determine input path
    input_path = args.input_file
    if not input_path:
        input_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\input of top 100 style"
        print(f"[INFO] No input file specified. Looking in {input_dir}...")
        input_path = get_latest_file(input_dir)
        if not input_path:
            print("[INFO] Not found in input of top 100 style. Looking in D:\\downloads...")
            input_path = get_latest_file(r"D:\downloads")
            if not input_path:
                print("[INFO] Not found in D:\\downloads. Looking in new style performance...")
                input_path = get_latest_file(r"D:\INCREFF ORDER PUNCH\new style performance", patterns=['Style wise sales amount*.csv', 'Style wise sales amount*.xlsx'])
                if not input_path:
                    print("[INFO] Looking in current directory...")
                    input_path = get_latest_file(".")
            
    if not input_path:
        print("[ERROR] Could not find any sales Excel/CSV file to run. Please specify the path as an argument.")
        sys.exit(1)
        
    print(f"Reading Sales Data from: {input_path}")
    
    # Load data
    if input_path.lower().endswith('.csv'):
        df = pd.read_csv(input_path)
    else:
        df = pd.read_excel(input_path)
        
    print(f"Loaded {len(df)} rows. Mapping columns...")
    cols = df.columns.tolist()
    
    store_code_candidates = ['store code', 'site code', 'store_code', 'site_code', 'sitecode', 'storecode']
    ebo_name_candidates = ['ebo name', 'store name', 'owner site', 'owner_site', 'ebo_name', 'store_name']
    bill_date_candidates = ['bill date', 'date', 'bill_date', 'transaction date', 'transaction_date', 'date of bill']
    style_candidates = ['style', 'style name', 'style_code', 'icode']
    qty_candidates = ['bill quantity', 'quantity', 'qty', 'bill_quantity', 'sold qty', 'sold quantity', 'bill qty']
    
    store_code_col = find_column(cols, store_code_candidates)
    ebo_name_col = find_column(cols, ebo_name_candidates)
    bill_date_col = find_column(cols, bill_date_candidates)
    style_col = find_column(cols, style_candidates)
    qty_col = find_column(cols, qty_candidates)

    # Check for aggregated style-level sales columns (support amount and qty)
    sales_10_col = find_column(cols, ['last 10 days sale amt', 'last 10 days sales amt', 'last 10 days sales', '10 days sales', 'sales 10', '10 days', '10d', ' last 10 days sales ', 'sales_10'])
    sales_30_col = find_column(cols, ['last 30 days sales amt', 'last 30 days sales', '30 days sales', 'sales 30', '30 days', '30d', ' last 30 days sales ', 'sales_30', 'sales_30_days'])
    
    qty_10_col = None
    qty_30_col = None
    for c in cols:
        c_lower = str(c).strip().lower()
        if '10' in c_lower and ('qty' in c_lower or 'quantity' in c_lower):
            qty_10_col = c
        if '30' in c_lower and ('qty' in c_lower or 'quantity' in c_lower):
            qty_30_col = c
            
    # Fallback by names and position
    if not qty_30_col:
        for c in cols:
            if str(c).strip() == 'Item Qty':
                qty_30_col = c
                break
    if not qty_30_col and len(cols) > 2:
        qty_30_col = cols[2]
        
    if not qty_10_col:
        for c in cols:
            if str(c).strip() == 'Item qty':
                qty_10_col = c
                break
    if not qty_10_col and len(cols) > 4:
        qty_10_col = cols[4]
        
    is_aggregated = False
    if style_col and (sales_10_col or sales_30_col or qty_10_col or qty_30_col) and not bill_date_col:
        is_aggregated = True
        
    if is_aggregated:
        print("[INFO] Aggregated style-level sales format detected.")
        print(f"Mapped columns: 30d Qty={qty_30_col}, 10d Qty={qty_10_col}, 30d Sales Amt={sales_30_col}, 10d Sales Amt={sales_10_col}")
        df[style_col] = df[style_col].astype(str).str.strip().str.upper()
        if sales_10_col:
            df[sales_10_col] = df[sales_10_col].apply(clean_currency).astype(int)
        if sales_30_col:
            df[sales_30_col] = df[sales_30_col].apply(clean_currency).astype(int)
        if qty_10_col:
            df[qty_10_col] = pd.to_numeric(df[qty_10_col], errors='coerce').fillna(0).astype(int)
        if qty_30_col:
            df[qty_30_col] = pd.to_numeric(df[qty_30_col], errors='coerce').fillna(0).astype(int)

        # Rank by qty (preferred) or sales amount as fallback
        sort_col = qty_30_col if qty_30_col else (qty_10_col if qty_10_col else (sales_30_col if sales_30_col else sales_10_col))
        df_sorted = df.sort_values(by=sort_col, ascending=False)
        top_styles = df_sorted.head(args.top)[style_col].tolist()
        
        summary_10d_amt = df_sorted.set_index(style_col).reindex(top_styles)[sales_10_col].rename("Last 10 Days Sales") if sales_10_col else pd.Series(0, index=top_styles, name="Last 10 Days Sales")
        summary_10d_qty = df_sorted.set_index(style_col).reindex(top_styles)[qty_10_col].rename("Last 10 Days Qty") if qty_10_col else pd.Series(0, index=top_styles, name="Last 10 Days Qty")
        summary_30d_amt = df_sorted.set_index(style_col).reindex(top_styles)[sales_30_col].rename("Last 30 Days Sales") if sales_30_col else pd.Series(0, index=top_styles, name="Last 30 Days Sales")
        summary_30d_qty = df_sorted.set_index(style_col).reindex(top_styles)[qty_30_col].rename("Last 30 Days Qty") if qty_30_col else pd.Series(0, index=top_styles, name="Last 30 Days Qty")
        
        summary_total = summary_30d_qty.rename("Total Qty Sold")
        
        df_10d_pivot = pd.DataFrame(0, index=top_styles, columns=['Grand Total'])
        df_10d_pivot.index.name = 'Style'
        df_100d_pivot = pd.DataFrame(0, index=top_styles, columns=['Grand Total'])
        df_100d_pivot.index.name = 'Style'
        df_flat = pd.DataFrame(columns=['Style', 'Store Code', 'EBO Name', 'Total Qty Sold', 'Last 10 Days Sales', 'Last 10 Days Qty', 'Last 30 Days Sales', 'Last 30 Days Qty'])
        store_priorities = {}
        
    else:
        # Standard Transaction-level processing
        missing = []
        if not store_code_col: missing.append("Store Code")
        if not bill_date_col: missing.append("Bill Date")
        if not style_col: missing.append("Style")
        if not qty_col: missing.append("Bill Quantity")
        
        if missing:
            print(f"[ERROR] Missing required columns in input sheet: {', '.join(missing)}")
            print(f"Available columns are: {cols}")
            sys.exit(1)
            
        print(f"Mapped Columns:")
        print(f"  - Store Code: {store_code_col}")
        if ebo_name_col:
             print(f"  - EBO Name: {ebo_name_col}")
        print(f"  - Bill Date: {bill_date_col}")
        print(f"  - Style: {style_col}")
        print(f"  - Bill Qty: {qty_col}")
        
        df[style_col] = df[style_col].astype(str).str.strip().str.upper()
        df[qty_col] = pd.to_numeric(df[qty_col], errors='coerce').fillna(0)
        df[bill_date_col] = pd.to_datetime(df[bill_date_col], errors='coerce')
        df = df.dropna(subset=[bill_date_col]).copy()
        
        if len(df) == 0:
            print("[ERROR] No valid sales records with dates found. Cannot continue.")
            sys.exit(1)
            
        max_date = df[bill_date_col].max()
        min_date = df[bill_date_col].min()
        start_date_10d = max_date - pd.Timedelta(days=args.days1 - 1)
        start_date_100d = max_date - pd.Timedelta(days=args.days2 - 1)
        
        print(f"\nTransaction Dates in File: {min_date.strftime('%Y-%m-%d')} to {max_date.strftime('%Y-%m-%d')}")
        store_priorities = load_store_priorities()
        
        def row_store_label(row):
            code = clean_store_code(row[store_code_col])
            name = str(row[ebo_name_col]).strip() if ebo_name_col and pd.notna(row[ebo_name_col]) else ""
            if code and name:
                return f"{code} - {name}"
            elif code:
                return code
            elif name:
                return name
            return "UNKNOWN"
            
        df['_store_label'] = df.apply(row_store_label, axis=1)
        
        style_totals = df.groupby(style_col)[qty_col].sum().sort_values(ascending=False)
        top_styles = style_totals.head(args.top).index.tolist()
        df_top = df[df[style_col].isin(top_styles)].copy()
        
        summary_total = df_top.groupby(style_col)[qty_col].sum().rename("Total Qty Sold")
        summary_10d_amt = pd.Series(0, index=top_styles, name="Last 10 Days Sales")
        summary_10d_qty = df_top[df_top[bill_date_col] >= start_date_10d].groupby(style_col)[qty_col].sum().rename("Last 10 Days Qty")
        summary_30d_amt = pd.Series(0, index=top_styles, name="Last 30 Days Sales")
        summary_30d_qty = df_top[df_top[bill_date_col] >= start_date_100d].groupby(style_col)[qty_col].sum().rename("Last 30 Days Qty")
        
        print("Generating Store-wise Pivot sheets...")
        unique_stores = df_top['_store_label'].unique().tolist()
        sorted_stores = sorted(unique_stores, key=lambda s: (get_label_priority(s, store_priorities), s.lower()))
        
        df_10d_filtered = df_top[df_top[bill_date_col] >= start_date_10d]
        if not df_10d_filtered.empty:
            df_10d_pivot = df_10d_filtered.pivot_table(index=style_col, columns='_store_label', values=qty_col, aggfunc='sum', fill_value=0)
            df_10d_pivot = df_10d_pivot.reindex(index=top_styles, columns=sorted_stores, fill_value=0).astype(int)
            df_10d_pivot['Grand Total'] = df_10d_pivot.sum(axis=1)
        else:
            df_10d_pivot = pd.DataFrame(0, index=top_styles, columns=sorted_stores + ['Grand Total'])
            df_10d_pivot.index.name = style_col
            
        df_100d_filtered = df_top[df_top[bill_date_col] >= start_date_100d]
        if not df_100d_filtered.empty:
            df_100d_pivot = df_100d_filtered.pivot_table(index=style_col, columns='_store_label', values=qty_col, aggfunc='sum', fill_value=0)
            df_100d_pivot = df_100d_pivot.reindex(index=top_styles, columns=sorted_stores, fill_value=0).astype(int)
            df_100d_pivot['Grand Total'] = df_100d_pivot.sum(axis=1)
        else:
            df_100d_pivot = pd.DataFrame(0, index=top_styles, columns=sorted_stores + ['Grand Total'])
            df_100d_pivot.index.name = style_col

        print("Generating Flat details sheet...")
        flat_total = df_top.groupby([style_col, store_code_col, '_store_label'])[qty_col].sum().rename("Total Qty Sold")
        flat_10d = df_top[df_top[bill_date_col] >= start_date_10d].groupby([style_col, store_code_col, '_store_label'])[qty_col].sum().rename(f"Last {args.days1} Days Qty")
        flat_100d = df_top[df_top[bill_date_col] >= start_date_100d].groupby([style_col, store_code_col, '_store_label'])[qty_col].sum().rename(f"Last {args.days2} Days Qty")
        
        df_flat = pd.DataFrame(index=df_top.groupby([style_col, store_code_col, '_store_label']).groups.keys())
        df_flat.index.names = ['Style', 'Store Code', 'Store Label']
        df_flat = df_flat.join(flat_total, how='left')
        df_flat = df_flat.join(flat_10d, how='left').fillna(0)
        df_flat = df_flat.join(flat_100d, how='left').fillna(0)
        df_flat = df_flat.reset_index()
        
        if ebo_name_col:
            def extract_ebo_name(label):
                if " - " in label:
                    return label.split(" - ", 1)[1]
                return label
            df_flat['EBO Name'] = df_flat['Store Label'].apply(extract_ebo_name)
        else:
            df_flat['EBO Name'] = ""
            
        df_flat = df_flat[['Style', 'Store Code', 'EBO Name', 'Total Qty Sold', f'Last {args.days1} Days Qty', f'Last {args.days2} Days Qty']]
        
        for c in ['Total Qty Sold', f'Last {args.days1} Days Qty', f'Last {args.days2} Days Qty']:
            df_flat[c] = df_flat[c].astype(int)
            
        style_rank_dict = {s: r for r, s in enumerate(top_styles)}
        df_flat['_style_rank'] = df_flat['Style'].map(style_rank_dict)
        
        def get_row_store_pri(row):
            code = clean_store_code(row['Store Code'])
            ebo = str(row['EBO Name']).strip().lower()
            if code in store_priorities: return store_priorities[code]
            if ebo in store_priorities: return store_priorities[ebo]
            return 999999
            
        df_flat['_store_pri'] = df_flat.apply(get_row_store_pri, axis=1)
        df_flat = df_flat.sort_values(by=['_style_rank', '_store_pri']).drop(columns=['_style_rank', '_store_pri']).reset_index(drop=True)

    # ── Load EBO current stock, transit, allocation and pools for enrichment ──
    print("Loading enrichment data (EBO stock, transit, allocation, OMS pools)...")
    style_stores_curr = {}
    style_stores_tran = {}
    style_stores_alloc = {}
    
    # 1. Current Stock
    try:
        ebo_curr_path = get_latest_file(r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock")
        if ebo_curr_path:
            df_curr = pd.read_excel(ebo_curr_path, engine='calamine')
            df_curr.columns = [str(c).strip() for c in df_curr.columns]
            sc_col_curr = next((c for c in df_curr.columns if str(c).strip().lower() in ['site_code', 'site code', 'sitecode', 'store code', 'store_code', 'storecode']), None)
            qty_col_curr = next((c for c in df_curr.columns if str(c).strip().lower() in ['quantity', 'qty', 'stock quantity', 'stock_quantity']), None)
            style_col_curr = next((c for c in df_curr.columns if str(c).strip().lower() == 'style'), None)
            if sc_col_curr and qty_col_curr and style_col_curr:
                df_curr[sc_col_curr] = df_curr[sc_col_curr].apply(clean_store_code)
                df_curr[qty_col_curr] = pd.to_numeric(df_curr[qty_col_curr], errors='coerce').fillna(0)
                df_curr[style_col_curr] = df_curr[style_col_curr].astype(str).str.strip().str.upper()
                df_curr_active = df_curr[df_curr[qty_col_curr] > 0]
                for style_code, group in df_curr_active.groupby(style_col_curr):
                    style_stores_curr[style_code] = set(group[sc_col_curr].dropna().unique())
    except Exception as e:
        print(f"[WARN] Error loading Current Stock: {e}")

    # 2. Transit Stock
    try:
        ebo_tran_path = get_latest_file(r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit")
        if ebo_tran_path:
            df_tran = pd.read_excel(ebo_tran_path, engine='calamine')
            df_tran.columns = [str(c).strip() for c in df_tran.columns]
            sc_col_tran = next((c for c in df_tran.columns if str(c).strip().lower() in ['store code', 'store_code', 'storecode', 'site code', 'site_code', 'sitecode']), None)
            qty_col_tran = next((c for c in df_tran.columns if str(c).strip().lower() in ['transit qty', 'transit_qty', 'quantity', 'qty']), None)
            style_col_tran = next((c for c in df_tran.columns if str(c).strip().lower() == 'style'), None)
            if sc_col_tran and qty_col_tran and style_col_tran:
                df_tran[sc_col_tran] = df_tran[sc_col_tran].apply(clean_store_code)
                df_tran[qty_col_tran] = pd.to_numeric(df_tran[qty_col_tran], errors='coerce').fillna(0)
                df_tran[style_col_tran] = df_tran[style_col_tran].astype(str).str.strip().str.upper()
                df_tran_active = df_tran[df_tran[qty_col_tran] > 0]
                for style_code, group in df_tran_active.groupby(style_col_tran):
                    style_stores_tran[style_code] = set(group[sc_col_tran].dropna().unique())
    except Exception as e:
        print(f"[WARN] Error loading Transit Stock: {e}")

    # 3. Allocation Stock
    try:
        alloc_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\ALLOCATION"
        alloc_files = glob.glob(os.path.join(alloc_dir, "*.xlsx"))
        alloc_files = [f for f in alloc_files if not os.path.basename(f).startswith('~')]
        if alloc_files:
            latest_alloc = max(alloc_files, key=os.path.getmtime)
            df_alloc = pd.read_excel(latest_alloc, engine='calamine')
            df_alloc.columns = [str(c).strip() for c in df_alloc.columns]
            sc_col_alloc = next((c for c in df_alloc.columns if 'store' in c.lower() and 'code' in c.lower()), None)
            style_col_alloc = next((c for c in df_alloc.columns if c.strip().lower() == 'style'), None)
            qty_col_alloc = next((c for c in df_alloc.columns if 'allocated' in c.lower() and 'qty' in c.lower()), None)
            if sc_col_alloc and style_col_alloc and qty_col_alloc:
                df_alloc[sc_col_alloc] = df_alloc[sc_col_alloc].apply(clean_store_code)
                df_alloc[style_col_alloc] = df_alloc[style_col_alloc].astype(str).str.strip().str.upper()
                df_alloc[qty_col_alloc] = pd.to_numeric(df_alloc[qty_col_alloc], errors='coerce').fillna(0)
                df_alloc_active = df_alloc[df_alloc[qty_col_alloc] > 0]
                for style_code, group in df_alloc_active.groupby(style_col_alloc):
                    style_stores_alloc[style_code] = set(group[sc_col_alloc].dropna().unique())
    except Exception as e:
        print(f"[WARN] Error loading Allocation Stock: {e}")

    # 4. Valid Options average
    style_valid_opts_map = {}
    try:
        path_validation = r"D:\INCREFF ORDER PUNCH\VALID STYLE OUTPUT\AG_Validation_Output_v2_LATEST.xlsx"
        if os.path.exists(path_validation):
            df_style_val = pd.read_excel(path_validation, sheet_name="STYLE WISE", engine="calamine")
            df_style_val.columns = [str(c).strip() for c in df_style_val.columns]
            for col in df_style_val.columns:
                if col.endswith("_Allocation"):
                    style_name = col[:-11].upper()
                    mean_val = df_style_val[col].mean()
                    style_valid_opts_map[style_name] = round(float(mean_val), 2)
    except Exception as e:
        print(f"[WARN] Error loading Valid Options: {e}")

    # 5. OMS Reservation Pools and GT Pool
    d2c_map, cred_map, ebo_map, common_pool_map, gt_pool_map = {}, {}, {}, {}, {}
    try:
        ns_perf_dir = r"D:\INCREFF ORDER PUNCH\new style performance"
        csv_files = glob.glob(os.path.join(ns_perf_dir, "Inventory Available for Sales - OMS*.csv"))
        latest_csv = max(csv_files, key=os.path.getmtime) if csv_files else None
        if latest_csv and os.path.exists(latest_csv):
            df_inv = pd.read_csv(latest_csv, usecols=['Style', 'Reservation Pool', 'Total Available Quantity'])
            df_inv['Style'] = df_inv['Style'].astype(str).str.strip().str.upper()
            df_inv['Reservation Pool'] = df_inv['Reservation Pool'].astype(str).str.strip()
            df_inv['Total Available Quantity'] = pd.to_numeric(df_inv['Total Available Quantity'], errors='coerce').fillna(0).astype(int)
            grouped_inv = df_inv.groupby(['Style', 'Reservation Pool'])['Total Available Quantity'].sum().reset_index()
            for _, row in grouped_inv.iterrows():
                style_p = row['Style']
                pool = row['Reservation Pool']
                qty = int(row['Total Available Quantity'])
                if pool == 'D2C-Marketplaces':
                    d2c_map[style_p] = d2c_map.get(style_p, 0) + qty
                elif pool == 'CRED':
                    cred_map[style_p] = cred_map.get(style_p, 0) + qty
                elif pool == 'EBO':
                    ebo_map[style_p] = ebo_map.get(style_p, 0) + qty
                elif pool == 'Common_Pool-TECHNO SPORTSWEAR PRIVATE LIMITED-wms_tiruppur':
                    common_pool_map[style_p] = common_pool_map.get(style_p, 0) + qty
    except Exception as e:
        print(f"[WARN] Error loading OMS pools: {e}")

    # Load GT inventory pool
    try:
        gt_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\gt inventory pool"
        if os.path.exists(gt_dir):
            gt_files = glob.glob(os.path.join(gt_dir, "*.xlsx"))
            gt_files = [f for f in gt_files if not os.path.basename(f).startswith('~')]
            if gt_files:
                latest_gt = max(gt_files, key=os.path.getmtime)
                df_gt = pd.read_excel(latest_gt)
                df_gt.columns = [str(c).strip() for c in df_gt.columns]
                style_col_gt = next((c for c in df_gt.columns if c.strip().lower() == 'style'), None)
                qty_col_gt   = next((c for c in df_gt.columns if c.strip().lower() == 'grand total'), None)
                if style_col_gt and qty_col_gt:
                    df_gt[style_col_gt] = df_gt[style_col_gt].astype(str).str.strip().str.upper()
                    df_gt[qty_col_gt] = pd.to_numeric(df_gt[qty_col_gt], errors='coerce').fillna(0).astype(int)
                    for _, row in df_gt.iterrows():
                        style_p = row[style_col_gt]
                        qty = int(row[qty_col_gt])
                        gt_pool_map[style_p] = gt_pool_map.get(style_p, 0) + qty
    except Exception as e:
        print(f"[WARN] Error loading GT pool: {e}")

    # Build metric lists
    valid_options_list = []
    stores_live_list = []
    d2c_list, cred_list, ebo_list, common_list, gt_list = [], [], [], [], []
    drr_list = []
    doh_list = []
    
    for style in top_styles:
        s_upper = style.upper()
        
        # Valid options
        valid_options_list.append(style_valid_opts_map.get(s_upper, 0.0))
        
        # Live stores
        live_set = set()
        if s_upper in style_stores_curr: live_set.update(style_stores_curr[s_upper])
        if s_upper in style_stores_tran: live_set.update(style_stores_tran[s_upper])
        if s_upper in style_stores_alloc: live_set.update(style_stores_alloc[s_upper])
        stores_live_list.append(len(live_set))
        
        # Pools
        d2c_val = d2c_map.get(s_upper, 0)
        cred_val = cred_map.get(s_upper, 0)
        ebo_val = ebo_map.get(s_upper, 0)
        common_val = common_pool_map.get(s_upper, 0)
        gt_val = gt_pool_map.get(s_upper, 0)
        
        d2c_list.append(d2c_val)
        cred_list.append(cred_val)
        ebo_list.append(ebo_val)
        common_list.append(common_val)
        gt_list.append(gt_val)
        
        # DRR is 30 days qty / 30.0
        qty_30_val = summary_30d_qty.get(style, 0) if hasattr(summary_30d_qty, 'get') else int(summary_30d_qty.reindex([style]).fillna(0).iloc[0])
        drr = round(float(qty_30_val) / 30.0, 2)
        drr_list.append(drr)
        
        total_stock = d2c_val + cred_val + ebo_val + common_val + gt_val
        if drr > 0:
            doh = round(float(total_stock) / drr, 1)
        else:
            doh = 999.0 if total_stock > 0 else 0.0
        doh_list.append(doh)

    # Combine Summary Sheet
    df_summary = pd.DataFrame(index=top_styles)
    df_summary.index.name = "Style"
    df_summary = df_summary.join(summary_10d_amt, how='left').fillna(0)
    df_summary = df_summary.join(summary_10d_qty, how='left').fillna(0)
    df_summary = df_summary.join(summary_30d_amt, how='left').fillna(0)
    df_summary = df_summary.join(summary_30d_qty, how='left').fillna(0)
    
    # Cast quantity/amount columns to integer
    for col in df_summary.columns:
        df_summary[col] = df_summary[col].astype(int)
        
    df_summary['DRR'] = drr_list
    df_summary['DOH'] = doh_list
    df_summary['Valid Options'] = valid_options_list
    df_summary['No. of Stores Live'] = stores_live_list
    df_summary['D2C Pool'] = d2c_list
    df_summary['CRED Pool'] = cred_list
    df_summary['EBO Pool'] = ebo_list
    df_summary['Common Pool'] = common_list
    df_summary['GT Pool'] = gt_list
    df_summary = df_summary.join(summary_total, how='left').fillna(0)
    for col in ('Total Sales', 'Total Qty Sold'):
        if col in df_summary.columns:
            df_summary[col] = df_summary[col].astype(int)
        
    df_summary = df_summary.reset_index()
    df_summary.insert(0, "Rank", range(1, len(df_summary) + 1))

    # ──────────────────────────────────────────────────────────
    #  WRITE WORKBOOK WITH OPENPYXL STYLING
    # ──────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\output of top 100"
    os.makedirs(output_dir, exist_ok=True)
    out_file = os.path.join(output_dir, f"Top_{args.top}_Sales_Report_{timestamp}.xlsx")
    
    print(f"\nWriting styled workbook to {out_file}...")
    save_workbook_with_styles(out_file, df_summary, df_10d_pivot, df_100d_pivot, df_flat)
    
    # Save LATEST copy
    try:
        latest_file = os.path.join(output_dir, f"Top_{args.top}_Sales_Report_LATEST.xlsx")
        import shutil
        shutil.copy(out_file, latest_file)
        print(f"Copied LATEST report to: {latest_file}")
    except Exception as e:
        print(f"[WARN] Could not copy to LATEST path: {e}")
        
    print("\nProcessing complete! Done.")

def save_workbook_with_styles(out_file, df_summary, df_10d_pivot, df_100d_pivot, df_flat):
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        df_10d_pivot.to_excel(writer, sheet_name='Store Sales - Last 10 Days', index=True)
        df_100d_pivot.to_excel(writer, sheet_name='Store Sales - Last 30 Days', index=True)
        df_flat.to_excel(writer, sheet_name='Flat Store Style Details', index=False)
        
        # Style Sheets
        style_sheet(writer.sheets['Summary'], has_index=False)
        style_sheet(writer.sheets['Store Sales - Last 10 Days'], has_index=True, index_name='Style')
        style_sheet(writer.sheets['Store Sales - Last 30 Days'], has_index=True, index_name='Style')
        style_sheet(writer.sheets['Flat Store Style Details'], has_index=False)

if __name__ == "__main__":
    main()
