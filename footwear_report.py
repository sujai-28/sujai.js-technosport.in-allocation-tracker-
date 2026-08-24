"""
footwear_report.py
-------------------
Generates a Footwear (Shoes) inventory report directly from EBO stock and transit files,
without filtering by warehouse availability.

Output Excel has two sheets:
  1. Footwear Report – Style | SKU | Size | per-store stock and transit cols (sorted by priority)
  2. Raw Stock Data  – Raw footwear stock records from current stock
"""

import os
import sys
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
STOCK_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock"
TRANSIT_DIR   = r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit"
OUTPUT_DIR    = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE"

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def get_latest(directory, pattern="*.xlsx"):
    files = glob.glob(os.path.join(directory, pattern))
    files = [f for f in files if not os.path.basename(f).startswith("~")]
    return max(files, key=os.path.getmtime) if files else None

def apply_header_style(ws, row_num, fill_hex="7F6000"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

# ─────────────────────────────────────────────
# STEP 1 — Load current store stock and filter footwear
# ─────────────────────────────────────────────
stock_file = get_latest(STOCK_DIR)
if not stock_file:
    print("ERROR: No current stock file found.")
    sys.exit(1)
print(f"Loading stock: {os.path.basename(stock_file)}")

df_stock = pd.read_excel(stock_file)
df_stock.columns = [c.strip() for c in df_stock.columns]

# Detect columns
sc_col   = next((c for c in df_stock.columns if c.strip().lower() in ["store code", "store_code", "site_code", "owner site code"]), None)
sn_col   = next((c for c in df_stock.columns if c.strip().lower() in ["owner site", "store name", "ebo name"]), None)
sku_s    = next((c for c in df_stock.columns if c.strip().lower() == "sku"), None) or \
           next((c for c in df_stock.columns if c.strip().lower() in ["barcode", "client sku id / ean"]), None)
qty_s    = next((c for c in df_stock.columns if c.strip().lower() in ["stock quantity", "quantity", "qty"]), None)
style_s  = next((c for c in df_stock.columns if c.strip().lower() == "style"), "STYLE")
size_s   = next((c for c in df_stock.columns if c.strip().lower() == "size"), "SIZE")
dept_s   = next((c for c in df_stock.columns if c.strip().lower() == "department"), "DEPARTMENT")

# Filter for Shoes department or style starts with 'S'
is_footwear_stock = (df_stock[dept_s].astype(str).str.strip().str.upper() == "MENS-SHOES") | \
                    (df_stock[style_s].astype(str).str.strip().str.upper().str.startswith("S"))

df_stock_foot = df_stock[is_footwear_stock].copy()
print(f"Found {len(df_stock_foot)} footwear stock rows.")

# ─────────────────────────────────────────────
# STEP 2 — Load in-transit data and filter footwear
# ─────────────────────────────────────────────
transit_file = get_latest(TRANSIT_DIR)
df_tran_foot = pd.DataFrame()

if not transit_file:
    print("WARNING: No transit file found. Skipping transit data.")
else:
    print(f"Loading transit: {os.path.basename(transit_file)}")
    df_tran = pd.read_excel(transit_file)
    df_tran.columns = [c.strip() for c in df_tran.columns]

    sku_t   = next((c for c in df_tran.columns if c.strip().lower() == "sku"), None) or \
              next((c for c in df_tran.columns if c.strip().lower() in ["barcode", "client sku id / ean"]), None)
    sn_t    = next((c for c in df_tran.columns if c.strip().lower() in ["ebo name", "store name", "owner site"]), None)
    sc_t    = next((c for c in df_tran.columns if c.strip().lower() in ["store code", "store_code"]), None)
    qty_t   = next((c for c in df_tran.columns if c.strip().lower() in ["transit qty", "quantity", "qty"]), None)
    style_t = next((c for c in df_tran.columns if c.strip().lower() == "style"), "STYLE")
    size_t  = next((c for c in df_tran.columns if c.strip().lower() == "size"), "SIZE")
    dept_t  = next((c for c in df_tran.columns if c.strip().lower() == "department"), "DEPARTMENT")

    is_footwear_tran = (df_tran[dept_t].astype(str).str.strip().str.upper() == "MENS-SHOES") | \
                       (df_tran[style_t].astype(str).str.strip().str.upper().str.startswith("S"))

    df_tran_foot = df_tran[is_footwear_tran].copy()
    print(f"Found {len(df_tran_foot)} footwear transit rows.")

# Exit if no footwear found anywhere
if df_stock_foot.empty and df_tran_foot.empty:
    print("No footwear records found in stock or transit files.")
    sys.exit(0)

# ─────────────────────────────────────────────
# STEP 3 — Build unique list of all footwear SKUs
# ─────────────────────────────────────────────
skus_stock = df_stock_foot[[style_s, sku_s, size_s]].rename(columns={style_s: "Style", sku_s: "SKU", size_s: "Size"})
skus_tran = pd.DataFrame()
if not df_tran_foot.empty:
    skus_tran = df_tran_foot[[style_t, sku_t, size_t]].rename(columns={style_t: "Style", sku_t: "SKU", size_t: "Size"})

df_sku = pd.concat([skus_stock, skus_tran], ignore_index=True).drop_duplicates()
df_sku["SKU"] = df_sku["SKU"].astype(str).str.strip()
df_sku["Style"] = df_sku["Style"].astype(str).str.strip()
df_sku["Size"] = df_sku["Size"].astype(str).str.strip()

print(f"Unique footwear SKUs in EBO files: {len(df_sku)}")

# ─────────────────────────────────────────────
# STEP 4 — Pivot Stock and Transit
# ─────────────────────────────────────────────
stock_pivot = pd.DataFrame()
if not df_stock_foot.empty:
    if sn_col:
        df_stock_foot["_store_label"] = df_stock_foot[sn_col].astype(str).str.strip()
    else:
        df_stock_foot["_store_label"] = df_stock_foot[sc_col].astype(str).str.strip()
    
    df_stock_foot[qty_s] = pd.to_numeric(df_stock_foot[qty_s], errors="coerce").fillna(0)
    df_stock_foot["_sku_up"] = df_stock_foot[sku_s].astype(str).str.strip().str.upper()

    stock_pivot = (
        df_stock_foot.groupby(["_sku_up", "_store_label"])[qty_s]
        .sum()
        .unstack(fill_value=0)
    )
    stock_pivot.columns = [f"Stock_{c}" for c in stock_pivot.columns]

transit_pivot = pd.DataFrame()
if not df_tran_foot.empty:
    label_col = sn_t if sn_t else sc_t
    df_tran_foot["_store_label"] = df_tran_foot[label_col].astype(str).str.strip()
    
    df_tran_foot[qty_t] = pd.to_numeric(df_tran_foot[qty_t], errors="coerce").fillna(0)
    df_tran_foot["_sku_up"] = df_tran_foot[sku_t].astype(str).str.strip().str.upper()

    transit_pivot = (
        df_tran_foot.groupby(["_sku_up", "_store_label"])[qty_t]
        .sum()
        .unstack(fill_value=0)
    )
    transit_pivot.columns = [f"Transit_{c}" for c in transit_pivot.columns]

# Merge pivots with our SKU master list
df_sku["_sku_up"] = df_sku["SKU"].str.upper()
df_merged = df_sku.set_index("_sku_up")

if not stock_pivot.empty:
    df_merged = df_merged.join(stock_pivot, how="left")
if not transit_pivot.empty:
    df_merged = df_merged.join(transit_pivot, how="left")

df_merged = df_merged.fillna(0)
df_merged = df_merged.reset_index(drop=True)

# Sort rows by Style, SKU, Size
df_merged = df_merged.sort_values(["Style", "SKU", "Size"]).reset_index(drop=True)

# ─────────────────────────────────────────────
# STEP 5 — Sort store columns by priority list
# ─────────────────────────────────────────────
PRIORITY_FILE = r"priority list\Priority list.xlsx"
priority_map = {}
if os.path.exists(PRIORITY_FILE):
    try:
        from check_allocation import PRIORITY_TO_EBO
        df_pri = pd.read_excel(PRIORITY_FILE)
        p_store_col = next((c for c in df_pri.columns if str(c).strip().lower() in ["store name", "store_name", "storename"]), "Store Name")
        p_num_col = next((c for c in df_pri.columns if str(c).strip().lower() in ["priority number", "priority", "priority list", "priority_number", "priority_list"]), "priority list")
        
        for _, row in df_pri.iterrows():
            store_name = str(row[p_store_col]).strip()
            try:
                pri_num = int(row[p_num_col])
            except ValueError:
                pri_num = 999999
            priority_map[store_name.lower()] = pri_num
            mapped_ebo = PRIORITY_TO_EBO.get(store_name)
            if mapped_ebo:
                priority_map[str(mapped_ebo).strip().lower()] = pri_num
        print(f"Loaded priority list mapping for {len(priority_map)} name variations.")
    except Exception as e:
        print(f"Warning: Failed to load priority list mapping ({e}). Falling back to alphabetical order.")

def get_store_priority(store_name):
    s_clean = str(store_name).strip().lower()
    if s_clean in priority_map:
        return priority_map[s_clean]
    for k, pri in priority_map.items():
        if k in s_clean:
            return pri
    return 999999

# The 51 target stores requested by the user
TARGET_STORES_MAP = {
    "PONDICHERRY": "TSPL PONDICHERRY",
    "BESANT NAGAR": "TSPL BESANT NAGAR EBO",
    "NAMAKKAL": "TSPL NAMAKKAL EBO",
    "VELLORE": "TSPL VELLORE EBO",
    "TEX VALLEY ERODE": "TSPL TEX VALLEY EBO",
    "(Salem 2) Seelanaickenpatti": "TSPL SALEM-2 EBO",
    "Selayur": "TSPL SELAYUR EBO",
    "TIRUPPUR": "TSPL TIRUPPUR",
    "SALEM": "TSPL SALEM",
    "ERODE": "TSPL ERODE STORE",
    "HSR LAYOUT": "TSPL HSR STORE",
    "CHIKKAJALA": "TSPL CHIKKAJALA EBO",
    "DODDABALAPUR": "TSPL DODDABALLA EBO",
    "Mysore Road - Bangalore": "TSPL MYSORE ROAD EBO",
    "ATTIBELLE": "TSPL ATTIBELE EBO",
    "KUVEMPUNAGAR - mysore 2": "TSPL MYSORE 2 EBO",
    "INORBIT MALL - HUBALI": "TSPL HUBBALI EBO",
    "HASSAN": "TSPL HASSAN EBO",
    "UDUPI": "TSPL UDUPI EBO",
    "HUBLI -2": "TSPL SHIRUR_PARK EBO",
    "Belgaum": "TSPL BELGAUM EBO",
    "Davangiri": "TSPL DAVANAGERE EBO",
    "Sarath City Mall": "TSPL SARATH CITY MALL",
    "Lakeshore": "TSPL HYDERABAD",
    "AS Rao": "TSPL AS NAGAR EBO",
    "Vijayawada": "TSPL VIJAYAWADA EBO",
    "Ananthpur": "TSPL ANANTAPUR EBO",
    "Vizianagram": "TSPL VIZIANAGARAM EBO",
    "In-orbit Vizag": "TSPL VIZAG EBO",
    "Khammam": "TSPL KHAMMAM EBO",
    "Pune Wagholi": "TSPL WAGHOLI EBO",
    "Moshi": "TSPL MOUSHI EBO",
    "Katraj": "TSPL KATRAJ EBO",
    "Pune - Kharadi": "TSPL PUNE KH",
    "Pune - Pimple Saudhagar": "TSPL PUNE PIMPLE",
    "Pune Hadapsar": "TSPL HADAPSAR EBO",
    "KOLHAPUR": "TSPL KOLHAPUR EBO",
    "Nashik": "TSPL NASHIK",
    "Capital Mall Vasai-Virar": "TSPL VASAI_1 EBO",
    "Bhumi": "TSPL BHIWANDI EBO",
    "Sambhajinagar": "TSPL DIVINITY MALL",
    "Bilashpur": "TSPL BILASPUR EBO",
    "Bhopal": "TSPL BHOPAL STORE",
    "Jabalpur": "TSPL JABALPUR EBO",
    "Raipur": "TSPL RAIPUR EBO",
    "Mani square mall": "TSPL MANI SQUARE MALL",
    "Sentrum mall": "TSPL SENTRUM MALL",
    "Rourkela": "TSPL ROURKELA EBO",
    "Bhagalpur": "TSPL BHAGALPUR EBO",
    "Kharagpur": "TSPL KHARAGPUR EBO",
    "Katagram": "TSPL KATARGAM EBO"
}

REVERSE_MAP = {v: k for k, v in TARGET_STORES_MAP.items()}

# Sort stores by priority rank
sorted_stores = sorted(list(TARGET_STORES_MAP.values()), key=lambda s: (get_store_priority(s), s.lower()))

base_cols    = ["Style", "SKU", "Size"]
final_cols   = base_cols.copy()
for store in sorted_stores:
    s_col = f"Stock_{store}"
    t_col = f"Transit_{store}"
    if s_col not in df_merged.columns:
        df_merged[s_col] = 0
    if t_col not in df_merged.columns:
        df_merged[t_col] = 0
    final_cols.append(s_col)
    final_cols.append(t_col)

df_out = df_merged[final_cols].copy()

# Friendly display names mapping back to user's 51 target store names
display_names = {}
display_names["Style"] = "Style"
display_names["SKU"] = "SKU"
display_names["Size"] = "Size"
for store in sorted_stores:
    clean_name = REVERSE_MAP.get(store, store)
    display_names[f"Stock_{store}"] = clean_name
    display_names[f"Transit_{store}"] = f"🚚 {clean_name}"

# ─────────────────────────────────────────────
# STEP 6 — Write Excel Report
# ─────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, "Footwear_SKU_Report.xlsx")

def save_report(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Footwear Report", startrow=1)
        ws = writer.sheets["Footwear Report"]

        # Group header row (row 1) — colour bands
        # Warm golden/bronze theme for Footwear to look professional and distinct
        for idx, col_name in enumerate(final_cols, start=1):
            cell = ws.cell(row=1, column=idx)
            if col_name in base_cols:
                cell.fill = PatternFill("solid", fgColor="5B3F11") # dark bronze
            elif col_name.startswith("Stock_"):
                cell.fill = PatternFill("solid", fgColor="8C6239") # medium bronze
            else:
                cell.fill = PatternFill("solid", fgColor="7F6000") # light gold/transit
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.value = display_names.get(col_name, col_name)

        # Data rows formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="FBF8F2") # soft gold tint for zebra striping

        # Column widths
        for idx, col_name in enumerate(final_cols, start=1):
            col_letter = get_column_letter(idx)
            if col_name == "SKU":
                ws.column_dimensions[col_letter].width = 22
            elif col_name == "Style":
                ws.column_dimensions[col_letter].width = 14
            elif col_name in base_cols:
                ws.column_dimensions[col_letter].width = 14
            else:
                ws.column_dimensions[col_letter].width = 16

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "E2"

        # Raw sheet (Stock Data)
        if not df_stock_foot.empty:
            df_raw = df_stock_foot[[style_s, sku_s, size_s, qty_s, sc_col, sn_col]].copy()
            df_raw.columns = ["Style", "SKU", "Size", "Qty", "Store Code", "Store Name"]
            df_raw.to_excel(writer, index=False, sheet_name="Raw Stock Data", startrow=1)
            ws2 = writer.sheets["Raw Stock Data"]
            apply_header_style(ws2, 1, fill_hex="5B3F11")
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
            for col in ws2.columns:
                ws2.column_dimensions[col[0].column_letter].width = 20

try:
    save_report(output_path)
except PermissionError:
    import time
    output_path = os.path.join(OUTPUT_DIR, f"Footwear_SKU_Report_{time.strftime('%H%M%S')}.xlsx")
    print(f"Warning: Footwear_SKU_Report.xlsx is locked. Saving as {os.path.basename(output_path)}")
    save_report(output_path)

print(f"\nDone! Output saved to:\n   {output_path}")
print(f"   Footwear SKUs  : {len(df_sku)}")
print(f"   Store columns  : {len(sorted_stores)} stores (stock & transit columns)")
