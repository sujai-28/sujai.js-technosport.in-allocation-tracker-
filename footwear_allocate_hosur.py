import os
import sys
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
STOCK_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock"
TRANSIT_DIR   = r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit"
PRIORITY_FILE = r"D:\INCREFF ORDER PUNCH\priority list\Priority list.xlsx"
HOSUR_STOCK_FILE = r"D:\downloads\Inventory Available for Sales - S103.xlsx"
OUTPUT_DIR    = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE"
OUTPUT_PATH   = os.path.join(OUTPUT_DIR, "Footwear_Hosur_Allocation_Report.xlsx")

TARGET_STYLES = {"S101", "S102", "S103"}

# 51 Target Stores Mapping
TARGET_STORES_MAP = {
    "PONDICHERRY": "TSPL PONDICHERRY",
    "BESANT NAGAR": "TSPL BESANT NAGAR EBO",
    "NAMAKKAL": "TSPL NAMAKKAL EBO",
    "VELLORE": "TSPL VELLORE EBO",
    "TEX VALLEY ERODE": "TSPL TEX VALLEY EBO",
    "(Salem 2) Seelanaickenpatti": "TSPL SALEM-2 EBO",
    "Selayur": "TSPL SELAYUR EBO",
    "TIRUPPUR": "TSPL TIRUPPUR",
    "SALEM": "TSPL SALEM",
    "ERODE": "TSPL ERODE STORE",
    "HSR LAYOUT": "TSPL HSR STORE",
    "CHIKKAJALA": "TSPL CHIKKAJALA EBO",
    "DODDABALAPUR": "TSPL DODDABALLA EBO",
    "Mysore Road - Bangalore": "TSPL MYSORE ROAD EBO",
    "ATTIBELLE": "TSPL ATTIBELE EBO",
    "KUVEMPUNAGAR - mysore 2": "TSPL MYSORE 2 EBO",
    "INORBIT MALL - HUBALI": "TSPL HUBBALI EBO",
    "HASSAN": "TSPL HASSAN EBO",
    "UDUPI": "TSPL UDUPI EBO",
    "HUBLI -2": "TSPL SHIRUR_PARK EBO",
    "Belgaum": "TSPL BELGAUM EBO",
    "Davangiri": "TSPL DAVANAGERE EBO",
    "Sarath City Mall": "TSPL SARATH CITY MALL",
    "Lakeshore": "TSPL HYDERABAD",
    "AS Rao": "TSPL AS NAGAR EBO",
    "Vijayawada": "TSPL VIJAYAWADA EBO",
    "Ananthpur": "TSPL ANANTAPUR EBO",
    "Vizianagram": "TSPL VIZIANAGARAM EBO",
    "In-orbit Vizag": "TSPL VIZAG EBO",
    "Khammam": "TSPL KHAMMAM EBO",
    "Pune Wagholi": "TSPL WAGHOLI EBO",
    "Moshi": "TSPL MOUSHI EBO",
    "Katraj": "TSPL KATRAJ EBO",
    "Pune - Kharadi": "TSPL PUNE KH",
    "Pune - Pimple Saudhagar": "TSPL PUNE PIMPLE",
    "Pune Hadapsar": "TSPL HADAPSAR EBO",
    "KOLHAPUR": "TSPL KOLHAPUR EBO",
    "Nashik": "TSPL NASHIK",
    "Capital Mall Vasai-Virar": "TSPL VASAI_1 EBO",
    "Bhumi": "TSPL BHIWANDI EBO",
    "Sambhajinagar": "TSPL DIVINITY MALL",
    "Bilashpur": "TSPL BILASPUR EBO",
    "Bhopal": "TSPL BHOPAL STORE",
    "Jabalpur": "TSPL JABALPUR EBO",
    "Raipur": "TSPL RAIPUR EBO",
    "Mani square mall": "TSPL MANI SQUARE MALL",
    "Sentrum mall": "TSPL SENTRUM MALL",
    "Rourkela": "TSPL ROURKELA EBO",
    "Bhagalpur": "TSPL BHAGALPUR EBO",
    "Kharagpur": "TSPL KHARAGPUR EBO",
    "Katagram": "TSPL KATARGAM EBO"
}

REVERSE_MAP = {v: k for k, v in TARGET_STORES_MAP.items()}

def get_latest_file(directory, ext='*.xlsx'):
    files = glob.glob(os.path.join(directory, ext))
    files = [f for f in files if not os.path.basename(f).startswith('~')]
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def clean_store_code(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        return s[:-2]
    return s

def clean_size(s):
    try:
        val = str(s).strip()
        if val.endswith('.0'):
            val = val[:-2]
        return val.upper()
    except:
        return str(s).strip().upper()

def get_base_stock(rank, size):
    sz_str = clean_size(size)
    try:
        sz = int(sz_str)
    except ValueError:
        return 0
        
    if 1 <= rank <= 4:
        if sz == 7:
            return 2
        elif sz in (8, 9, 10):
            return 3
        elif sz == 11:
            return 2
    elif rank >= 5:
        if sz == 7:
            return 1
        elif sz in (8, 9, 10):
            return 2
        elif sz == 11:
            return 1
    return 0

def main():
    print("=" * 65)
    print("  Footwear Priority-Based Allocation Simulator (Hosur)")
    print("=" * 65)

    # 1. Load Files
    stock_file = get_latest_file(STOCK_DIR, '*.xlsx')
    transit_file = get_latest_file(TRANSIT_DIR, '*.xlsx')
    
    if not stock_file:
        print("Error: Could not find current stock file.")
        sys.exit(1)
    if not os.path.exists(HOSUR_STOCK_FILE):
        print(f"Error: Could not find Hosur stock file at '{HOSUR_STOCK_FILE}'.")
        sys.exit(1)

    print(f"Store Stock  : {os.path.basename(stock_file)}")
    print(f"Store Transit: {os.path.basename(transit_file) if transit_file else 'None'}")
    print(f"Hosur Stock  : {os.path.basename(HOSUR_STOCK_FILE)}")

    # 2. Build Store Mapping
    store_map = {}
    df_stock_raw = pd.read_excel(stock_file)
    df_stock_raw.columns = [c.strip() for c in df_stock_raw.columns]
    
    s_sc = next((c for c in df_stock_raw.columns if c.lower() in ["store code", "store_code", "site_code", "site code", "owner site code"]), None)
    s_sn = next((c for c in df_stock_raw.columns if c.lower() in ["owner site", "store name", "ebo name"]), None)
    s_sku = next((c for c in df_stock_raw.columns if c.lower() == "sku"), None) or next((c for c in df_stock_raw.columns if c.lower() in ["barcode", "client sku id / ean"]), None)
    s_style = next((c for c in df_stock_raw.columns if c.lower() == "style"), "Style")
    s_color = next((c for c in df_stock_raw.columns if c.lower() in ["color", "colour"]), "Colour")
    s_size = next((c for c in df_stock_raw.columns if c.lower() == "size"), "Size")
    s_qty = next((c for c in df_stock_raw.columns if c.lower() in ["stock quantity", "quantity", "qty"]), "Qty")

    if s_sc and s_sn:
        for _, row in df_stock_raw.iterrows():
            code = clean_store_code(row[s_sc])
            name = str(row[s_sn]).strip()
            if code and name and name.lower() != 'nan':
                store_map[code] = name

    df_tran_raw = pd.DataFrame()
    if transit_file:
        df_tran_raw = pd.read_excel(transit_file)
        df_tran_raw.columns = [c.strip() for c in df_tran_raw.columns]
        t_sc = next((c for c in df_tran_raw.columns if c.lower() in ["store code", "store_code", "site code", "site_code"]), None)
        t_sn = next((c for c in df_tran_raw.columns if c.lower() in ["ebo name", "store name", "owner site"]), None)
        t_sku = next((c for c in df_tran_raw.columns if c.lower() == "sku"), None) or next((c for c in df_tran_raw.columns if c.lower() in ["barcode", "client sku id / ean"]), None)
        t_style = next((c for c in df_tran_raw.columns if c.lower() == "style"), "Style")
        t_color = next((c for c in df_tran_raw.columns if c.lower() in ["color", "colour"]), "Color")
        t_size = next((c for c in df_tran_raw.columns if c.lower() == "size"), "Size")
        t_qty = next((c for c in df_tran_raw.columns if c.lower() in ["transit qty", "quantity", "qty"]), "Transit Qty")

        if t_sc and t_sn:
            for _, row in df_tran_raw.iterrows():
                code = clean_store_code(row[t_sc])
                name = str(row[t_sn]).strip()
                if code and name and name.lower() != 'nan' and code not in store_map:
                    store_map[code] = name

    # Load store priorities
    from check_allocation import PRIORITY_TO_EBO
    priority_map = {}
    if os.path.exists(PRIORITY_FILE):
        try:
            df_pri = pd.read_excel(PRIORITY_FILE)
            df_pri.columns = [c.strip() for c in df_pri.columns]
            p_store_col = next((c for c in df_pri.columns if c.lower() in ["store name", "store_name", "storename"]), "Store Name")
            p_num_col = next((c for c in df_pri.columns if c.lower() in ["priority number", "priority", "priority list", "priority_number", "priority_list"]), "priority list")
            
            for _, row in df_pri.iterrows():
                store_name = str(row[p_store_col]).strip()
                try:
                    pri_num = int(row[p_num_col])
                except ValueError:
                    pri_num = 999999
                priority_map[store_name.lower()] = pri_num
                mapped_ebo = PRIORITY_TO_EBO.get(store_name)
                if mapped_ebo:
                    priority_map[str(mapped_ebo).strip().lower()] = pri_num
        except Exception as e:
            print(f"Warning: Failed to load priority list mapping ({e}).")

    def get_store_priority(store_name):
        s_clean = str(store_name).strip().lower()
        if s_clean in priority_map:
            return priority_map[s_clean]
        for k, pri in priority_map.items():
            if k in s_clean or s_clean in k:
                return pri
        return 999999

    # Filter stock and transit for S101, S102, S103
    df_stock_filtered = df_stock_raw[df_stock_raw[s_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
    df_stock_filtered["Style_Clean"] = df_stock_filtered[s_style].astype(str).str.strip().str.upper()
    df_stock_filtered["Color_Clean"] = df_stock_filtered[s_color].astype(str).str.strip().str.upper()
    df_stock_filtered["Size_Clean"] = df_stock_filtered[s_size].apply(clean_size)
    df_stock_filtered["SKU_Clean"] = df_stock_filtered[s_sku].astype(str).str.strip().str.upper()
    df_stock_filtered["Store_Code_Clean"] = df_stock_filtered[s_sc].apply(clean_store_code)
    df_stock_filtered["Store_Name_Clean"] = df_stock_filtered.apply(
        lambda r: store_map.get(r["Store_Code_Clean"], str(r[s_sn]).strip() if s_sn else r["Store_Code_Clean"]),
        axis=1
    )
    df_stock_filtered["Qty_Clean"] = pd.to_numeric(df_stock_filtered[s_qty], errors="coerce").fillna(0).astype(int)

    df_tran_filtered = pd.DataFrame()
    if not df_tran_raw.empty:
        df_tran_filtered = df_tran_raw[df_tran_raw[t_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
        df_tran_filtered["Style_Clean"] = df_tran_filtered[t_style].astype(str).str.strip().str.upper()
        df_tran_filtered["Color_Clean"] = df_tran_filtered[t_color].astype(str).str.strip().str.upper()
        df_tran_filtered["Size_Clean"] = df_tran_filtered[t_size].apply(clean_size)
        df_tran_filtered["SKU_Clean"] = df_tran_filtered[t_sku].astype(str).str.strip().str.upper()
        df_tran_filtered["Store_Code_Clean"] = df_tran_filtered[t_sc].apply(clean_store_code)
        df_tran_filtered["Store_Name_Clean"] = df_tran_filtered.apply(
            lambda r: store_map.get(r["Store_Code_Clean"], str(r[t_sn]).strip() if t_sn else r["Store_Code_Clean"]),
            axis=1
        )
        df_tran_filtered["Qty_Clean"] = pd.to_numeric(df_tran_filtered[t_qty], errors="coerce").fillna(0).astype(int)

    # 3. Load Hosur Stock
    df_hosur = pd.read_excel(HOSUR_STOCK_FILE)
    df_hosur.columns = [c.strip() for c in df_hosur.columns]
    
    wh_sku_col = next(c for c in df_hosur.columns if c.strip().lower() in ["client sku id / ean", "client sku id/ean", "barcode", "sku"])
    wh_qty_col = next(c for c in df_hosur.columns if c.strip().lower() in ["total available quantity", "available qty", "quantity", "qty"])
    wh_style_col = next(c for c in df_hosur.columns if c.strip().lower() == "style")
    wh_color_col = next(c for c in df_hosur.columns if c.strip().lower() == "color")
    wh_size_col = next(c for c in df_hosur.columns if c.strip().lower() == "size")

    # Clean warehouse inventory
    df_hosur["SKU_Clean"] = df_hosur[wh_sku_col].astype(str).str.strip().str.upper()
    df_hosur["Available_Clean"] = pd.to_numeric(df_hosur[wh_qty_col], errors="coerce").fillna(0).astype(int)
    
    # Store initial available quantities for remaining sheet
    wh_initial_stock = df_hosur.groupby("SKU_Clean")["Available_Clean"].sum().to_dict()
    wh_stock = wh_initial_stock.copy()

    # Create SKU info map for details sheet
    df_hosur_styles = df_hosur[[wh_style_col, wh_color_col, wh_size_col, "SKU_Clean"]].drop_duplicates()
    sku_info = {}
    for _, r in df_hosur_styles.iterrows():
        sku_info[r["SKU_Clean"]] = {
            "Style": str(r[wh_style_col]).strip().upper(),
            "Color": str(r[wh_color_col]).strip().upper(),
            "Size": clean_size(r[wh_size_col])
        }

    # 4. Resolve Target Stores list
    sorted_stores = sorted(list(TARGET_STORES_MAP.values()), key=lambda s: (get_store_priority(s), s.lower()))

    # Build Master list of all SKUs from store stock, transit and warehouse stock
    skus_stock = df_stock_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
    skus_tran = pd.DataFrame()
    if not df_tran_filtered.empty:
        skus_tran = df_tran_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
    skus_wh = pd.DataFrame()
    wh_recs = []
    for s_clean, info in sku_info.items():
        wh_recs.append({
            "Style_Clean": info["Style"],
            "Color_Clean": info["Color"],
            "Size_Clean": info["Size"],
            "SKU_Clean": s_clean
        })
    if wh_recs:
        skus_wh = pd.DataFrame(wh_recs)

    df_skus_master = pd.concat([skus_stock, skus_tran, skus_wh], ignore_index=True).drop_duplicates()
    
    # Sort Master list by Style, Color, Size
    def size_sort_key(sz):
        try:
            return int(sz)
        except:
            return 999
    df_skus_master["size_int"] = df_skus_master["Size_Clean"].apply(size_sort_key)
    df_skus_master = df_skus_master.sort_values(by=["Style_Clean", "Color_Clean", "size_int"]).drop(columns=["size_int"]).reset_index(drop=True)

    # 5. Pivot Stock and Transit data
    stock_pivot = (
        df_stock_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
        .sum()
        .unstack(fill_value=0)
    )
    stock_pivot.columns = [f"Stock_{c}" for c in stock_pivot.columns]
    
    transit_pivot = pd.DataFrame()
    if not df_tran_filtered.empty:
        transit_pivot = (
            df_tran_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
            .sum()
            .unstack(fill_value=0)
        )
        transit_pivot.columns = [f"Transit_{c}" for c in transit_pivot.columns]

    df_merged = df_skus_master.set_index(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"])
    if not stock_pivot.empty:
        df_merged = df_merged.join(stock_pivot, how="left")
    if not transit_pivot.empty:
        df_merged = df_merged.join(transit_pivot, how="left")
    df_merged = df_merged.fillna(0).reset_index()

    # 6. Priority-Based Allocation simulation loop
    print("Simulating allocation...")
    allocations = {} # store -> sku -> allocated_qty
    for store in sorted_stores:
        allocations[store] = {}
        for sku in df_skus_master["SKU_Clean"].unique():
            allocations[store][sku] = 0

    detailed_rows = []
    
    for rank, store in enumerate(sorted_stores, start=1):
        for _, r in df_skus_master.iterrows():
            style = r["Style_Clean"]
            color = r["Color_Clean"]
            size = r["Size_Clean"]
            sku = r["SKU_Clean"]

            base_stock = get_base_stock(rank, size)
            if base_stock == 0:
                continue

            # Fetch current stock and transit
            s_col = f"Stock_{store}"
            t_col = f"Transit_{store}"
            
            curr_stock = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                s_col
            ].iloc[0]) if s_col in df_merged.columns else 0

            transit_qty = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                t_col
            ].iloc[0]) if t_col in df_merged.columns else 0

            # Calculate gap/shortfall
            gap = max(0, base_stock - (curr_stock + transit_qty))
            allocated_qty = 0

            if gap > 0:
                available_wh = wh_stock.get(sku, 0)
                allocated_qty = min(gap, available_wh)
                wh_stock[sku] = available_wh - allocated_qty
                allocations[store][sku] = allocated_qty

            remaining_shortfall = gap - allocated_qty
            detailed_rows.append({
                "Priority Rank": rank,
                "Store Name": store,
                "Style": style,
                "Color": color,
                "Size": size,
                "SKU": sku,
                "Base Stock": base_stock,
                "Current Stock": curr_stock,
                "Transit Qty": transit_qty,
                "Shortfall Qty": gap,
                "Allocated Qty (Hosur)": allocated_qty,
                "Remaining Shortfall": remaining_shortfall
            })

    df_detail_alloc = pd.DataFrame(detailed_rows)

    # 7. Create Store Allocation Summary
    summary_rows = []
    for rank, store in enumerate(sorted_stores, start=1):
        store_rows = df_detail_alloc[df_detail_alloc["Store Name"] == store]
        s_7 = store_rows[store_rows["Size"] == "7"]["Allocated Qty (Hosur)"].sum()
        s_8 = store_rows[store_rows["Size"] == "8"]["Allocated Qty (Hosur)"].sum()
        s_9 = store_rows[store_rows["Size"] == "9"]["Allocated Qty (Hosur)"].sum()
        s_10 = store_rows[store_rows["Size"] == "10"]["Allocated Qty (Hosur)"].sum()
        s_11 = store_rows[store_rows["Size"] == "11"]["Allocated Qty (Hosur)"].sum()
        total_alloc = store_rows["Allocated Qty (Hosur)"].sum()
        group = "Top 4" if rank <= 4 else "Remaining"
        
        summary_rows.append({
            "Priority Rank": rank,
            "Group": group,
            "Store Name": store,
            "Size 7 Allocated": s_7,
            "Size 8 Allocated": s_8,
            "Size 9 Allocated": s_9,
            "Size 10 Allocated": s_10,
            "Size 11 Allocated": s_11,
            "Total Allocated": total_alloc
        })
    df_summary_alloc = pd.DataFrame(summary_rows)

    # 8. Create Pivoted Sheet
    df_pivot_out = df_skus_master.copy()
    df_pivot_out.rename(columns={
        "Style_Clean": "Style",
        "Color_Clean": "Color",
        "Size_Clean": "Size",
        "SKU_Clean": "SKU"
    }, inplace=True)

    for store in sorted_stores:
        s_col = f"Stock_{store}"
        t_col = f"Transit_{store}"
        
        # Add stock & transit to pivot
        if s_col not in df_merged.columns:
            df_pivot_out[s_col] = 0
        else:
            df_pivot_out[s_col] = df_merged[s_col]
            
        if t_col not in df_merged.columns:
            df_pivot_out[t_col] = 0
        else:
            df_pivot_out[t_col] = df_merged[t_col]
            
        # Add allocation column
        alloc_col = f"Alloc_{store}"
        alloc_list = []
        for _, row in df_pivot_out.iterrows():
            alloc_list.append(allocations[store].get(row["SKU"], 0))
        df_pivot_out[alloc_col] = alloc_list

    # 9. Create Warehouse remaining inventory sheet
    wh_rem_rows = []
    for sku, initial in wh_initial_stock.items():
        rem = wh_stock.get(sku, 0)
        allocated = initial - rem
        info = sku_info.get(sku, {"Style": "-", "Color": "-", "Size": "-"})
        wh_rem_rows.append({
            "Style": info["Style"],
            "Color": info["Color"],
            "Size": info["Size"],
            "SKU": sku,
            "Initial Stock": initial,
            "Allocated Stock": allocated,
            "Remaining Stock": rem
        })
    df_wh_rem = pd.DataFrame(wh_rem_rows)
    df_wh_rem.sort_values(by=["Style", "Color", "Size"], inplace=True)

    # Apply REVERSE_MAP to make Store Name values user friendly in sheets
    df_summary_alloc["Store Name"] = df_summary_alloc["Store Name"].map(lambda s: REVERSE_MAP.get(s, s))
    df_detail_alloc["Store Name"] = df_detail_alloc["Store Name"].map(lambda s: REVERSE_MAP.get(s, s))

    # 10. Write Excel
    print(f"Saving Excel report to: {OUTPUT_PATH}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Check for file locks
    actual_path = OUTPUT_PATH
    try:
        writer = pd.ExcelWriter(actual_path, engine="openpyxl")
    except PermissionError:
        import time
        actual_path = OUTPUT_PATH.replace(".xlsx", f"_{time.strftime('%H%M%S')}.xlsx")
        print(f"Warning: File is locked. Saving to: {actual_path}")
        writer = pd.ExcelWriter(actual_path, engine="openpyxl")

    def style_sheet(ws, df, title, freeze_pane="D3"):
        bronze_fill = PatternFill("solid", fgColor="5B3F11")
        alt_fill = PatternFill("solid", fgColor="FBF8F2")
        title_font = Font(name="Calibri", size=14, bold=True, color="5B3F11")
        hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_side = Side(style="thin", color="E0D0C0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = left_align
        ws.row_dimensions[1].height = 30
        
        ws.row_dimensions[2].height = 25
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = bronze_fill
            cell.font = hdr_font
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx in range(3, len(df) + 3):
            ws.row_dimensions[row_idx].height = 18
            use_alt = (row_idx % 2 == 0)
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if use_alt:
                    cell.fill = alt_fill
                
                col_name = df.columns[col_idx - 1]
                if col_name in ("Store Name", "SKU", "Style", "Color"):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                
                if isinstance(cell.value, (int, float)):
                    cell.value = int(cell.value)
                
                if "Allocated" in col_name or "Total" in col_name:
                    if cell.value and cell.value > 0:
                        cell.font = bold_data_font
        
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.freeze_panes = freeze_pane

    with writer:
        df_summary_alloc.to_excel(writer, sheet_name="Allocation Summary", index=False, startrow=1)
        df_detail_alloc.to_excel(writer, sheet_name="Detailed Allocation", index=False, startrow=1)
        df_pivot_out.to_excel(writer, sheet_name="Pivoted Allocation Report", index=False, startrow=1)
        df_wh_rem.to_excel(writer, sheet_name="Warehouse Remaining Stock", index=False, startrow=1)

        style_sheet(writer.sheets["Allocation Summary"], df_summary_alloc, "Footwear Store-Wise Allocation Summary Sheet", "D3")
        style_sheet(writer.sheets["Detailed Allocation"], df_detail_alloc, "Footwear Detailed Option-Wise Allocation Sheet", "C3")
        style_sheet(writer.sheets["Warehouse Remaining Stock"], df_wh_rem, "Hosur Warehouse Initial vs Remaining Inventory Sheet", "E3")

        # Style pivoted sheet specially
        ws_piv = writer.sheets["Pivoted Allocation Report"]
        ws_piv.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_pivot_out.columns))
        title_cell = ws_piv.cell(row=1, column=1)
        title_cell.value = "Footwear Store-Wise Stock, Transit, and Simulated Allocations (Hosur)"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="5B3F11")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_piv.row_dimensions[1].height = 30
        ws_piv.row_dimensions[2].height = 35

        # Format column headers with emojis
        display_cols = []
        for c in df_pivot_out.columns:
            if c.startswith("Stock_"):
                orig_name = c.replace("Stock_", "")
                display_cols.append(REVERSE_MAP.get(orig_name, orig_name))
            elif c.startswith("Transit_"):
                orig_name = c.replace("Transit_", "")
                display_cols.append(f"🚚 {REVERSE_MAP.get(orig_name, orig_name)}")
            elif c.startswith("Alloc_"):
                orig_name = c.replace("Alloc_", "")
                display_cols.append(f"📦 {REVERSE_MAP.get(orig_name, orig_name)}")
            else:
                display_cols.append(c)

        for idx, col_name in enumerate(df_pivot_out.columns, start=1):
            cell = ws_piv.cell(row=2, column=idx)
            cell.value = display_cols[idx - 1]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            
            if col_name in ["Style", "Color", "Size", "SKU"]:
                cell.fill = PatternFill("solid", fgColor="5B3F11")
            elif col_name.startswith("Stock_"):
                cell.fill = PatternFill("solid", fgColor="8C6239")
            elif col_name.startswith("Transit_"):
                cell.fill = PatternFill("solid", fgColor="7F6000")
            else:
                cell.fill = PatternFill("solid", fgColor="2e7d32") # dark green for allocation

        # Style data rows
        thin_side = Side(style="thin", color="E0D0C0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for row_idx in range(3, len(df_pivot_out) + 3):
            ws_piv.row_dimensions[row_idx].height = 18
            use_alt = (row_idx % 2 == 0)
            for col_idx in range(1, len(df_pivot_out.columns) + 1):
                cell = ws_piv.cell(row=row_idx, column=col_idx)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border
                if use_alt:
                    cell.fill = PatternFill("solid", fgColor="FBF8F2")
                    
                col_name = df_pivot_out.columns[col_idx - 1]
                if col_name in ("Style", "Color", "SKU"):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                if isinstance(cell.value, (int, float)):
                    cell.value = int(cell.value)

        # Column widths for Pivoted sheet
        for col in ws_piv.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_piv.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws_piv.freeze_panes = "E3"

    print(f"SUCCESS! Allocation file saved to:\n  {actual_path}")
    print(f"Total allocated units: {df_detail_alloc['Allocated Qty (Hosur)'].sum()} units")

if __name__ == "__main__":
    main()
