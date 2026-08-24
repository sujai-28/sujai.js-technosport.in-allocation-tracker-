import os
import sys
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
INVENTORY_CSV = r"D:\downloads\Inventory Available for Sales - OMS-2026-07-07T10_29_29.434+05_30.csv"
REPORT_PATH   = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE\Plus_Size_SKU_Report.xlsx"
OUTPUT_PATH   = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE\Plus_Size_Shortfall_Report.xlsx"

# ─────────────────────────────────────────────
# BASE STOCK SETTINGS
# ─────────────────────────────────────────────
def clean_size(s):
    return str(s).strip().upper().replace(" ", "")

# Top 5 stores: 3XL=4, 4XL=3, 5XL=3
# Next 6 stores: 3XL=3, 4XL=2, 5XL=2
def get_base_stock(rank, size):
    sz = clean_size(size)
    if 1 <= rank <= 5:
        if sz == "3XL":
            return 4
        elif sz in ("4XL", "5XL"):
            return 3
    elif rank >= 6:
        if sz == "3XL":
            return 3
        elif sz in ("4XL", "5XL"):
            return 2
    return 0

def main():
    print("Loading plus size report...")
    if not os.path.exists(REPORT_PATH):
        print(f"Error: Base report not found at {REPORT_PATH}")
        sys.exit(1)
        
    # Read the Plus Size Report sheet, using row 2 (index 1) as the header
    df = pd.read_excel(REPORT_PATH, sheet_name="Plus Size Report", header=1)
    
    # Identify ordered store names from Stock columns
    stock_cols = [c for c in df.columns if str(c).startswith("Stock_")]
    ordered_stores = [c.replace("Stock_", "") for c in stock_cols]
    
    if not ordered_stores:
        print("Error: No store stock columns found in the report.")
        sys.exit(1)
        
    print(f"Found {len(ordered_stores)} stores in priority order.")
    
    # Evaluate all stores in priority list
    target_stores = ordered_stores
    print(f"Evaluating shortfalls for all {len(target_stores)} stores...")
    
    # List to store detailed shortfall rows
    shortfall_rows = []
    
    # Evaluate shortfall row-by-row
    for _, row in df.iterrows():
        style = str(row["Style"]).strip()
        sku = str(row["SKU"]).strip()
        size = str(row["Size"]).strip()
        wh_avail = int(row["WH_Available_Qty"])
        
        for rank, store in enumerate(target_stores, start=1):
            base_stock = get_base_stock(rank, size)
            if base_stock == 0:
                continue
                
            s_col = f"Stock_{store}"
            t_col = f"Transit_{store}"
            a_col = f"Alloc_{store}"
            
            curr_stock = int(row[s_col]) if s_col in df.columns else 0
            transit_qty = int(row[t_col]) if t_col in df.columns else 0
            alloc_qty = int(row[a_col]) if a_col in df.columns else 0
            
            net_stock = curr_stock + transit_qty + alloc_qty
            
            # Shortfall relative to net inventory (stock + transit + alloc)
            shortfall = max(0, base_stock - net_stock)
            
            if shortfall > 0:
                shortfall_rows.append({
                    "Priority Rank": rank,
                    "Store Name": store,
                    "Style": style,
                    "SKU": sku,
                    "Size": size,
                    "Base Stock": base_stock,
                    "Current Stock": curr_stock,
                    "Transit Qty": transit_qty,
                    "Allocated Qty": alloc_qty,
                    "Total Store Qty": net_stock,
                    "Shortfall Qty": shortfall,
                    "WH Available Qty": wh_avail
                })
                
    df_detail = pd.DataFrame(shortfall_rows)
    
    # ─────────────────────────────────────────────
    # GENERATE POOL-WISE INVENTORY DATA
    # ─────────────────────────────────────────────
    print("Loading OMS inventory for pool-wise analysis...")
    if not os.path.exists(INVENTORY_CSV):
        print(f"Error: OMS Inventory file not found at {INVENTORY_CSV}")
        sys.exit(1)
        
    df_inv = pd.read_csv(INVENTORY_CSV)
    df_inv.columns = [c.strip() for c in df_inv.columns]
    
    # Filter plus sizes
    df_inv["_size_clean"] = df_inv["Size"].apply(clean_size)
    plus_sizes_clean = {clean_size(p) for p in ["3XL", "4XL", "5XL", "3 XL", "4 XL", "5 XL", "XXXL", "XXXXL", "XXXXXL"]}
    plus_mask = df_inv["_size_clean"].isin(plus_sizes_clean)
    df_inv_plus = df_inv[plus_mask].copy()
    
    # Rename columns for clarity
    df_inv_plus = df_inv_plus.rename(columns={
        "Client SKU Id / EAN": "SKU",
        "Total Available Quantity": "Available Qty",
        "Total Allocated Quantity": "Allocated Qty",
        "Total Quantity": "Total Qty"
    })
    
    # Detail by SKU & Pool
    df_pool_detail = (
        df_inv_plus.groupby(["Style", "SKU", "Size", "Reservation Pool"], as_index=False)[
            ["Available Qty", "Allocated Qty", "Total Qty"]
        ].sum()
    )
    df_pool_detail = df_pool_detail.sort_values(["Style", "SKU", "Reservation Pool"]).reset_index(drop=True)
    
    # Summary by Pool
    df_pool_summary = (
        df_inv_plus.groupby("Reservation Pool", as_index=False)[
            ["Available Qty", "Allocated Qty", "Total Qty"]
        ].sum()
    )
    df_pool_summary = df_pool_summary.sort_values("Available Qty", ascending=False).reset_index(drop=True)
    
    # Build lookup for SKU pool quantities: (SKU, Pool) -> Available Qty
    pool_qty_lookup = {}
    for _, row in df_pool_detail.iterrows():
        sku_val = str(row["SKU"]).strip()
        pool_val = str(row["Reservation Pool"]).strip()
        avail = int(row["Available Qty"])
        pool_qty_lookup[(sku_val, pool_val)] = avail

    # ─────────────────────────────────────────────
    # PERFORM POOL-PRIORITY WAREHOUSE ALLOCATION
    # ─────────────────────────────────────────────
    if df_detail.empty:
        print("No shortfalls found. All target stores meet their base stock requirements.")
        df_detail = pd.DataFrame(columns=[
            "Priority Rank", "Store Name", "Style", "SKU", "Size", "Base Stock",
            "Current Stock", "Transit Qty", "Allocated Qty", "Total Store Qty",
            "Shortfall Qty", "WH Available Qty"
        ])
        df_alloc_plan = pd.DataFrame(columns=[
            "Priority Rank", "Store Name", "Style", "SKU", "Size",
            "Shortfall Qty", "Allocated from EBO", "Allocated from D2C", 
            "Allocated from Other", "Total Allocated", "Remaining Shortfall",
            "WH EBO Qty", "WH D2C Qty", "WH Other Qty"
        ])
    else:
        print("Performing warehouse inventory allocation (EBO -> D2C -> Other) by store priority...")
        alloc_rows = []
        
        # Group by SKU to allocate available stock in pool order and store priority order
        for sku_val, group in df_detail.groupby("SKU"):
            group_sorted = group.sort_values("Priority Rank")
            
            # Available stock in each pool for this SKU
            wh_ebo = pool_qty_lookup.get((sku_val, "EBO"), 0)
            wh_d2c = pool_qty_lookup.get((sku_val, "D2C-Marketplaces"), 0)
            
            wh_other = 0
            for (s, p), qty in pool_qty_lookup.items():
                if s == sku_val and p not in ("EBO", "D2C-Marketplaces"):
                    wh_other += qty
            
            rem_ebo = wh_ebo
            rem_d2c = wh_d2c
            rem_other = wh_other
            
            # Initialize remaining shortfall and allocations for each store in the group
            store_needs = {}
            allocated_ebo = {}
            allocated_d2c = {}
            allocated_other = {}
            
            for _, row in group_sorted.iterrows():
                store_name = row["Store Name"]
                store_needs[store_name] = int(row["Shortfall Qty"])
                allocated_ebo[store_name] = 0
                allocated_d2c[store_name] = 0
                allocated_other[store_name] = 0
            
            # Pass 1: Try to satisfy all stores using EBO pool only
            for _, row in group_sorted.iterrows():
                store_name = row["Store Name"]
                alloc = min(rem_ebo, store_needs[store_name])
                rem_ebo -= alloc
                store_needs[store_name] -= alloc
                allocated_ebo[store_name] = alloc
                
            # Pass 2: Try to satisfy remaining needs using D2C pool only
            for _, row in group_sorted.iterrows():
                store_name = row["Store Name"]
                alloc = min(rem_d2c, store_needs[store_name])
                rem_d2c -= alloc
                store_needs[store_name] -= alloc
                allocated_d2c[store_name] = alloc
                
            # Pass 3: Try to satisfy remaining needs using Other pools
            for _, row in group_sorted.iterrows():
                store_name = row["Store Name"]
                alloc = min(rem_other, store_needs[store_name])
                rem_other -= alloc
                store_needs[store_name] -= alloc
                allocated_other[store_name] = alloc
                
            # Append rows with results
            for _, row in group_sorted.iterrows():
                store_name = row["Store Name"]
                alloc_ebo = allocated_ebo[store_name]
                alloc_d2c = allocated_d2c[store_name]
                alloc_other = allocated_other[store_name]
                tot_alloc = alloc_ebo + alloc_d2c + alloc_other
                rem_sf = store_needs[store_name]
                
                alloc_rows.append({
                    "Priority Rank": row["Priority Rank"],
                    "Store Name": store_name,
                    "Style": row["Style"],
                    "SKU": row["SKU"],
                    "Size": row["Size"],
                    "Shortfall Qty": row["Shortfall Qty"],
                    "Allocated from EBO": alloc_ebo,
                    "Allocated from D2C": alloc_d2c,
                    "Allocated from Other": alloc_other,
                    "Total Allocated": tot_alloc,
                    "Remaining Shortfall": rem_sf,
                    "WH EBO Qty": wh_ebo,
                    "WH D2C Qty": wh_d2c,
                    "WH Other Qty": wh_other
                })
                
        df_alloc_plan = pd.DataFrame(alloc_rows)
        df_alloc_plan = df_alloc_plan.sort_values(["Priority Rank", "Style", "SKU"]).reset_index(drop=True)

    print(f"Generated {len(df_detail)} shortfall rows.")
    print(f"Generated {len(df_alloc_plan)} allocation plan rows.")
    
    # ─────────────────────────────────────────────
    # GENERATE HIGH-LEVEL SUMMARY BY STORE
    # ─────────────────────────────────────────────
    summary_data = []
    for rank, store in enumerate(target_stores, start=1):
        store_rows = df_detail[df_detail["Store Name"] == store]
        s_3xl = store_rows[store_rows["Size"] == "3XL"]["Shortfall Qty"].sum()
        s_4xl = store_rows[store_rows["Size"] == "4XL"]["Shortfall Qty"].sum()
        s_5xl = store_rows[store_rows["Size"] == "5XL"]["Shortfall Qty"].sum()
        total_shortfall = store_rows["Shortfall Qty"].sum()
        
        if not df_alloc_plan.empty:
            store_alloc = df_alloc_plan[df_alloc_plan["Store Name"] == store]
            alloc_ebo = store_alloc["Allocated from EBO"].sum()
            alloc_d2c = store_alloc["Allocated from D2C"].sum()
            alloc_other = store_alloc["Allocated from Other"].sum()
            total_allocated = store_alloc["Total Allocated"].sum()
            total_remaining = store_alloc["Remaining Shortfall"].sum()
        else:
            alloc_ebo = alloc_d2c = alloc_other = total_allocated = total_remaining = 0
            
        group = "Top 5" if rank <= 5 else "Remaining"
        
        summary_data.append({
            "Priority Rank": rank,
            "Group": group,
            "Store Name": store,
            "3XL Shortfall": s_3xl,
            "4XL Shortfall": s_4xl,
            "5XL Shortfall": s_5xl,
            "Total Shortfall": total_shortfall,
            "Allocated from EBO": alloc_ebo,
            "Allocated from D2C": alloc_d2c,
            "Allocated from Other": alloc_other,
            "Total Allocated": total_allocated,
            "Total Rem Shortfall": total_remaining
        })
        
    df_summary = pd.DataFrame(summary_data)
    
    # ─────────────────────────────────────────────
    # WRITE EXCEL WITH OPENPYXL STYLING
    # ─────────────────────────────────────────────
    print(f"Writing output report to {OUTPUT_PATH}...")
    
    def do_save(p):
        with pd.ExcelWriter(p, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Shortfall Summary", index=False, startrow=1)
            df_detail.to_excel(writer, sheet_name="Detailed Shortfall", index=False, startrow=1)
            df_alloc_plan.to_excel(writer, sheet_name="Final Allocation Plan", index=False, startrow=1)
            df_pool_summary.to_excel(writer, sheet_name="Pool Summary", index=False, startrow=1)
            df_pool_detail.to_excel(writer, sheet_name="Pool Detail", index=False, startrow=1)
            
            ws_sum = writer.sheets["Shortfall Summary"]
            style_sheet(ws_sum, df_summary, title="Plus-Size Allocation Shortfall & WH Allocation Summary")
            
            ws_det = writer.sheets["Detailed Shortfall"]
            style_sheet(ws_det, df_detail, title="Detailed SKU Shortfall Report (Target Base Stock)")
            
            ws_alc = writer.sheets["Final Allocation Plan"]
            style_sheet(ws_alc, df_alloc_plan, title="Final Priority-Based Allocation Plan from Warehouse Stock")
            
            ws_p_sum = writer.sheets["Pool Summary"]
            style_sheet(ws_p_sum, df_pool_summary, title="Plus-Size Warehouse Inventory Pool Summary")
            
            ws_p_det = writer.sheets["Pool Detail"]
            style_sheet(ws_p_det, df_pool_detail, title="Detailed SKU Inventory by Reservation Pool")

    try:
        do_save(OUTPUT_PATH)
        print(f"Report completed successfully! Saved to: {OUTPUT_PATH}")
    except PermissionError:
        import time
        alt_path = OUTPUT_PATH.replace(".xlsx", f"_{time.strftime('%H%M%S')}.xlsx")
        print(f"Warning: {OUTPUT_PATH} is locked. Saving to {alt_path} instead.")
        do_save(alt_path)
        print(f"Report completed successfully! Saved to: {alt_path}")

    # ─────────────────────────────────────────────
    # WRITE SIMPLIFIED REPLENISHMENT ORDER FILE
    # ─────────────────────────────────────────────
    ORDER_OUTPUT_PATH = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE\Plus_Size_Store_Replenishment_Order.xlsx"
    ORDER_CSV_PATH = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE\Plus_Size_Store_Replenishment_Order.csv"
    
    if not df_alloc_plan.empty:
        df_order = df_alloc_plan[df_alloc_plan["Total Allocated"] > 0][["Store Name", "SKU", "Total Allocated"]].copy()
        df_order.columns = ["Store Name", "SKU", "Req Rep Qty"]
    else:
        df_order = pd.DataFrame(columns=["Store Name", "SKU", "Req Rep Qty"])
        
    print(f"Writing simplified order file with {len(df_order)} rows...")
    
    def save_order(p_xlsx, p_csv):
        df_order.to_csv(p_csv, index=False)
        with pd.ExcelWriter(p_xlsx, engine="openpyxl") as writer:
            df_order.to_excel(writer, sheet_name="Replenishment Order", index=False, startrow=1)
            ws = writer.sheets["Replenishment Order"]
            style_sheet(ws, df_order, title="Plus-Size Store Replenishment Order")
            
    try:
        save_order(ORDER_OUTPUT_PATH, ORDER_CSV_PATH)
        print(f"Order file saved to: {ORDER_OUTPUT_PATH}")
    except PermissionError:
        import time
        alt_xlsx = ORDER_OUTPUT_PATH.replace(".xlsx", f"_{time.strftime('%H%M%S')}.xlsx")
        alt_csv = ORDER_CSV_PATH.replace(".csv", f"_{time.strftime('%H%M%S')}.csv")
        print(f"Warning: Order files are locked. Saving to {alt_xlsx} instead.")
        save_order(alt_xlsx, alt_csv)

def style_sheet(ws, df, title):
    # Colors
    navy_fill = PatternFill("solid", fgColor="1F3864")
    alt_fill = PatternFill("solid", fgColor="F2F4F8")
    
    # Soft distinct fills for pool sources
    ebo_fill = PatternFill("solid", fgColor="E2EFDA")    # soft green
    d2c_fill = PatternFill("solid", fgColor="DDEBF7")    # soft blue
    other_fill = PatternFill("solid", fgColor="FFF2CC")  # soft yellow/orange
    
    # Fonts
    title_font = Font(name="Calibri", size=14, bold=True, color="1F3864")
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    bold_data_font = Font(name="Calibri", size=10, bold=True)
    
    ebo_font = Font(name="Calibri", size=10, color="375623", bold=True)
    d2c_font = Font(name="Calibri", size=10, color="1F4E79", bold=True)
    other_font = Font(name="Calibri", size=10, color="7F6000", bold=True)
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Borders
    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Write Title Block in row 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = left_align
    ws.row_dimensions[1].height = 30
    
    # Style Header Row (row 2)
    ws.row_dimensions[2].height = 25
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = navy_fill
        cell.font = hdr_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Style Data Rows
    for row_idx in range(3, len(df) + 3):
        ws.row_dimensions[row_idx].height = 18
        use_alt = (row_idx % 2 == 0)
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            
            # Default backgrounds
            if use_alt:
                cell.fill = alt_fill
                
            # Alignments
            col_name = df.columns[col_idx - 1]
            if col_name in ("Store Name", "SKU", "Style", "Reservation Pool"):
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
            # Formatting numeric values as integers if possible
            if isinstance(cell.value, (int, float)):
                if cell.value == 0:
                    cell.value = 0
                else:
                    cell.value = int(cell.value)
                    
            # Highlight shortfall and allocation columns
            if "Shortfall" in col_name or "Total" in col_name or "Allocated" in col_name:
                if cell.value and cell.value > 0:
                    cell.font = bold_data_font
            
            # Color-code allocation source columns
            if "Allocated from EBO" in col_name:
                if cell.value and cell.value > 0:
                    cell.fill = ebo_fill
                    cell.font = ebo_font
            elif "Allocated from D2C" in col_name:
                if cell.value and cell.value > 0:
                    cell.fill = d2c_fill
                    cell.font = d2c_font
            elif "Allocated from Other" in col_name:
                if cell.value and cell.value > 0:
                    cell.fill = other_fill
                    cell.font = other_font
                    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Check lengths starting from header row
        for cell in col[1:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.freeze_panes = "D3"

if __name__ == "__main__":
    main()
