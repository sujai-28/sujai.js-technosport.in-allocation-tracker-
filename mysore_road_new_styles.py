"""
mysore_road_new_styles.py
--------------------------
Generates a NEW STYLES allocation plan for TSPL MYSORE ROAD EBO.

Logic:
  1. Load stock + in-transit + allocation for TSPL MYSORE ROAD EBO.
  2. Load warehouse available inventory (only SML / MED / LAR / XLR sizes).
  3. Load AG validation → find AGs where this store has fewer valid options
     than its Final Options cap (shortfall).
  4. For each shortfall AG, find NEW styles (not already present at the store)
     from warehouse and plan base-stock quantities:
       SML = 2,  MED = 2,  LAR = 4,  XLR = 4
  5. Output Excel (D:\\INCREFF ORDER PUNCH\\OUTPUTFILE):
       Sheet 1 – AG Shortfall Summary
       Sheet 2 – New Style Plan        (Style | Color | Size | Suggest Qty | WH Avail | AG)
       Sheet 3 – Current Store Position
"""

import os, sys, glob, shutil
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────
STOCK_DIR   = r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock"
TRANSIT_DIR = r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit"
ALLOC_DIR   = r"D:\INCREFF ORDER PUNCH\ebo stock track data\ALLOCATION"
AG_MAP_DIR  = r"D:\INCREFF ORDER PUNCH\ebo stock track data\STYLE WISE AG"
AG_VALID    = r"D:\INCREFF ORDER PUNCH\VALID STYLE OUTPUT\AG_Validation_Output_v2_LATEST.xlsx"
WH_CSV      = r"D:\downloads\Inventory Available for Sales - OMS-2026-07-23T15_40_21.531+05_30 (1).csv"
OUTPUT_DIR  = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE"

TARGET_STORE_KEYWORDS = ["mysore road"]
TARGET_STORE_LABEL    = "TSPL MYSORE ROAD EBO"

# Base stock targets – only these 4 sizes will be planned
BASE_STOCK = {"SML": 2, "MED": 2, "LAR": 4, "XLR": 4, "2XL": 2}
PLAN_SIZES = set(BASE_STOCK.keys())

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def get_latest(directory, pattern="*.xlsx"):
    files = glob.glob(os.path.join(directory, pattern))
    files = [f for f in files if not os.path.basename(f).startswith("~")]
    return max(files, key=os.path.getmtime) if files else None

def is_mysore(name):
    n = str(name).strip().lower()
    return any(kw in n for kw in TARGET_STORE_KEYWORDS)

def styled_sheet(writer, df, sheet_name, hdr_hex, zebra_hex, freeze="A2"):
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    fill_h = PatternFill("solid", fgColor=hdr_hex)
    font_h = Font(bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill      = fill_h
        cell.font      = font_h
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_z = PatternFill("solid", fgColor=zebra_hex)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row % 2 == 0:
                cell.fill = fill_z
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 42)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze

# ─────────────────────────────────────────────────────────────
# STEP 1 – AG mapping (Style → AG Name)
# ─────────────────────────────────────────────────────────────
ag_map_file = get_latest(AG_MAP_DIR)
if not ag_map_file:
    sys.exit("ERROR: No AG mapping file found.")
print(f"AG mapping   : {os.path.basename(ag_map_file)}")
df_ag_map = pd.read_excel(ag_map_file, usecols=["Style", "AG Name"])
df_ag_map["Style"]   = df_ag_map["Style"].astype(str).str.strip().str.upper()
df_ag_map["AG Name"] = df_ag_map["AG Name"].astype(str).str.strip()
ag_lookup = dict(zip(df_ag_map["Style"], df_ag_map["AG Name"]))

# ─────────────────────────────────────────────────────────────
# STEP 2 – AG Validation → Mysore Road shortfall
# ─────────────────────────────────────────────────────────────
print(f"AG Validation: {os.path.basename(AG_VALID)}")
df_ag_raw = pd.read_excel(AG_VALID, sheet_name="STORE AG RAW")
df_ag_raw.columns = [str(c).strip() for c in df_ag_raw.columns]

mysore_ag = df_ag_raw[df_ag_raw["Store Name"].apply(is_mysore)].copy()
if mysore_ag.empty:
    sys.exit(f"ERROR: '{TARGET_STORE_LABEL}' not found in STORE AG RAW sheet.")

mysore_ag["Gap"] = (
    pd.to_numeric(mysore_ag["Final Options"],     errors="coerce").fillna(0) -
    pd.to_numeric(mysore_ag["All Valid Options"],  errors="coerce").fillna(0)
).astype(int)

print(f"\n{'AG Name':48s} {'Cap':>4}  {'Valid':>5}  {'Gap':>4}")
print("-" * 67)
for _, r in mysore_ag.sort_values("Gap", ascending=False).iterrows():
    print(f"  {r['AG Name']:46s} {int(r['Final Options']):>4}  {int(r['All Valid Options']):>5}  {int(r['Gap']):>4}")

# Shortfall AGs
ag_shortfall = (
    mysore_ag[mysore_ag["Gap"] > 0]
    [["AG Name", "Final Options", "All Valid Options", "Gap"]]
    .copy()
    .sort_values("Gap", ascending=False)
    .reset_index(drop=True)
)
ag_shortfall.columns = ["AG Name", "Final Options (Cap)", "Current All Valid", "Shortfall (Gap)"]
shortfall_ag_names = set(ag_shortfall["AG Name"].str.strip())

# ─────────────────────────────────────────────────────────────
# STEP 3 – Store current stock
# ─────────────────────────────────────────────────────────────
stock_file = get_latest(STOCK_DIR)
if not stock_file:
    sys.exit("ERROR: No stock file found.")
print(f"\nStock file   : {os.path.basename(stock_file)}")
df_stock = pd.read_excel(stock_file)
df_stock.columns = [c.strip() for c in df_stock.columns]

sc_col  = next((c for c in df_stock.columns if c.lower() in ["owner site", "store name"]), None)
sku_col = next((c for c in df_stock.columns if c.lower() == "sku"), None)
sty_col = next((c for c in df_stock.columns if c.lower() == "style"), None)
col_col = next((c for c in df_stock.columns if c.lower() == "color"), None)
sz_col  = next((c for c in df_stock.columns if c.lower() == "size"), None)
qty_col = next((c for c in df_stock.columns if c.lower() in ["stock quantity", "quantity", "qty"]), None)

df_mysore_stk = df_stock[df_stock[sc_col].apply(is_mysore)].copy()
df_mysore_stk["_style"] = df_mysore_stk[sty_col].astype(str).str.strip().str.upper()
df_mysore_stk["_color"] = df_mysore_stk[col_col].astype(str).str.strip().str.upper()
df_mysore_stk["_size"]  = df_mysore_stk[sz_col].astype(str).str.strip().str.upper()
df_mysore_stk["_sku"]   = df_mysore_stk[sku_col].astype(str).str.strip().str.upper()
df_mysore_stk[qty_col]  = pd.to_numeric(df_mysore_stk[qty_col], errors="coerce").fillna(0)
print(f"  Stock rows : {len(df_mysore_stk)}")

# ─────────────────────────────────────────────────────────────
# STEP 4 – Store current transit
# ─────────────────────────────────────────────────────────────
transit_file = get_latest(TRANSIT_DIR)
df_mysore_tran = pd.DataFrame()
if transit_file:
    print(f"Transit file : {os.path.basename(transit_file)}")
    df_tran = pd.read_excel(transit_file)
    df_tran.columns = [c.strip() for c in df_tran.columns]
    tn_col = next((c for c in df_tran.columns if c.lower() in ["ebo name","store name","owner site"]), None)
    ts_col = next((c for c in df_tran.columns if c.lower() == "sku"), None)
    tsty   = next((c for c in df_tran.columns if c.lower() == "style"), None)
    tcol   = next((c for c in df_tran.columns if c.lower() == "color"), None)
    tsz    = next((c for c in df_tran.columns if c.lower() == "size"), None)
    tqty   = next((c for c in df_tran.columns if c.lower() in ["transit qty","quantity","qty"]), None)
    df_mysore_tran = df_tran[df_tran[tn_col].apply(is_mysore)].copy()
    df_mysore_tran["_style"] = df_mysore_tran[tsty].astype(str).str.strip().str.upper()
    df_mysore_tran["_color"] = df_mysore_tran[tcol].astype(str).str.strip().str.upper()
    df_mysore_tran["_size"]  = df_mysore_tran[tsz].astype(str).str.strip().str.upper()
    df_mysore_tran["_sku"]   = df_mysore_tran[ts_col].astype(str).str.strip().str.upper()
    df_mysore_tran[tqty]     = pd.to_numeric(df_mysore_tran[tqty], errors="coerce").fillna(0)
    print(f"  Transit rows: {len(df_mysore_tran)}")

# ─────────────────────────────────────────────────────────────
# STEP 5 – Store current allocation
# ─────────────────────────────────────────────────────────────
alloc_file = get_latest(ALLOC_DIR)
df_mysore_alloc = pd.DataFrame()
if alloc_file:
    print(f"Alloc file   : {os.path.basename(alloc_file)}")
    df_alloc = pd.read_excel(alloc_file)
    df_alloc.columns = [c.strip() for c in df_alloc.columns]
    a_name = next((c for c in df_alloc.columns if c.lower() == "store name"), None)
    a_sku  = next((c for c in df_alloc.columns if c.lower() in ["client sku id / ean","sku","barcode"]), None)
    a_sty  = next((c for c in df_alloc.columns if c.lower() == "style"), None)
    a_col  = next((c for c in df_alloc.columns if c.lower() == "color"), None)
    a_sz   = next((c for c in df_alloc.columns if c.lower() == "size"), None)
    a_qty  = next((c for c in df_alloc.columns if c.lower() in ["max allocated qty","allocated qty","qty"]), None)
    df_mysore_alloc = df_alloc[df_alloc[a_name].apply(is_mysore)].copy()
    df_mysore_alloc["_style"] = df_mysore_alloc[a_sty].astype(str).str.strip().str.upper()
    df_mysore_alloc["_color"] = df_mysore_alloc[a_col].astype(str).str.strip().str.upper()
    df_mysore_alloc["_size"]  = df_mysore_alloc[a_sz].astype(str).str.strip().str.upper()
    df_mysore_alloc["_sku"]   = df_mysore_alloc[a_sku].astype(str).str.strip().str.upper()
    df_mysore_alloc[a_qty]    = pd.to_numeric(df_mysore_alloc[a_qty], errors="coerce").fillna(0)
    print(f"  Alloc rows  : {len(df_mysore_alloc)}")

# ─────────────────────────────────────────────────────────────
# STEP 6 – Existing STYLES at this store (stock + transit + alloc)
#   If a style appears in ANY color/size at the store, the entire style
#   is considered present – do NOT plan any new colors for it.
# ─────────────────────────────────────────────────────────────
existing_styles = set()
for df_ in [df_mysore_stk, df_mysore_tran, df_mysore_alloc]:
    if not df_.empty:
        for style in df_["_style"]:
            existing_styles.add(str(style).strip().upper())

print(f"\nExisting STYLES at {TARGET_STORE_LABEL}: {len(existing_styles)}")

# ─────────────────────────────────────────────────────────────
# STEP 7 – Load warehouse inventory
#   NOTE: CSV columns are:
#     Each Qty        → SKU
#     SKU Description → (ignore, it's "1")
#     Style           → Item Description text
#     Color           → Style code (e.g. OR10)
#     Size            → Color code (e.g. BLK)
#     MRP             → Actual Size (SML/MED/LAR/XLR …)
#     Total Available Quantity → Qty
# ─────────────────────────────────────────────────────────────
print(f"\nWarehouse    : {os.path.basename(WH_CSV)}")
df_wh = pd.read_csv(WH_CSV)
df_wh.columns = [c.strip() for c in df_wh.columns]

# Rename to meaningful names based on actual content
df_wh = df_wh.rename(columns={
    "Each Qty"               : "SKU",
    "SKU Description"        : "_drop1",
    "Style"                  : "Description",
    "Color"                  : "Style",
    "Size"                   : "Color",
    "MRP"                    : "Size",
    "Total Available Quantity": "WH Qty",
})

df_wh["Style"]  = df_wh["Style"].astype(str).str.strip().str.upper()
df_wh["Color"]  = df_wh["Color"].astype(str).str.strip().str.upper()
df_wh["Size"]   = df_wh["Size"].astype(str).str.strip().str.upper()
df_wh["WH Qty"] = pd.to_numeric(df_wh["WH Qty"], errors="coerce").fillna(0)

# Keep only our 4 target sizes with positive stock
df_wh_plan = df_wh[
    (df_wh["Size"].isin(PLAN_SIZES)) &
    (df_wh["WH Qty"] > 0)
].copy()

print(f"  WH rows (SML/MED/LAR/XLR with stock): {len(df_wh_plan)}")

# Keep SKU-level rows (each row = one SKU, one size)
# Normalise SKU to uppercase for matching
df_wh_plan["SKU"] = df_wh_plan["SKU"].astype(str).str.strip().str.upper()
wh_agg = df_wh_plan[["SKU", "Style", "Color", "Size", "WH Qty"]].copy()
wh_agg["AG Name"] = wh_agg["Style"].map(ag_lookup).fillna("UNKNOWN AG")

# ─────────────────────────────────────────────────────────────
# STEP 8 – Find NEW Styles (not present at store at all) from shortfall AGs
# ─────────────────────────────────────────────────────────────
# Exclude entire style if it appears in stock / transit / allocation
wh_agg["Is_New"] = ~wh_agg["Style"].isin(existing_styles)

# Candidate: completely new style AND belongs to a shortfall AG
wh_candidates = wh_agg[
    wh_agg["Is_New"] &
    wh_agg["AG Name"].isin(shortfall_ag_names)
].copy()

print(f"  New Style rows from shortfall AGs (SML/MED/LAR/XLR): {len(wh_candidates)}")

# Gap lookup for sorting
gap_lookup = dict(zip(
    ag_shortfall["AG Name"].str.strip(),
    ag_shortfall["Shortfall (Gap)"].astype(int)
))

# ─────────────────────────────────────────────────────────────
# STEP 9 – Build size-level plan with base stock quantities
# ─────────────────────────────────────────────────────────────
plan_rows = []
for _, r in wh_candidates.iterrows():
    base    = BASE_STOCK[r["Size"]]           # 2 or 4
    suggest = min(base, int(r["WH Qty"]))     # can't exceed WH stock
    if suggest == 0:
        continue
    plan_rows.append({
        "SKU"             : r["SKU"],
        "Style"           : r["Style"],
        "Color"           : r["Color"],
        "Size"            : r["Size"],
        "AG Name"         : r["AG Name"],
        "AG Shortfall"    : gap_lookup.get(r["AG Name"], 0),
        "WH Available"    : int(r["WH Qty"]),
        "Base Stock Target": base,
        "Suggest Qty"     : suggest,
    })

df_plan = pd.DataFrame(plan_rows)

if not df_plan.empty:
    df_plan = df_plan.sort_values(
        ["AG Shortfall", "Style", "Color", "Size"],
        ascending=[False, True, True, True]
    ).reset_index(drop=True)
    # Re-order columns: SKU first
    df_plan = df_plan[["SKU", "Style", "Color", "Size", "AG Name", "AG Shortfall", "WH Available", "Base Stock Target", "Suggest Qty"]]

print(f"  Plan rows (size level)               : {len(df_plan)}")

# Summary by Style-Color for quick view
if not df_plan.empty:
    sc_summary = df_plan.groupby(["AG Name", "Style", "Color", "AG Shortfall"]).agg(
        Sizes_Available=("Size", lambda x: " | ".join(sorted(x))),
        Total_Suggest_Qty=("Suggest Qty", "sum"),
        WH_Total=("WH Available", "sum")
    ).reset_index().sort_values(["AG Shortfall", "Style"], ascending=[False, True])
    print(f"\nNew style-color options planned: {len(sc_summary)}")
    for _, r in sc_summary.head(20).iterrows():
        print(f"  [{r['AG Name'][:30]:30s}] gap={r['AG Shortfall']:3d} | {r['Style']} {r['Color']:6s} | sizes={r['Sizes_Available']} | suggestQty={r['Total_Suggest_Qty']}")
else:
    sc_summary = pd.DataFrame()
    print("\nWARNING: No new style-color options found in shortfall AGs with SML/MED/LAR/XLR stock.")

# ─────────────────────────────────────────────────────────────
# STEP 10 – Current store position sheet
# ─────────────────────────────────────────────────────────────
pos_rows = []
for _, r in df_mysore_stk.iterrows():
    pos_rows.append({"Source":"Stock",  "Style":r["_style"], "Color":r["_color"], "Size":r["_size"],
                     "SKU":r["_sku"],   "Qty":int(r[qty_col]),
                     "AG Name": ag_lookup.get(r["_style"], "UNKNOWN AG")})
for _, r in df_mysore_tran.iterrows():
    pos_rows.append({"Source":"Transit","Style":r["_style"], "Color":r["_color"], "Size":r["_size"],
                     "SKU":r["_sku"],   "Qty":int(r[tqty]),
                     "AG Name": ag_lookup.get(r["_style"], "UNKNOWN AG")})
for _, r in df_mysore_alloc.iterrows():
    pos_rows.append({"Source":"Alloc",  "Style":r["_style"], "Color":r["_color"], "Size":r["_size"],
                     "SKU":r["_sku"],   "Qty":int(r[a_qty]),
                     "AG Name": ag_lookup.get(r["_style"], "UNKNOWN AG")})

df_pos = pd.DataFrame(pos_rows).sort_values(
    ["AG Name", "Style", "Color", "Size", "Source"]
).reset_index(drop=True) if pos_rows else pd.DataFrame()

# ─────────────────────────────────────────────────────────────
# STEP 11 – Write Excel
# ─────────────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
out = os.path.join(OUTPUT_DIR, f"MysoreRoad_NewStyles_{ts}.xlsx")

print(f"\nWriting -> {out}")
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    styled_sheet(writer, ag_shortfall,          "AG Shortfall Summary",   "1F4E79", "DDEEFF")
    if not df_plan.empty:
        styled_sheet(writer, df_plan,           "New Style Plan",         "375623", "E8F5E9")
    if not sc_summary.empty:
        styled_sheet(writer, sc_summary,        "New Style Summary",      "4A235A", "F5E6FF")
    if not df_pos.empty:
        styled_sheet(writer, df_pos,            "Current Store Position", "843C0C", "FFF3E0")

latest = os.path.join(OUTPUT_DIR, "MysoreRoad_NewStyles_LATEST.xlsx")
try:
    shutil.copy(out, latest)
except PermissionError:
    print(f"  Note: Could not overwrite LATEST copy (file may be open in Excel).")

print(f"\n{'='*62}")
print(f"  DONE - {TARGET_STORE_LABEL}")
print(f"  AG Shortfall AGs     : {len(ag_shortfall)}")
print(f"  New Style Plan rows  : {len(df_plan)}")
print(f"  Store Position rows  : {len(df_pos)}")
print(f"  Output  -> {out}")
print(f"  Latest  -> {latest}")
print(f"{'='*62}")
