"""
plus_size_report.py
-------------------
Generates a Plus-Size (3XL / 4XL / 5XL) inventory report.

Output Excel has two sheets:
  1. Summary   – Style | SKU | WH Available Qty | per-store stock cols | per-store transit cols
  2. Raw        – Raw plus-size rows from the OMS inventory CSV
"""

import os
import sys
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
INVENTORY_CSV = r"D:\downloads\Inventory Available for Sales - OMS-2026-07-07T10_29_29.434+05_30.csv"
STOCK_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock"
TRANSIT_DIR   = r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit"
ALLOC_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\ALLOCATION"
OUTPUT_DIR    = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE"

PLUS_SIZES = {"3XL", "4XL", "5XL", "3 XL", "4 XL", "5 XL", "XXXL", "XXXXL", "XXXXXL"}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def get_latest(directory, pattern="*.xlsx"):
    files = glob.glob(os.path.join(directory, pattern))
    files = [f for f in files if not os.path.basename(f).startswith("~")]
    return max(files, key=os.path.getmtime) if files else None

def clean_size(s):
    return str(s).strip().upper().replace(" ", "")

def clean_store_code(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        return s[:-2]
    return s

def is_plus_size(s):
    return clean_size(s) in {clean_size(p) for p in PLUS_SIZES}

def apply_header_style(ws, row_num, fill_hex="1F3864"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align

def apply_thin_border(ws):
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

# ─────────────────────────────────────────────
# STEP 1 — Load OMS inventory, filter plus sizes
# ─────────────────────────────────────────────
print("Loading OMS inventory...")
df_inv = pd.read_csv(INVENTORY_CSV)

# Standardise column names
df_inv.columns = [c.strip() for c in df_inv.columns]

df_inv["_size_clean"] = df_inv["Size"].apply(clean_size)
plus_mask = df_inv["_size_clean"].isin({clean_size(p) for p in PLUS_SIZES})
df_plus = df_inv[plus_mask].copy()

if df_plus.empty:
    print("No plus-size SKUs found in inventory CSV.")
    sys.exit(0)

print(f"Found {len(df_plus)} plus-size inventory rows.")

# Aggregate WH available qty per SKU
sku_col   = "Client SKU Id / EAN"
style_col = "Style"
size_col  = "Size"
avail_col = "Total Available Quantity"

df_sku = (
    df_plus.groupby([style_col, sku_col, size_col], as_index=False)
    .agg(WH_Available_Qty=(avail_col, "sum"))
    .rename(columns={style_col: "Style", sku_col: "SKU", size_col: "Size"})
)
df_sku["SKU"] = df_sku["SKU"].astype(str).str.strip()
df_sku["Style"] = df_sku["Style"].astype(str).str.strip()
all_skus = set(df_sku["SKU"].str.upper())

print(f"Unique plus-size SKUs: {len(all_skus)}")

# ─────────────────────────────────────────────
# STEP 2 — Load current store stock and build mapping
# ─────────────────────────────────────────────
stock_file = get_latest(STOCK_DIR)
if not stock_file:
    print("ERROR: No current stock file found.")
    sys.exit(1)
print(f"Loading stock: {os.path.basename(stock_file)}")

df_stock = pd.read_excel(stock_file)
df_stock.columns = [c.strip() for c in df_stock.columns]

# Detect columns
sc_col   = next((c for c in df_stock.columns if c.strip().lower() in ["store code", "store_code", "site_code", "owner site code"]), None)
sn_col   = next((c for c in df_stock.columns if c.strip().lower() in ["owner site", "store name", "ebo name"]), None)
sku_s    = next((c for c in df_stock.columns if c.strip().lower() == "sku"), None) or \
           next((c for c in df_stock.columns if c.strip().lower() in ["barcode", "client sku id / ean"]), None)
qty_s    = next((c for c in df_stock.columns if c.strip().lower() in ["stock quantity", "quantity", "qty"]), None)

# Build a store_map dynamically
store_map = {}
if sc_col and sn_col:
    for _, row in df_stock.iterrows():
        code = clean_store_code(row[sc_col])
        name = str(row[sn_col]).strip()
        if code and name and name.lower() != 'nan':
            store_map[code] = name

# Load transit to help with store mapping
transit_file = get_latest(TRANSIT_DIR)
df_tran = None
if transit_file:
    print(f"Reading transit for mapping: {os.path.basename(transit_file)}")
    df_tran = pd.read_excel(transit_file)
    df_tran.columns = [c.strip() for c in df_tran.columns]
    
    sku_t  = next((c for c in df_tran.columns if c.strip().lower() == "sku"), None) or \
               next((c for c in df_tran.columns if c.strip().lower() in ["barcode", "client sku id / ean"]), None)
    sn_t   = next((c for c in df_tran.columns if c.strip().lower() in ["ebo name", "store name", "owner site"]), None)
    sc_t   = next((c for c in df_tran.columns if c.strip().lower() in ["store code", "store_code"]), None)
    qty_t  = next((c for c in df_tran.columns if c.strip().lower() in ["transit qty", "quantity", "qty"]), None)
    
    if sc_t and sn_t:
        for _, row in df_tran.iterrows():
            code = clean_store_code(row[sc_t])
            name = str(row[sn_t]).strip()
            if code and name and name.lower() != 'nan' and code not in store_map:
                store_map[code] = name

# Load allocation to help with store mapping
alloc_csv = get_latest(ALLOC_DIR, '*.csv')
alloc_xlsx = get_latest(ALLOC_DIR, '*.xlsx')
if alloc_csv and alloc_xlsx:
    alloc_file = alloc_csv if os.path.getmtime(alloc_csv) > os.path.getmtime(alloc_xlsx) else alloc_xlsx
else:
    alloc_file = alloc_csv or alloc_xlsx

df_alloc = None
if alloc_file:
    print(f"Reading allocation for mapping: {os.path.basename(alloc_file)}")
    if alloc_file.endswith('.csv'):
        df_alloc = pd.read_csv(alloc_file)
    else:
        df_alloc = pd.read_excel(alloc_file)
    df_alloc.columns = [c.strip() for c in df_alloc.columns]
    
    sku_a = next((c for c in df_alloc.columns if c.strip().lower() in ["client sku id / ean", "client sku id/ean", "barcode", "sku"]), None)
    sn_a  = next((c for c in df_alloc.columns if c.strip().lower() in ["store name", "store_name", "ebo name"]), None)
    sc_a  = next((c for c in df_alloc.columns if c.strip().lower() in ["store code", "store_code", "site code", "site_code"]), None)
    qty_a = next((c for c in df_alloc.columns if c.strip().lower() in ["max allocated qty", "allocated qty", "quantity", "qty"]), None)
    
    if sc_a and sn_a:
        for _, row in df_alloc.iterrows():
            code = clean_store_code(row[sc_a])
            name = str(row[sn_a]).strip()
            if code and name and name.lower() != 'nan' and code not in store_map:
                store_map[code] = name

# Process and pivot stock
df_stock["_sku_up"] = df_stock[sku_s].astype(str).str.strip().str.upper()
df_stock_plus = df_stock[df_stock["_sku_up"].isin(all_skus)].copy()
df_stock_plus["_store_code"] = df_stock_plus[sc_col].apply(clean_store_code) if sc_col else ""
df_stock_plus["_store_label"] = df_stock_plus.apply(
    lambda r: store_map.get(r["_store_code"], str(r[sn_col]).strip() if sn_col else r["_store_code"]),
    axis=1
)
df_stock_plus[qty_s] = pd.to_numeric(df_stock_plus[qty_s], errors="coerce").fillna(0)

stock_pivot = (
    df_stock_plus.groupby(["_sku_up", "_store_label"])[qty_s]
    .sum()
    .unstack(fill_value=0)
)
stock_pivot.columns = [f"Stock_{c}" for c in stock_pivot.columns]
print(f"Store stock columns: {len(stock_pivot.columns)}")

# ─────────────────────────────────────────────
# STEP 3 — Load and pivot transit and allocation
# ─────────────────────────────────────────────
if df_tran is None:
    print("WARNING: No transit file found. Skipping transit data.")
    transit_pivot = pd.DataFrame()
else:
    df_tran["_sku_up"] = df_tran[sku_t].astype(str).str.strip().str.upper()
    df_tran_plus = df_tran[df_tran["_sku_up"].isin(all_skus)].copy()
    df_tran_plus["_store_code"] = df_tran_plus[sc_t].apply(clean_store_code) if sc_t else ""
    df_tran_plus["_store_label"] = df_tran_plus.apply(
        lambda r: store_map.get(r["_store_code"], str(r[sn_t]).strip() if sn_t else r["_store_code"]),
        axis=1
    )
    df_tran_plus[qty_t] = pd.to_numeric(df_tran_plus[qty_t], errors="coerce").fillna(0)
    transit_pivot = (
        df_tran_plus.groupby(["_sku_up", "_store_label"])[qty_t]
        .sum()
        .unstack(fill_value=0)
    )
    transit_pivot.columns = [f"Transit_{c}" for c in transit_pivot.columns]
    print(f"Transit columns: {len(transit_pivot.columns)}")

if df_alloc is None:
    print("WARNING: No allocation file found. Skipping allocation data.")
    alloc_pivot = pd.DataFrame()
else:
    df_alloc["_sku_up"] = df_alloc[sku_a].astype(str).str.strip().str.upper()
    df_alloc_plus = df_alloc[df_alloc["_sku_up"].isin(all_skus)].copy()
    df_alloc_plus["_store_code"] = df_alloc_plus[sc_a].apply(clean_store_code) if sc_a else ""
    df_alloc_plus["_store_label"] = df_alloc_plus.apply(
        lambda r: store_map.get(r["_store_code"], str(r[sn_a]).strip() if sn_a else r["_store_code"]),
        axis=1
    )
    df_alloc_plus[qty_a] = pd.to_numeric(df_alloc_plus[qty_a], errors="coerce").fillna(0)
    alloc_pivot = (
        df_alloc_plus.groupby(["_sku_up", "_store_label"])[qty_a]
        .sum()
        .unstack(fill_value=0)
    )
    alloc_pivot.columns = [f"Alloc_{c}" for c in alloc_pivot.columns]
    print(f"Allocation columns: {len(alloc_pivot.columns)}")

# ─────────────────────────────────────────────
# STEP 4 — Merge everything
# ─────────────────────────────────────────────
df_sku["_sku_up"] = df_sku["SKU"].str.upper()

df_merged = df_sku.set_index("_sku_up")

if not stock_pivot.empty:
    df_merged = df_merged.join(stock_pivot, how="left")

if not transit_pivot.empty:
    df_merged = df_merged.join(transit_pivot, how="left")

if not alloc_pivot.empty:
    df_merged = df_merged.join(alloc_pivot, how="left")

df_merged = df_merged.fillna(0)
df_merged = df_merged.reset_index(drop=True)

# Sort: Style → Size
size_order = {"3XL": 1, "4XL": 2, "5XL": 3}
df_merged["_sort"] = df_merged["Size"].apply(lambda x: size_order.get(clean_size(x), 9))
df_merged = df_merged.sort_values(["Style", "_sort"]).drop(columns=["_sort"]).reset_index(drop=True)

# ─────────────────────────────────────────────
# STEP 5 — Write Excel
# ─────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
output_path = os.path.join(OUTPUT_DIR, "Plus_Size_SKU_Report.xlsx")

# Load priority list to sort stores by priority number
PRIORITY_FILE = r"priority list\Priority list.xlsx"
priority_map = {}
if os.path.exists(PRIORITY_FILE):
    try:
        from check_allocation import PRIORITY_TO_EBO
        df_pri = pd.read_excel(PRIORITY_FILE)
        p_store_col = next((c for c in df_pri.columns if str(c).strip().lower() in ["store name", "store_name", "storename"]), "Store Name")
        p_num_col = next((c for c in df_pri.columns if str(c).strip().lower() in ["priority number", "priority", "priority list", "priority_number", "priority_list"]), "priority list")
        
        for _, row in df_pri.iterrows():
            store_name = str(row[p_store_col]).strip()
            try:
                pri_num = int(row[p_num_col])
            except ValueError:
                pri_num = 999999
            priority_map[store_name.lower()] = pri_num
            mapped_ebo = PRIORITY_TO_EBO.get(store_name)
            if mapped_ebo:
                priority_map[str(mapped_ebo).strip().lower()] = pri_num
        print(f"Loaded priority list mapping for {len(priority_map)} name variations.")
    except Exception as e:
        print(f"Warning: Failed to load priority list mapping ({e}). Falling back to alphabetical order.")

def get_store_priority(store_name):
    s_clean = str(store_name).strip().lower()
    if s_clean in priority_map:
        return priority_map[s_clean]
    for k, pri in priority_map.items():
        if k in s_clean:
            return pri
    return 999999

# Gather all stores from stock, transit, and allocation
stock_cols   = [c for c in df_merged.columns if c.startswith("Stock_")]
transit_cols = [c for c in df_merged.columns if c.startswith("Transit_")]
alloc_cols   = [c for c in df_merged.columns if c.startswith("Alloc_")]
all_stores   = set(
    [c.replace("Stock_", "") for c in stock_cols] + 
    [c.replace("Transit_", "") for c in transit_cols] +
    [c.replace("Alloc_", "") for c in alloc_cols]
)

# Sort stores by priority rank, then alphabetically for ties
sorted_stores = sorted(all_stores, key=lambda s: (get_store_priority(s), s.lower()))

base_cols    = ["Style", "SKU", "Size", "WH_Available_Qty"]
final_cols   = base_cols.copy()
for store in sorted_stores:
    s_col = f"Stock_{store}"
    t_col = f"Transit_{store}"
    a_col = f"Alloc_{store}"
    if s_col in df_merged.columns:
        final_cols.append(s_col)
    if t_col in df_merged.columns:
        final_cols.append(t_col)
    if a_col in df_merged.columns:
        final_cols.append(a_col)

df_out = df_merged[final_cols].copy()

# Friendly display names
display_names = {
    c: c.replace("Stock_", "").replace("Transit_", "🚚 ").replace("Alloc_", "📦 ")
    for c in final_cols
}
display_names.update({"Style": "Style", "SKU": "SKU", "Size": "Size", "WH_Available_Qty": "WH Avail Qty"})

def save_report(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Plus Size Report", startrow=1)
        ws = writer.sheets["Plus Size Report"]

        # Group header row (row 1) — colour bands
        base_fill    = PatternFill("solid", fgColor="1F3864")   # dark blue for base
        stock_fill   = PatternFill("solid", fgColor="2E75B6")   # blue for stock
        transit_fill = PatternFill("solid", fgColor="375623")   # green for transit
        alloc_fill   = PatternFill("solid", fgColor="7030A0")   # purple for allocation

        for idx, col_name in enumerate(final_cols, start=1):
            cell = ws.cell(row=1, column=idx)
            if col_name in base_cols:
                cell.fill = base_fill
            elif col_name.startswith("Stock_"):
                cell.fill = stock_fill
            elif col_name.startswith("Transit_"):
                cell.fill = transit_fill
            elif col_name.startswith("Alloc_"):
                cell.fill = alloc_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.value = display_names.get(col_name, col_name)

        # Data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F0F4F8")

        # Column widths
        for idx, col_name in enumerate(final_cols, start=1):
            col_letter = get_column_letter(idx)
            if col_name == "SKU":
                ws.column_dimensions[col_letter].width = 22
            elif col_name == "Style":
                ws.column_dimensions[col_letter].width = 14
            elif col_name in base_cols:
                ws.column_dimensions[col_letter].width = 14
            else:
                ws.column_dimensions[col_letter].width = 16

        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "E2"

        # Raw sheet
        df_raw = df_plus[["Style", sku_col, "Size", avail_col, "Total Allocated Quantity", "Total Quantity"]].copy()
        df_raw.columns = ["Style", "SKU", "Size", "Available Qty", "Allocated Qty", "Total Qty"]
        df_raw.to_excel(writer, index=False, sheet_name="Raw OMS Data", startrow=1)
        ws2 = writer.sheets["Raw OMS Data"]
        apply_header_style(ws2, 1)
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 20

try:
    save_report(output_path)
except PermissionError:
    import time
    output_path = os.path.join(OUTPUT_DIR, f"Plus_Size_SKU_Report_{time.strftime('%H%M%S')}.xlsx")
    print(f"Warning: Plus_Size_SKU_Report.xlsx is locked. Saving as {os.path.basename(output_path)}")
    save_report(output_path)

print(f"\nDone! Output saved to:\n   {output_path}")
print(f"   Plus-size SKUs : {len(df_sku)}")
print(f"   Store columns  : {len(stock_cols)} (stock) + {len(transit_cols)} (transit) + {len(alloc_cols)} (alloc)")
