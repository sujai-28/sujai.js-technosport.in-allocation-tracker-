import os
import sys
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
STOCK_DIR       = r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock"
TRANSIT_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit"
PRIORITY_FILE   = r"D:\INCREFF ORDER PUNCH\priority list\Priority list.xlsx"
HOSUR_FILE      = r"D:\downloads\Inventory Available for Sales - OMS-2026-08-12T15_33_14.590+05_30.csv"
TIRUPUR_FILE    = r"D:\downloads\TUP STOCK DETAILS - 31.07.26 (2).xlsx"
OUTPUT_DIR      = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE"
OUTPUT_PATH     = os.path.join(OUTPUT_DIR, "Footwear_Warehouse_Comparison_Report.xlsx")

TARGET_STYLES = {"S102", "S103"}

# 51 Target Stores Mapping
TARGET_STORES_MAP = {
    "PONDICHERRY": "TSPL PONDICHERRY",
    "BESANT NAGAR": "TSPL BESANT NAGAR EBO",
    "NAMAKKAL": "TSPL NAMAKKAL EBO",
    "VELLORE": "TSPL VELLORE EBO",
    "TEX VALLEY ERODE": "TSPL TEX VALLEY EBO",
    "(Salem 2) Seelanaickenpatti": "TSPL SALEM-2 EBO",
    "Selayur": "TSPL SELAYUR EBO",
    "TIRUPPUR": "TSPL TIRUPPUR",
    "SALEM": "TSPL SALEM",
    "ERODE": "TSPL ERODE STORE",
    "HSR LAYOUT": "TSPL HSR STORE",
    "CHIKKAJALA": "TSPL CHIKKAJALA EBO",
    "DODDABALAPUR": "TSPL DODDABALLA EBO",
    "Mysore Road - Bangalore": "TSPL MYSORE ROAD EBO",
    "ATTIBELLE": "TSPL ATTIBELE EBO",
    "KUVEMPUNAGAR - mysore 2": "TSPL MYSORE 2 EBO",
    "INORBIT MALL - HUBALI": "TSPL HUBBALI EBO",
    "HASSAN": "TSPL HASSAN EBO",
    "UDUPI": "TSPL UDUPI EBO",
    "HUBLI -2": "TSPL SHIRUR_PARK EBO",
    "Belgaum": "TSPL BELGAUM EBO",
    "Davangiri": "TSPL DAVANAGERE EBO",
    "Sarath City Mall": "TSPL SARATH CITY MALL",
    "Lakeshore": "TSPL HYDERABAD",
    "AS Rao": "TSPL AS NAGAR EBO",
    "Vijayawada": "TSPL VIJAYAWADA EBO",
    "Ananthpur": "TSPL ANANTAPUR EBO",
    "Vizianagram": "TSPL VIZIANAGARAM EBO",
    "In-orbit Vizag": "TSPL VIZAG EBO",
    "Khammam": "TSPL KHAMMAM EBO",
    "Pune Wagholi": "TSPL WAGHOLI EBO",
    "Moshi": "TSPL MOUSHI EBO",
    "Katraj": "TSPL KATRAJ EBO",
    "Pune - Kharadi": "TSPL PUNE KH",
    "Pune - Pimple Saudhagar": "TSPL PUNE PIMPLE",
    "Pune Hadapsar": "TSPL HADAPSAR EBO",
    "KOLHAPUR": "TSPL KOLHAPUR EBO",
    "Nashik": "TSPL NASHIK",
    "Capital Mall Vasai-Virar": "TSPL VASAI_1 EBO",
    "Bhumi": "TSPL BHIWANDI EBO",
    "Sambhajinagar": "TSPL DIVINITY MALL",
    "Bilashpur": "TSPL BILASPUR EBO",
    "Bhopal": "TSPL BHOPAL STORE",
    "Jabalpur": "TSPL JABALPUR EBO",
    "Raipur": "TSPL RAIPUR EBO",
    "Mani square mall": "TSPL MANI SQUARE MALL",
    "Sentrum mall": "TSPL SENTRUM MALL",
    "Rourkela": "TSPL ROURKELA EBO",
    "Bhagalpur": "TSPL BHAGALPUR EBO",
    "Kharagpur": "TSPL KHARAGPUR EBO",
    "Katagram": "TSPL KATARGAM EBO"
}

REVERSE_MAP = {v: k for k, v in TARGET_STORES_MAP.items()}

def get_latest_file(directory, ext='*.xlsx'):
    files = glob.glob(os.path.join(directory, ext))
    files = [f for f in files if not os.path.basename(f).startswith('~')]
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def clean_store_code(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        return s[:-2]
    return s

def clean_size(s):
    try:
        val = str(s).strip()
        if val.endswith('.0'):
            val = val[:-2]
        return val.upper()
    except:
        return str(s).strip().upper()

import re as _re
_SKU_PATTERN = _re.compile(r'^(MFW[A-Z0-9]+[A-Z]{2,}\d{2})', _re.IGNORECASE)

def clean_sku(sku):
    """Strip price suffix from SKUs like MFWS101BLK071499 → MFWS101BLK07"""
    s = str(sku).strip().upper()
    m = _SKU_PATTERN.match(s)
    return m.group(1) if m else s

def clean_color(c):
    if pd.isna(c):
        return ""
    val = str(c).strip().upper()
    if val in ("BLACK", "BLK"):
        return "BLK"
    if val in ("GREY", "GRY"):
        return "GRY"
    if val in ("LIGHT GREY", "LIGHT GRAY", "LTG"):
        return "LTG"
    if val in ("NAVY", "NVY"):
        return "NVY"
    if val in ("WHITE", "WHT"):
        return "WHT"
    if val in ("COBALT", "CARBON BLACK", "CBL"):
        return "CBL"
    return val

def get_base_stock(rank, size):
    sz_str = clean_size(size)
    try:
        sz = int(sz_str)
    except ValueError:
        return 0
        
    if 1 <= rank <= 4:
        if sz == 7:
            return 2
        elif sz in (8, 9, 10):
            return 3
        elif sz == 11:
            return 2
    elif rank >= 5:
        if sz == 7:
            return 1
        elif sz in (8, 9, 10):
            return 2
        elif sz == 11:
            return 1
    return 0

def main():
    global TIRUPUR_FILE
    print("=" * 70)
    print("  Footwear Warehouse Allocation Comparison Simulator (Hosur vs Tirupur)")
    print("=" * 70)

    # 1. Load Files
    stock_file = get_latest_file(STOCK_DIR, '*.xlsx')
    transit_file = get_latest_file(TRANSIT_DIR, '*.xlsx')
    
    if not stock_file:
        print("Error: Could not find current store stock file.")
        sys.exit(1)
    if not os.path.exists(HOSUR_FILE):
        print(f"Error: Could not find Hosur stock file at '{HOSUR_FILE}'.")
        sys.exit(1)
    if not os.path.exists(TIRUPUR_FILE):
        print(f"Error: Could not find Tirupur stock file at '{TIRUPUR_FILE}'.")
        sys.exit(1)

    print(f"Store Stock  : {os.path.basename(stock_file)}")
    print(f"Store Transit: {os.path.basename(transit_file) if transit_file else 'None'}")
    print(f"Hosur Warehouse Stock   : {os.path.basename(HOSUR_FILE)}")
    print(f"Tirupur Warehouse Stock : {os.path.basename(TIRUPUR_FILE)}")

    # 2. Build Store Code to Name Map
    store_map = {}
    df_stock_raw = pd.read_excel(stock_file)
    df_stock_raw.columns = [c.strip() for c in df_stock_raw.columns]
    
    s_sc = next((c for c in df_stock_raw.columns if c.lower() in ["store code", "store_code", "site_code", "site code", "owner site code"]), None)
    s_sn = next((c for c in df_stock_raw.columns if c.lower() in ["owner site", "store name", "ebo name"]), None)
    s_sku = next((c for c in df_stock_raw.columns if c.lower() == "sku"), None) or next((c for c in df_stock_raw.columns if c.lower() in ["barcode", "client sku id / ean"]), None)
    s_style = next((c for c in df_stock_raw.columns if c.lower() == "style"), "Style")
    s_color = next((c for c in df_stock_raw.columns if c.lower() in ["color", "colour"]), "Colour")
    s_size = next((c for c in df_stock_raw.columns if c.lower() == "size"), "Size")
    s_qty = next((c for c in df_stock_raw.columns if c.lower() in ["stock quantity", "quantity", "qty"]), "Qty")

    if s_sc and s_sn:
        for _, row in df_stock_raw.iterrows():
            code = clean_store_code(row[s_sc])
            name = str(row[s_sn]).strip()
            if code and name and name.lower() != 'nan':
                store_map[code] = name

    df_tran_raw = pd.DataFrame()
    if transit_file:
        df_tran_raw = pd.read_excel(transit_file)
        df_tran_raw.columns = [c.strip() for c in df_tran_raw.columns]
        t_sc = next((c for c in df_tran_raw.columns if c.lower() in ["store code", "store_code", "site code", "site_code"]), None)
        t_sn = next((c for c in df_tran_raw.columns if c.lower() in ["ebo name", "store name", "owner site"]), None)
        t_sku = next((c for c in df_tran_raw.columns if c.lower() == "sku"), None) or next((c for c in df_tran_raw.columns if c.lower() in ["barcode", "client sku id / ean"]), None)
        t_style = next((c for c in df_tran_raw.columns if c.lower() == "style"), "Style")
        t_color = next((c for c in df_tran_raw.columns if c.lower() in ["color", "colour"]), "Color")
        t_size = next((c for c in df_tran_raw.columns if c.lower() == "size"), "Size")
        t_qty = next((c for c in df_tran_raw.columns if c.lower() in ["transit qty", "quantity", "qty"]), "Transit Qty")

        if t_sc and t_sn:
            for _, row in df_tran_raw.iterrows():
                code = clean_store_code(row[t_sc])
                name = str(row[t_sn]).strip()
                if code and name and name.lower() != 'nan' and code not in store_map:
                    store_map[code] = name

    # Load store priorities
    from check_allocation import PRIORITY_TO_EBO
    priority_map = {}
    if os.path.exists(PRIORITY_FILE):
        try:
            df_pri = pd.read_excel(PRIORITY_FILE)
            df_pri.columns = [c.strip() for c in df_pri.columns]
            p_store_col = next((c for c in df_pri.columns if c.lower() in ["store name", "store_name", "storename"]), "Store Name")
            p_num_col = next((c for c in df_pri.columns if c.lower() in ["priority number", "priority", "priority list", "priority_number", "priority_list"]), "priority list")
            
            for _, row in df_pri.iterrows():
                store_name = str(row[p_store_col]).strip()
                try:
                    pri_num = int(row[p_num_col])
                except ValueError:
                    pri_num = 999999
                priority_map[store_name.lower()] = pri_num
                mapped_ebo = PRIORITY_TO_EBO.get(store_name)
                if mapped_ebo:
                    priority_map[str(mapped_ebo).strip().lower()] = pri_num
        except Exception as e:
            print(f"Warning: Failed to load priority list mapping ({e}).")

    def get_store_priority(store_name):
        return 999999

    # Filter stock and transit for S101, S102, S103
    df_stock_filtered = df_stock_raw[df_stock_raw[s_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
    df_stock_filtered["Style_Clean"] = df_stock_filtered[s_style].astype(str).str.strip().str.upper()
    df_stock_filtered["Color_Clean"] = df_stock_filtered[s_color].apply(clean_color)
    df_stock_filtered["Size_Clean"] = df_stock_filtered[s_size].apply(clean_size)
    df_stock_filtered["SKU_Clean"] = df_stock_filtered[s_sku].astype(str).apply(clean_sku)
    df_stock_filtered["Store_Code_Clean"] = df_stock_filtered[s_sc].apply(clean_store_code)
    df_stock_filtered["Store_Name_Clean"] = df_stock_filtered.apply(
        lambda r: store_map.get(r["Store_Code_Clean"], str(r[s_sn]).strip() if s_sn else r["Store_Code_Clean"]),
        axis=1
    )
    df_stock_filtered["Qty_Clean"] = pd.to_numeric(df_stock_filtered[s_qty], errors="coerce").fillna(0).astype(int)

    df_tran_filtered = pd.DataFrame()
    if not df_tran_raw.empty:
        df_tran_filtered = df_tran_raw[df_tran_raw[t_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
        df_tran_filtered["Style_Clean"] = df_tran_filtered[t_style].astype(str).str.strip().str.upper()
        df_tran_filtered["Color_Clean"] = df_tran_filtered[t_color].apply(clean_color)
        df_tran_filtered["Size_Clean"] = df_tran_filtered[t_size].apply(clean_size)
        df_tran_filtered["SKU_Clean"] = df_tran_filtered[t_sku].astype(str).apply(clean_sku)
        df_tran_filtered["Store_Code_Clean"] = df_tran_filtered[t_sc].apply(clean_store_code)
        df_tran_filtered["Store_Name_Clean"] = df_tran_filtered.apply(
            lambda r: store_map.get(r["Store_Code_Clean"], str(r[t_sn]).strip() if t_sn else r["Store_Code_Clean"]),
            axis=1
        )
        df_tran_filtered["Qty_Clean"] = pd.to_numeric(df_tran_filtered[t_qty], errors="coerce").fillna(0).astype(int)

    # 3. Load Hosur Stock
    if HOSUR_FILE.lower().endswith(".csv"):
        df_hosur = pd.read_csv(HOSUR_FILE)
    else:
        df_hosur = pd.read_excel(HOSUR_FILE)
    df_hosur.columns = [c.strip() for c in df_hosur.columns]
    wh_sku_col = next(c for c in df_hosur.columns if c.strip().lower() in ["client sku id / ean", "client sku id/ean", "barcode", "sku", "each client sku id"])
    wh_qty_col = next(c for c in df_hosur.columns if c.strip().lower() in ["total available quantity", "available qty", "quantity", "qty"])
    wh_style_col = next(c for c in df_hosur.columns if c.strip().lower() == "style")
    wh_color_col = next(c for c in df_hosur.columns if c.strip().lower() == "color")
    wh_size_col = next(c for c in df_hosur.columns if c.strip().lower() == "size")

    df_hosur["SKU_Clean"] = df_hosur[wh_sku_col].astype(str).apply(clean_sku)
    df_hosur["Available_Clean"] = pd.to_numeric(df_hosur[wh_qty_col], errors="coerce").fillna(0).astype(int)
    hosur_initial_stock = df_hosur.groupby("SKU_Clean")["Available_Clean"].sum().to_dict()
    
    sku_info = {}
    for _, r in df_hosur[[wh_style_col, wh_color_col, wh_size_col, "SKU_Clean"]].drop_duplicates().iterrows():
        sku_info[r["SKU_Clean"]] = {
            "Style": str(r[wh_style_col]).strip().upper(),
            "Color": str(r[wh_color_col]).strip().upper(),
            "Size": clean_size(r[wh_size_col])
        }

    # 4. Load Tirupur Stock
    xl_tup = pd.ExcelFile(TIRUPUR_FILE)
    tirupur_initial_stock = {}
    for sheet in xl_tup.sheet_names:
        df_sheet = xl_tup.parse(sheet)
        df_sheet.columns = [c.strip() for c in df_sheet.columns]
        if 'Product/Internal Reference' in df_sheet.columns and 'Inventoried Quantity' in df_sheet.columns:
            summary = df_sheet[['Product/Internal Reference','Inventoried Quantity']].dropna(subset=['Product/Internal Reference','Inventoried Quantity'])
            summary = summary[summary['Product/Internal Reference'].astype(str).str.startswith('MFW')]
            for _, r in summary.iterrows():
                sku_code = clean_sku(r['Product/Internal Reference'])
                qty_val = int(pd.to_numeric(r['Inventoried Quantity'], errors='coerce') or 0)
                tirupur_initial_stock[sku_code] = qty_val
                
                # Fill SKU info map if not present
                if sku_code not in sku_info:
                    # Extract from SKU code (e.g. MFWS102BLK09)
                    # MFWS102 BLK 09
                    style = "S102" if "S102" in sku_code else "S103" if "S103" in sku_code else "S101"
                    color = "BLK" if "BLK" in sku_code else "NVY" if "NVY" in sku_code else "GRY" if "GRY" in sku_code else "WHT" if "WHT" in sku_code else "CBL"
                    size = sku_code[-2:]
                    if size.startswith('0'):
                        size = size[1:]
                    sku_info[sku_code] = {
                        "Style": style,
                        "Color": color,
                        "Size": size
                    }

    # 5. Resolve Target Stores list
    sorted_stores = sorted(list(TARGET_STORES_MAP.values()), key=lambda s: (get_store_priority(s), s.lower()))

    # Build Master list of all SKUs
    skus_stock = df_stock_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
    skus_tran = pd.DataFrame()
    if not df_tran_filtered.empty:
        skus_tran = df_tran_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
    
    wh_recs = []
    for s_clean, info in sku_info.items():
        wh_recs.append({
            "Style_Clean": info["Style"],
            "Color_Clean": info["Color"],
            "Size_Clean": info["Size"],
            "SKU_Clean": s_clean
        })
    df_wh_skus = pd.DataFrame(wh_recs)
    
    df_skus_master = pd.concat([skus_stock, skus_tran, df_wh_skus], ignore_index=True).drop_duplicates()
    def size_sort_key(sz):
        try:
            return int(sz)
        except:
            return 999
    df_skus_master["size_int"] = df_skus_master["Size_Clean"].apply(size_sort_key)
    df_skus_master = df_skus_master.sort_values(by=["Style_Clean", "Color_Clean", "size_int"]).drop(columns=["size_int"]).reset_index(drop=True)

    # Pivot Stock and Transit data
    stock_pivot = (
        df_stock_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
        .sum()
        .unstack(fill_value=0)
    )
    stock_pivot.columns = [f"Stock_{c}" for c in stock_pivot.columns]
    
    transit_pivot = pd.DataFrame()
    if not df_tran_filtered.empty:
        transit_pivot = (
            df_tran_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
            .sum()
            .unstack(fill_value=0)
        )
        transit_pivot.columns = [f"Transit_{c}" for c in transit_pivot.columns]

    df_merged = df_skus_master.set_index(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"])
    if not stock_pivot.empty:
        df_merged = df_merged.join(stock_pivot, how="left")
    if not transit_pivot.empty:
        df_merged = df_merged.join(transit_pivot, how="left")
    df_merged = df_merged.fillna(0).reset_index()

    # 6. RUN SIMULATION: TIRUPUR WAREHOUSE ONLY
    print("Running Tirupur allocation simulation...")
    tirupur_stock = tirupur_initial_stock.copy()
    tirupur_allocations = {s: {sku: 0 for sku in df_skus_master["SKU_Clean"]} for s in sorted_stores}
    tirupur_detailed = []

    for rank, store in enumerate(sorted_stores, start=1):
        for _, r in df_skus_master.iterrows():
            style, color, size, sku = r["Style_Clean"], r["Color_Clean"], r["Size_Clean"], r["SKU_Clean"]
            base_stock = get_base_stock(rank, size)
            if base_stock == 0:
                continue

            s_col = f"Stock_{store}"
            t_col = f"Transit_{store}"
            
            curr_stock = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                s_col
            ].iloc[0]) if s_col in df_merged.columns else 0

            transit_qty = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                t_col
            ].iloc[0]) if t_col in df_merged.columns else 0

            gap = max(0, base_stock - (curr_stock + transit_qty))
            allocated_qty = 0

            if gap > 0:
                available_wh = tirupur_stock.get(sku, 0)
                allocated_qty = min(gap, available_wh)
                tirupur_stock[sku] = available_wh - allocated_qty
                tirupur_allocations[store][sku] = allocated_qty

            tirupur_detailed.append({
                "Store Name": store,
                "Style": style,
                "Color": color,
                "Size": size,
                "SKU": sku,
                "Base Stock": base_stock,
                "Current Stock": curr_stock,
                "Transit Qty": transit_qty,
                "Shortfall Qty": gap,
                "Allocated Qty": allocated_qty
            })
    df_tup_detail = pd.DataFrame(tirupur_detailed)

    df_comparison = pd.DataFrame()
    if not df_tup_detail.empty:
        comp_rows = []
        for rank, store in enumerate(sorted_stores, start=1):
            t_store = df_tup_detail[df_tup_detail["Store Name"] == store]
            t_tot = t_store["Allocated Qty"].sum()
            t_s102 = t_store[t_store["Style"] == "S102"]["Allocated Qty"].sum()
            t_s103 = t_store[t_store["Style"] == "S103"]["Allocated Qty"].sum()
            comp_rows.append({
                "Store Name": store,
                "Total Allocated (TUP)": t_tot,
                "S102 Allocated": t_s102,
                "S103 Allocated": t_s103,
            })
        df_comparison = pd.DataFrame(comp_rows)

    wh_comp = []
    for sku, info in sku_info.items():
        t_init = tirupur_initial_stock.get(sku, 0)
        t_rem = tirupur_stock.get(sku, 0)
        t_alloc = t_init - t_rem
        wh_comp.append({
            "Style": info["Style"],
            "Color": info["Color"],
            "Size": info["Size"],
            "SKU": sku,
            "TUP Initial Stock": t_init,
            "TUP Allocated": t_alloc,
            "TUP Remaining": t_rem
        })
    df_wh_comp = pd.DataFrame(wh_comp).sort_values(by=["Style", "Color", "Size"])

    # 10. Clean store names in output dataframes
    if not df_comparison.empty:
        df_comparison["Store Name"] = df_comparison["Store Name"].map(lambda s: REVERSE_MAP.get(s, s))
    df_tup_detail["Store Name"] = df_tup_detail["Store Name"].map(lambda s: REVERSE_MAP.get(s, s))

    # 11. Write Excel Report with Style
    print(f"Writing comparison report to: {OUTPUT_PATH}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    actual_path = OUTPUT_PATH
    try:
        writer = pd.ExcelWriter(actual_path, engine="openpyxl")
    except PermissionError:
        import time
        actual_path = OUTPUT_PATH.replace(".xlsx", f"_{time.strftime('%H%M%S')}.xlsx")
        print(f"Warning: Output file is locked. Saving as: {actual_path}")
        writer = pd.ExcelWriter(actual_path, engine="openpyxl")

    def style_sheet(ws, df, title, freeze_pane="C3"):
        bronze_fill = PatternFill("solid", fgColor="5B3F11")
        alt_fill = PatternFill("solid", fgColor="FBF8F2")
        title_font = Font(name="Calibri", size=14, bold=True, color="5B3F11")
        hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_side = Side(style="thin", color="E0D0C0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = left_align
        ws.row_dimensions[1].height = 30
        
        ws.row_dimensions[2].height = 25
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = bronze_fill
            cell.font = hdr_font
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx in range(3, len(df) + 3):
            ws.row_dimensions[row_idx].height = 18
            use_alt = (row_idx % 2 == 0)
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                if use_alt:
                    cell.fill = alt_fill
                
                col_name = df.columns[col_idx - 1]
                if col_name in ("Store Name", "SKU", "Style", "Color"):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                
                if isinstance(cell.value, (int, float)):
                    cell.value = int(cell.value)
                
                # Highlight columns of interest
                if "Allocated" in col_name or "Difference" in col_name:
                    if cell.value and cell.value != 0:
                        cell.font = bold_data_font
                        
                # Highlight positive/negative difference
                if col_name == "Difference (TUP - HOS)":
                    if cell.value > 0:
                        cell.fill = PatternFill("solid", fgColor="E8F5E9") # light green
                    elif cell.value < 0:
                        cell.fill = PatternFill("solid", fgColor="FFEBEE") # light red

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.freeze_panes = freeze_pane

    with writer:
        if not df_comparison.empty:
            df_comparison.to_excel(writer, sheet_name="Store Allocation (TUP)", index=False, startrow=1)
        df_wh_comp.to_excel(writer, sheet_name="TUP Stock Summary", index=False, startrow=1)
        df_tup_detail.to_excel(writer, sheet_name="Tirupur Detailed Allocation", index=False, startrow=1)

        if not df_comparison.empty:
            style_sheet(writer.sheets["Store Allocation (TUP)"], df_comparison, "Store-Wise Allocation from Tirupur (S102 & S103)", "B3")
        style_sheet(writer.sheets["TUP Stock Summary"], df_wh_comp, "Tirupur Warehouse Stock - Initial, Allocated & Remaining", "E3")
        style_sheet(writer.sheets["Tirupur Detailed Allocation"], df_tup_detail, "Detailed Store SKU Allocations from Tirupur", "C3")

    print("=" * 70)
    print("SUCCESS! Report saved!")
    print(f"File Path: {actual_path}")
    print(f"Tirupur Total Allocated : {df_comparison['Total Allocated (TUP)'].sum() if not df_comparison.empty else 0} units")
    tup_remaining_total = sum(tirupur_stock.values())
    print(f"Tirupur Remaining Stock : {tup_remaining_total} units")
    print("=" * 70)

if __name__ == "__main__":
    main()
