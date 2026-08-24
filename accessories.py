import pandas as pd
import numpy as np
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- PATHS ---
STORE_STOCK_PATH = max(glob.glob(r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock\*.xlsx"), key=os.path.getctime)
INVENTORY_PATH = r"D:\downloads\Inventory Available for Sales - OMS-2026-08-18T11_52_32.411+05_30 (1).csv"
OUTPUT_PATH = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE\Accessories_Replenishment_NewStyle.xlsx"

# --- CONSTANTS ---
TARGET_SIZES = ['MED', 'LAR', 'XLR', '2XL']
TARGET_STYLES = ['OR53', 'OR63', 'OR64', 'OR83']  # Only styles in the accessories inventory

def get_base_qty(size):
    return 4 if size in ['XLR', 'LAR'] else 3

def clean_store_name(name):
    # E.g., 'TSPL SARATH CITY MALL' -> 'SARATH CITY MALL'
    # 'TSPL KATRAJ EBO' -> 'KATRAJ EBO'
    name = str(name).strip()
    if name.startswith('TSPL '):
        name = name[5:]
    # Do not remove 'EBO' because priority list has it (e.g. 'KATRAJ EBO')
    return name.strip().upper()

def style_sheet(ws, df, title):
    navy_fill = PatternFill("solid", fgColor="1F3864")
    alt_fill = PatternFill("solid", fgColor="F2F4F8")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F3864")
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    bold_data_font = Font(name="Calibri", size=10, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_side = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = left_align
    ws.row_dimensions[1].height = 30
    
    ws.row_dimensions[2].height = 25
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = navy_fill
        cell.font = hdr_font
        cell.alignment = center_align
        cell.border = thin_border
        
    for row_idx in range(3, len(df) + 3):
        ws.row_dimensions[row_idx].height = 18
        use_alt = (row_idx % 2 == 0)
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            if use_alt:
                cell.fill = alt_fill
            
            val = cell.value
            if isinstance(val, (int, float)):
                if pd.isna(val):
                    cell.value = 0
                else:
                    cell.value = int(val)
                cell.alignment = center_align
            else:
                cell.alignment = left_align
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col[1:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    ws.freeze_panes = "A3"

def main():
    print("Loading datasets...")
    print(f"Using inventory: {INVENTORY_PATH}")
    
    df_stock = pd.read_excel(STORE_STOCK_PATH)
    df_inv = pd.read_csv(INVENTORY_PATH)
    
    # Standardize stock df - use CCODE (short code) to match inventory Color column
    df_stock.rename(columns={
        'STYLE': 'Style',
        'CCODE': 'Color',       # Use CCODE so it matches inventory color codes
        'SIZE': 'Size',
        'STOCK QUANTITY': 'StockQty',
        'OWNER SITE': 'Store'
    }, inplace=True)
    df_stock['Size'] = df_stock['Size'].astype(str).str.strip().str.upper()
    df_stock['Style'] = df_stock['Style'].astype(str).str.strip()
    df_stock['Color'] = df_stock['Color'].astype(str).str.strip()
    df_stock = df_stock[df_stock['Style'].isin(TARGET_STYLES)]  # Only accessories styles
    
    # Extract all unique stores and order them
    all_stores = df_stock['Store'].dropna().unique().tolist()
    
    # Define top priority stores
    top_1 = 'TSPL SARATH CITY MALL'
    top_2 = 'TSPL VIZAG EBO'
    
    priority_list = []
    if top_1 in all_stores:
        priority_list.append(top_1)
    if top_2 in all_stores:
        priority_list.append(top_2)
        
    for s in all_stores:
        if s not in priority_list:
            priority_list.append(s)
    
    # Standardize inventory df
    df_inv.columns = [str(c).strip() for c in df_inv.columns]
    # Check possible column names
    col_name = 'Sum of Total Available Quantity' if 'Sum of Total Available Quantity' in df_inv.columns else 'Total Available Quantity'
    df_inv.rename(columns={col_name: 'InvQty'}, inplace=True)
    df_inv['Size'] = df_inv['Size'].astype(str).str.strip().str.upper()
    df_inv['Style'] = df_inv['Style'].astype(str).str.strip()
    # Exclude Grand Total row and filter only accessories styles
    df_inv = df_inv[(df_inv['InvQty'] > 0) & (df_inv['Style'].isin(TARGET_STYLES))]
    
    # Precompute SKU mapping from inventory (primary source)
    df_inv['Color'] = df_inv['Color'].astype(str).str.strip()
    sku_mapping = df_inv[['Style', 'Color', 'Size', 'Client SKU Id / EAN']].drop_duplicates()
    sku_dict = {}
    for _, row in sku_mapping.iterrows():
        sku_dict[(row['Style'], row['Color'], row['Size'])] = row['Client SKU Id / EAN']
    
    # Secondary SKU lookup from store stock's own SKU column (fallback)
    store_sku_dict = {}
    if 'SKU' in df_stock.columns:
        for _, row in df_stock[['Style', 'Color', 'Size', 'SKU']].drop_duplicates().iterrows():
            key = (str(row['Style']).strip(), str(row['Color']).strip(), str(row['Size']).strip())
            if key not in store_sku_dict and pd.notna(row['SKU']) and str(row['SKU']).strip():
                store_sku_dict[key] = str(row['SKU']).strip()
    
    def get_sku(key):
        return sku_dict.get(key) or store_sku_dict.get(key) or f"{key[0]}-{key[1]}-{key[2]}"
        
    # Aggregate inventory
    inv_agg = df_inv.groupby(['Style', 'Color', 'Size'], as_index=False)['InvQty'].sum()
    
    # Create inventory pool for allocation
    inv_pool = {}
    for _, row in inv_agg.iterrows():
        inv_pool[(row['Style'], row['Color'], row['Size'])] = row['InvQty']

    # Identify styles in inventory
    inv_styles = inv_agg['Style'].unique()
    
    replenish_rows = []
    new_style_rows = []
    
    print("Processing stores in priority order...")
    # Step 1: REPLENISHMENT
    for store_name in priority_list:
        store_stock = df_stock[df_stock['Store'] == store_name]
        
        # Aggregate store stock by Style, Color, Size
        store_stock_agg = store_stock.groupby(['Style', 'Color', 'Size'], as_index=False)['StockQty'].sum()
        
        # Existing Style-Colors in this store
        existing_sc = set(zip(store_stock_agg['Style'], store_stock_agg['Color']))
        
        # We only care about replenishing TARGET_SIZES
        for style, color in existing_sc:
            sc_stock = store_stock_agg[(store_stock_agg['Style'] == style) & (store_stock_agg['Color'] == color)]
            
            # Check target sizes
            for size in TARGET_SIZES:
                key = (style, color, size)
                
                curr_qty_series = sc_stock[sc_stock['Size'] == size]['StockQty']
                curr_qty = curr_qty_series.values[0] if len(curr_qty_series) > 0 else 0
                
                base_qty = get_base_qty(size)
                if curr_qty < base_qty:
                    needed = base_qty - curr_qty
                    avail = inv_pool.get(key, 0)
                    allocated = min(needed, avail)
                    
                    if allocated > 0:
                        inv_pool[key] -= allocated
                        
                    replenish_rows.append({
                        'Store Name': store_name,
                        'SKU': get_sku(key),
                        'Style': style,
                        'Color': color,
                        'Size': size,
                        'Current Stock': curr_qty,
                        'Base Qty': base_qty,
                        'Needed Qty': needed,
                        'Allocated Qty': allocated,
                        'Remaining WH Inv': avail
                    })
                            
    # Step 2: NEW STYLE ALLOCATION
    for store_name in priority_list:
        store_stock = df_stock[df_stock['Store'] == store_name]
        existing_sc = set(zip(store_stock['Style'], store_stock['Color']))
        
        # All available style-colors in inventory
        inv_sc = set([(k[0], k[1]) for k in inv_pool.keys()])
        
        # New style-colors for this store
        new_sc = [sc for sc in inv_sc if sc not in existing_sc]
        
        for style, color in new_sc:
            # Check availability of target sizes
            available_target_sizes = [sz for sz in TARGET_SIZES if inv_pool.get((style, color, sz), 0) > 0]
            
            if len(available_target_sizes) == 0:
                continue # No target sizes available
                
            # Allocate
            allocated_something = False
            temp_rows = []
            for size in TARGET_SIZES:
                key = (style, color, size)
                avail = inv_pool.get(key, 0)
                
                if avail > 0:
                    base_qty = get_base_qty(size)
                    allocated = min(base_qty, avail)
                    inv_pool[key] -= allocated
                    allocated_something = True
                    
                    temp_rows.append({
                        'Store Name': store_name,
                        'SKU': sku_dict.get(key, f"{style}-{color}-{size}"),
                        'Style': style,
                        'Color': color,
                        'Size': size,
                        'Base Qty': base_qty,
                        'Allocated Qty': allocated,
                        'Remaining WH Inv': inv_pool[key]
                    })
                    
            if allocated_something:
                new_style_rows.extend(temp_rows)

    df_rep = pd.DataFrame(replenish_rows)
    df_new = pd.DataFrame(new_style_rows)
    
    print(f"Total replenishment allocations: {len(df_rep)}")
    print(f"Total new style allocations: {len(df_new)}")
    
    print(f"Saving reports to {OUTPUT_PATH}...")
    def save_report(path):
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            if not df_rep.empty:
                df_rep.to_excel(writer, sheet_name="Replenishment", index=False, startrow=1)
                style_sheet(writer.sheets["Replenishment"], df_rep, "Accessories Replenishment Allocation")
            else:
                pd.DataFrame([{'Message': 'No replenishment needed'}]).to_excel(writer, sheet_name='Replenishment', index=False)
                
            if not df_new.empty:
                df_new.to_excel(writer, sheet_name="New Styles", index=False, startrow=1)
                style_sheet(writer.sheets["New Styles"], df_new, "Accessories New Style Allocation")
            else:
                pd.DataFrame([{'Message': 'No new styles allocated'}]).to_excel(writer, sheet_name='New Styles', index=False)

    try:
        save_report(OUTPUT_PATH)
    except PermissionError:
        import time
        alt_path = OUTPUT_PATH.replace(".xlsx", f"_{time.strftime('%H%M%S')}.xlsx")
        print(f"Warning: {OUTPUT_PATH} is locked. Saving to {alt_path} instead.")
        save_report(alt_path)
            
    print("Done!")

if __name__ == "__main__":
    main()
