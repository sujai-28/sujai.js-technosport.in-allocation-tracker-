import os
import sys
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
STOCK_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\current stock"
TRANSIT_DIR   = r"D:\INCREFF ORDER PUNCH\ebo stock track data\intransit"
ALLOC_DIR     = r"D:\INCREFF ORDER PUNCH\ebo stock track data\ALLOCATION"
PRIORITY_FILE = r"D:\INCREFF ORDER PUNCH\priority list\Priority list.xlsx"
OUTPUT_DIR    = r"D:\INCREFF ORDER PUNCH\OUTPUTFILE"
OUTPUT_PATH   = os.path.join(OUTPUT_DIR, "Footwear_Shortfall_Report.xlsx")

# Footwear styles target
TARGET_STYLES = {"S101", "S102", "S103"}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def get_latest_file(directory, ext='*.xlsx'):
    files = glob.glob(os.path.join(directory, ext))
    files = [f for f in files if not os.path.basename(f).startswith('~')]
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def clean_store_code(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        return s[:-2]
    return s

def clean_size(s):
    # Normalize sizes to integers if they are float strings
    try:
        val = str(s).strip()
        if val.endswith('.0'):
            val = val[:-2]
        return val.upper()
    except:
        return str(s).strip().upper()

def clean_color(c):
    if pd.isna(c):
        return ""
    val = str(c).strip().upper()
    if val in ("BLACK", "BLK"):
        return "BLK"
    if val in ("GREY", "GRY"):
        return "GRY"
    if val in ("NAVY", "NVY"):
        return "NVY"
    if val in ("WHITE", "WHT"):
        return "WHT"
    if val in ("COBALT", "CBL"):
        return "CBL"
    return val

# Base stock function
def get_base_stock(rank, size):
    sz_str = clean_size(size)
    try:
        sz = int(sz_str)
    except ValueError:
        return 0
        
    if 1 <= rank <= 4:
        if sz == 7:
            return 2
        elif sz in (8, 9, 10):
            return 3
        elif sz == 11:
            return 2
    elif rank >= 5:
        if sz == 7:
            return 1
        elif sz in (8, 9, 10):
            return 2
        elif sz == 11:
            return 1
    return 0

def main():
    print("=" * 65)
    print("  Footwear Shortfall Analysis (S101, S102, S103)")
    print("=" * 65)
    
    # 1. Resolve Latest Files
    stock_file = get_latest_file(STOCK_DIR, '*.xlsx')
    transit_file = get_latest_file(TRANSIT_DIR, '*.xlsx')
    alloc_csv = get_latest_file(ALLOC_DIR, '*.csv')
    alloc_xlsx = get_latest_file(ALLOC_DIR, '*.xlsx')
    if alloc_csv and alloc_xlsx:
        alloc_file = alloc_csv if os.path.getmtime(alloc_csv) > os.path.getmtime(alloc_xlsx) else alloc_xlsx
    else:
        alloc_file = alloc_csv or alloc_xlsx
        
    if not stock_file:
        print("Error: Could not find Stock file.")
        sys.exit(1)
        
    print(f"Stock File  : {os.path.basename(stock_file)}")
    print(f"Transit File: {os.path.basename(transit_file) if transit_file else 'None'}")
    print(f"Alloc File  : {os.path.basename(alloc_file) if alloc_file else 'None'}")
    
    # 2. Build Store Mapping
    store_map = {}
    
    # Peek stock columns
    df_stock_raw = pd.read_excel(stock_file)
    df_stock_raw.columns = [c.strip() for c in df_stock_raw.columns]
    
    s_sc = next((c for c in df_stock_raw.columns if c.lower() in ["store code", "store_code", "site_code", "site code", "owner site code"]), None)
    s_sn = next((c for c in df_stock_raw.columns if c.lower() in ["owner site", "store name", "ebo name"]), None)
    s_sku = next((c for c in df_stock_raw.columns if c.lower() == "sku"), None) or next((c for c in df_stock_raw.columns if c.lower() in ["barcode", "client sku id / ean"]), None)
    s_style = next((c for c in df_stock_raw.columns if c.lower() == "style"), "Style")
    s_color = next((c for c in df_stock_raw.columns if c.lower() in ["color", "colour"]), "Colour")
    s_size = next((c for c in df_stock_raw.columns if c.lower() == "size"), "Size")
    s_qty = next((c for c in df_stock_raw.columns if c.lower() in ["stock quantity", "quantity", "qty"]), "Qty")
    
    if s_sc and s_sn:
        for _, row in df_stock_raw.iterrows():
            code = clean_store_code(row[s_sc])
            name = str(row[s_sn]).strip()
            if code and name and name.lower() != 'nan':
                store_map[code] = name
                
    # Read transit for additional mapping
    df_tran_raw = pd.DataFrame()
    if transit_file:
        df_tran_raw = pd.read_excel(transit_file)
        df_tran_raw.columns = [c.strip() for c in df_tran_raw.columns]
        t_sc = next((c for c in df_tran_raw.columns if c.lower() in ["store code", "store_code", "site code", "site_code"]), None)
        t_sn = next((c for c in df_tran_raw.columns if c.lower() in ["ebo name", "store name", "owner site"]), None)
        t_sku = next((c for c in df_tran_raw.columns if c.lower() == "sku"), None) or next((c for c in df_tran_raw.columns if c.lower() in ["barcode", "client sku id / ean"]), None)
        t_style = next((c for c in df_tran_raw.columns if c.lower() == "style"), "Style")
        t_color = next((c for c in df_tran_raw.columns if c.lower() in ["color", "colour"]), "Color")
        t_size = next((c for c in df_tran_raw.columns if c.lower() == "size"), "Size")
        t_qty = next((c for c in df_tran_raw.columns if c.lower() in ["transit qty", "quantity", "qty"]), "Transit Qty")
        
        if t_sc and t_sn:
            for _, row in df_tran_raw.iterrows():
                code = clean_store_code(row[t_sc])
                name = str(row[t_sn]).strip()
                if code and name and name.lower() != 'nan' and code not in store_map:
                    store_map[code] = name
                    
    # Read allocation for additional mapping
    df_alloc_raw = pd.DataFrame()
    if alloc_file:
        if alloc_file.endswith('.csv'):
            df_alloc_raw = pd.read_csv(alloc_file)
        else:
            df_alloc_raw = pd.read_excel(alloc_file)
        df_alloc_raw.columns = [c.strip() for c in df_alloc_raw.columns]
        a_sc = next((c for c in df_alloc_raw.columns if c.lower() in ["store code", "store_code", "site code", "site_code"]), None)
        a_sn = next((c for c in df_alloc_raw.columns if c.lower() in ["store name", "store_name", "ebo name"]), None)
        a_sku = next((c for c in df_alloc_raw.columns if c.lower() in ["client sku id / ean", "client sku id/ean", "barcode", "sku"]), None)
        a_style = next((c for c in df_alloc_raw.columns if c.lower() == "style"), "Style")
        a_color = next((c for c in df_alloc_raw.columns if c.lower() in ["color", "colour"]), "Color")
        a_size = next((c for c in df_alloc_raw.columns if c.lower() == "size"), "Size")
        a_qty = next((c for c in df_alloc_raw.columns if c.lower() in ["max allocated qty", "allocated qty", "quantity", "qty"]), "Max Allocated Qty")
        
        if a_sc and a_sn:
            for _, row in df_alloc_raw.iterrows():
                code = clean_store_code(row[a_sc])
                name = str(row[a_sn]).strip()
                if code and name and name.lower() != 'nan' and code not in store_map:
                    store_map[code] = name
                    
    # 3. Load Store Priorities
    from check_allocation import PRIORITY_TO_EBO
    priority_map = {}
    if os.path.exists(PRIORITY_FILE):
        try:
            df_pri = pd.read_excel(PRIORITY_FILE)
            df_pri.columns = [c.strip() for c in df_pri.columns]
            p_store_col = next((c for c in df_pri.columns if c.lower() in ["store name", "store_name", "storename"]), "Store Name")
            p_num_col = next((c for c in df_pri.columns if c.lower() in ["priority number", "priority", "priority list", "priority_number", "priority_list"]), "priority list")
            
            for _, row in df_pri.iterrows():
                store_name = str(row[p_store_col]).strip()
                try:
                    pri_num = int(row[p_num_col])
                except ValueError:
                    pri_num = 999999
                priority_map[store_name.lower()] = pri_num
                mapped_ebo = PRIORITY_TO_EBO.get(store_name)
                if mapped_ebo:
                    priority_map[str(mapped_ebo).strip().lower()] = pri_num
            print(f"Loaded store priorities for {len(priority_map)} name variations.")
        except Exception as e:
            print(f"Warning: Failed to load priority list mapping ({e}).")
            
    def get_store_priority(store_name):
        s_clean = str(store_name).strip().lower()
        if s_clean in priority_map:
            return priority_map[s_clean]
        for k, pri in priority_map.items():
            if k in s_clean or s_clean in k:
                return pri
        return 999999

    # 4. Filter and Process Stock
    print("\nFiltering stock for S101, S102, S103...")
    df_stock_filtered = df_stock_raw[df_stock_raw[s_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
    print(f"Filtered stock rows: {len(df_stock_filtered)}")
    
    df_stock_filtered["Style_Clean"] = df_stock_filtered[s_style].astype(str).str.strip().str.upper()
    df_stock_filtered["Color_Clean"] = df_stock_filtered[s_color].apply(clean_color)
    df_stock_filtered["Size_Clean"] = df_stock_filtered[s_size].apply(clean_size)
    df_stock_filtered["SKU_Clean"] = df_stock_filtered[s_sku].astype(str).str.strip().str.upper()
    df_stock_filtered["Store_Code_Clean"] = df_stock_filtered[s_sc].apply(clean_store_code)
    df_stock_filtered["Store_Name_Clean"] = df_stock_filtered.apply(
        lambda r: store_map.get(r["Store_Code_Clean"], str(r[s_sn]).strip() if s_sn else r["Store_Code_Clean"]),
        axis=1
    )
    df_stock_filtered["Qty_Clean"] = pd.to_numeric(df_stock_filtered[s_qty], errors="coerce").fillna(0).astype(int)

    # 5. Filter and Process Transit
    df_tran_filtered = pd.DataFrame()
    if not df_tran_raw.empty:
        print("Filtering transit for S101, S102, S103...")
        df_tran_filtered = df_tran_raw[df_tran_raw[t_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
        print(f"Filtered transit rows: {len(df_tran_filtered)}")
        
        df_tran_filtered["Style_Clean"] = df_tran_filtered[t_style].astype(str).str.strip().str.upper()
        df_tran_filtered["Color_Clean"] = df_tran_filtered[t_color].apply(clean_color)
        df_tran_filtered["Size_Clean"] = df_tran_filtered[t_size].apply(clean_size)
        df_tran_filtered["SKU_Clean"] = df_tran_filtered[t_sku].astype(str).str.strip().str.upper()
        df_tran_filtered["Store_Code_Clean"] = df_tran_filtered[t_sc].apply(clean_store_code)
        df_tran_filtered["Store_Name_Clean"] = df_tran_filtered.apply(
            lambda r: store_map.get(r["Store_Code_Clean"], str(r[t_sn]).strip() if t_sn else r["Store_Code_Clean"]),
            axis=1
        )
        df_tran_filtered["Qty_Clean"] = pd.to_numeric(df_tran_filtered[t_qty], errors="coerce").fillna(0).astype(int)

    # 6. Filter and Process Allocation
    df_alloc_filtered = pd.DataFrame()
    if not df_alloc_raw.empty:
        print("Filtering allocation for S101, S102, S103...")
        df_alloc_filtered = df_alloc_raw[df_alloc_raw[a_style].astype(str).str.strip().str.upper().isin(TARGET_STYLES)].copy()
        print(f"Filtered allocation rows: {len(df_alloc_filtered)}")
        
        df_alloc_filtered["Style_Clean"] = df_alloc_filtered[a_style].astype(str).str.strip().str.upper()
        df_alloc_filtered["Color_Clean"] = df_alloc_filtered[a_color].apply(clean_color)
        df_alloc_filtered["Size_Clean"] = df_alloc_filtered[a_size].apply(clean_size)
        df_alloc_filtered["SKU_Clean"] = df_alloc_filtered[a_sku].astype(str).str.strip().str.upper()
        df_alloc_filtered["Store_Code_Clean"] = df_alloc_filtered[a_sc].apply(clean_store_code)
        df_alloc_filtered["Store_Name_Clean"] = df_alloc_filtered.apply(
            lambda r: store_map.get(r["Store_Code_Clean"], str(r[a_sn]).strip() if a_sn else r["Store_Code_Clean"]),
            axis=1
        )
        df_alloc_filtered["Qty_Clean"] = pd.to_numeric(df_alloc_filtered[a_qty], errors="coerce").fillna(0).astype(int)

    # The 51 target stores requested by the user
    TARGET_STORES_MAP = {
        "PONDICHERRY": "TSPL PONDICHERRY",
        "BESANT NAGAR": "TSPL BESANT NAGAR EBO",
        "NAMAKKAL": "TSPL NAMAKKAL EBO",
        "VELLORE": "TSPL VELLORE EBO",
        "TEX VALLEY ERODE": "TSPL TEX VALLEY EBO",
        "(Salem 2) Seelanaickenpatti": "TSPL SALEM-2 EBO",
        "Selayur": "TSPL SELAYUR EBO",
        "TIRUPPUR": "TSPL TIRUPPUR",
        "SALEM": "TSPL SALEM",
        "ERODE": "TSPL ERODE STORE",
        "HSR LAYOUT": "TSPL HSR STORE",
        "CHIKKAJALA": "TSPL CHIKKAJALA EBO",
        "DODDABALAPUR": "TSPL DODDABALLA EBO",
        "Mysore Road - Bangalore": "TSPL MYSORE ROAD EBO",
        "ATTIBELLE": "TSPL ATTIBELE EBO",
        "KUVEMPUNAGAR - mysore 2": "TSPL MYSORE 2 EBO",
        "INORBIT MALL - HUBALI": "TSPL HUBBALI EBO",
        "HASSAN": "TSPL HASSAN EBO",
        "UDUPI": "TSPL UDUPI EBO",
        "HUBLI -2": "TSPL SHIRUR_PARK EBO",
        "Belgaum": "TSPL BELGAUM EBO",
        "Davangiri": "TSPL DAVANAGERE EBO",
        "Sarath City Mall": "TSPL SARATH CITY MALL",
        "Lakeshore": "TSPL HYDERABAD",
        "AS Rao": "TSPL AS NAGAR EBO",
        "Vijayawada": "TSPL VIJAYAWADA EBO",
        "Ananthpur": "TSPL ANANTAPUR EBO",
        "Vizianagram": "TSPL VIZIANAGARAM EBO",
        "In-orbit Vizag": "TSPL VIZAG EBO",
        "Khammam": "TSPL KHAMMAM EBO",
        "Pune Wagholi": "TSPL WAGHOLI EBO",
        "Moshi": "TSPL MOUSHI EBO",
        "Katraj": "TSPL KATRAJ EBO",
        "Pune - Kharadi": "TSPL PUNE KH",
        "Pune - Pimple Saudhagar": "TSPL PUNE PIMPLE",
        "Pune Hadapsar": "TSPL HADAPSAR EBO",
        "KOLHAPUR": "TSPL KOLHAPUR EBO",
        "Nashik": "TSPL NASHIK",
        "Capital Mall Vasai-Virar": "TSPL VASAI_1 EBO",
        "Bhumi": "TSPL BHIWANDI EBO",
        "Sambhajinagar": "TSPL DIVINITY MALL",
        "Bilashpur": "TSPL BILASPUR EBO",
        "Bhopal": "TSPL BHOPAL STORE",
        "Jabalpur": "TSPL JABALPUR EBO",
        "Raipur": "TSPL RAIPUR EBO",
        "Mani square mall": "TSPL MANI SQUARE MALL",
        "Sentrum mall": "TSPL SENTRUM MALL",
        "Rourkela": "TSPL ROURKELA EBO",
        "Bhagalpur": "TSPL BHAGALPUR EBO",
        "Kharagpur": "TSPL KHARAGPUR EBO",
        "Katagram": "TSPL KATARGAM EBO"
    }

    global REVERSE_MAP
    REVERSE_MAP = {v: k for k, v in TARGET_STORES_MAP.items()}

    # Sort stores by priority rank, then alphabetically
    sorted_stores = sorted(list(TARGET_STORES_MAP.values()), key=lambda s: (get_store_priority(s), s.lower()))
    
    # Print sorted stores ranking
    print("\n--- Resolved Store Priorities Rank ---")
    store_ranks = {}
    for idx, s in enumerate(sorted_stores, start=1):
        store_ranks[s] = idx
        print(f"  Rank {idx:>2}: {s} (Priority Rank: {get_store_priority(s)})")
            
    # 8. Create Master List of Style, Color, Size, SKU
    skus_stock = df_stock_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
    skus_tran = pd.DataFrame()
    if not df_tran_filtered.empty:
        skus_tran = df_tran_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
    skus_alloc = pd.DataFrame()
    if not df_alloc_filtered.empty:
        skus_alloc = df_alloc_filtered[["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]].drop_duplicates()
        
    df_skus_master = pd.concat([skus_stock, skus_tran, skus_alloc], ignore_index=True).drop_duplicates()
    # Sort Master list by Style, Color, Size (numeric sort)
    def size_sort_key(sz):
        try:
            return int(sz)
        except:
            return 999
    df_skus_master["size_int"] = df_skus_master["Size_Clean"].apply(size_sort_key)
    df_skus_master = df_skus_master.sort_values(by=["Style_Clean", "Color_Clean", "size_int"]).drop(columns=["size_int"]).reset_index(drop=True)
    
    print(f"\nUnique Footwear SKUs (S101-S103): {len(df_skus_master)}")
    
    # 9. Pivot Data per Store
    # Stock Pivot
    stock_pivot = (
        df_stock_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
        .sum()
        .unstack(fill_value=0)
    )
    stock_pivot.columns = [f"Stock_{c}" for c in stock_pivot.columns]
    
    # Transit Pivot
    transit_pivot = pd.DataFrame()
    if not df_tran_filtered.empty:
        transit_pivot = (
            df_tran_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
            .sum()
            .unstack(fill_value=0)
        )
        transit_pivot.columns = [f"Transit_{c}" for c in transit_pivot.columns]
        
    # Allocation Pivot
    alloc_pivot = pd.DataFrame()
    if not df_alloc_filtered.empty:
        alloc_pivot = (
            df_alloc_filtered.groupby(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean", "Store_Name_Clean"])["Qty_Clean"]
            .sum()
            .unstack(fill_value=0)
        )
        alloc_pivot.columns = [f"Alloc_{c}" for c in alloc_pivot.columns]

    # Merge into Master pivoted sheet
    df_merged = df_skus_master.set_index(["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"])
    if not stock_pivot.empty:
        df_merged = df_merged.join(stock_pivot, how="left")
    if not transit_pivot.empty:
        df_merged = df_merged.join(transit_pivot, how="left")
    if not alloc_pivot.empty:
        df_merged = df_merged.join(alloc_pivot, how="left")
        
    df_merged = df_merged.fillna(0).reset_index()
    
    # Rename columns to match Friendly output
    final_cols = ["Style_Clean", "Color_Clean", "Size_Clean", "SKU_Clean"]
    for store in sorted_stores:
        s_col = f"Stock_{store}"
        t_col = f"Transit_{store}"
        a_col = f"Alloc_{store}"
        if s_col not in df_merged.columns:
            df_merged[s_col] = 0
        if t_col not in df_merged.columns:
            df_merged[t_col] = 0
        if a_col not in df_merged.columns:
            df_merged[a_col] = 0
        final_cols.append(s_col)
        final_cols.append(t_col)
        final_cols.append(a_col)
        
    df_pivot_out = df_merged[final_cols].copy()
    df_pivot_out.rename(columns={
        "Style_Clean": "Style",
        "Color_Clean": "Color",
        "Size_Clean": "Size",
        "SKU_Clean": "SKU"
    }, inplace=True)
    
    # 10. Calculate Shortfall
    print("Calculating shortfall...")
    detailed_shortfall = []
    
    for _, row in df_skus_master.iterrows():
        style = row["Style_Clean"]
        color = row["Color_Clean"]
        size = row["Size_Clean"]
        sku = row["SKU_Clean"]
        
        for rank, store in enumerate(sorted_stores, start=1):
            base_stock = get_base_stock(rank, size)
            if base_stock == 0:
                continue # Only check sizes 7, 8, 9, 10, 11
                
            s_col = f"Stock_{store}"
            t_col = f"Transit_{store}"
            a_col = f"Alloc_{store}"
            
            curr_stock = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                s_col
            ].iloc[0]) if s_col in df_merged.columns else 0
            
            transit_qty = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                t_col
            ].iloc[0]) if t_col in df_merged.columns else 0
            
            alloc_qty = int(df_merged.loc[
                (df_merged["Style_Clean"] == style) & 
                (df_merged["Color_Clean"] == color) & 
                (df_merged["Size_Clean"] == size) & 
                (df_merged["SKU_Clean"] == sku),
                a_col
            ].iloc[0]) if a_col in df_merged.columns else 0
            
            net_qty = curr_stock + transit_qty + alloc_qty
            shortfall = max(0, base_stock - net_qty)
            
            if shortfall > 0:
                detailed_shortfall.append({
                    "Priority Rank": rank,
                    "Store Name": store,
                    "Style": style,
                    "Color": color,
                    "Size": size,
                    "SKU": sku,
                    "Base Stock": base_stock,
                    "Current Stock": curr_stock,
                    "Transit Qty": transit_qty,
                    "Allocated Qty": alloc_qty,
                    "Total Store Qty": net_qty,
                    "Shortfall Qty": shortfall
                })
                
    df_detail_short = pd.DataFrame(detailed_shortfall)
    if df_detail_short.empty:
        print("No shortfalls found!")
        df_detail_short = pd.DataFrame(columns=[
            "Priority Rank", "Store Name", "Style", "Color", "Size", "SKU",
            "Base Stock", "Current Stock", "Transit Qty", "Allocated Qty", "Total Store Qty", "Shortfall Qty"
        ])
        
    print(f"Total detailed shortfall rows: {len(df_detail_short)}")
    
    # 11. Create High-Level Summary by Store
    summary_rows = []
    for rank, store in enumerate(sorted_stores, start=1):
        store_rows = df_detail_short[df_detail_short["Store Name"] == store]
        
        s_7 = store_rows[store_rows["Size"] == "7"]["Shortfall Qty"].sum()
        s_8 = store_rows[store_rows["Size"] == "8"]["Shortfall Qty"].sum()
        s_9 = store_rows[store_rows["Size"] == "9"]["Shortfall Qty"].sum()
        s_10 = store_rows[store_rows["Size"] == "10"]["Shortfall Qty"].sum()
        s_11 = store_rows[store_rows["Size"] == "11"]["Shortfall Qty"].sum()
        total_shortfall = store_rows["Shortfall Qty"].sum()
        
        group = "Top 4" if rank <= 4 else "Remaining"
        
        summary_rows.append({
            "Priority Rank": rank,
            "Group": group,
            "Store Name": store,
            "Size 7 Shortfall": s_7,
            "Size 8 Shortfall": s_8,
            "Size 9 Shortfall": s_9,
            "Size 10 Shortfall": s_10,
            "Size 11 Shortfall": s_11,
            "Total Shortfall": total_shortfall
        })
        
    df_summary_short = pd.DataFrame(summary_rows)
    print(f"Total summary shortfall rows: {len(df_summary_short)}")
    
    # 12. Save Report with excel Styling
    print(f"\nWriting output to: {OUTPUT_PATH}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    def style_sheet(ws, df, title, freeze_pane="D3"):
        # Style Definitions
        navy_fill = PatternFill("solid", fgColor="5B3F11")    # bronze/gold theme
        alt_fill = PatternFill("solid", fgColor="FBF8F2")     # light beige alt row
        
        title_font = Font(name="Calibri", size=14, bold=True, color="5B3F11")
        hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=10)
        bold_data_font = Font(name="Calibri", size=10, bold=True)
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        thin_side = Side(style="thin", color="E0D0C0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Merge and set title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = left_align
        ws.row_dimensions[1].height = 30
        
        # Headers (row 2)
        ws.row_dimensions[2].height = 25
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = navy_fill
            cell.font = hdr_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # Data
        for row_idx in range(3, len(df) + 3):
            ws.row_dimensions[row_idx].height = 18
            use_alt = (row_idx % 2 == 0)
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                if use_alt:
                    cell.fill = alt_fill
                    
                col_name = df.columns[col_idx - 1]
                if col_name in ("Store Name", "SKU", "Style", "Color"):
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                    
                if isinstance(cell.value, (int, float)):
                    cell.value = int(cell.value)
                    
                if "Shortfall" in col_name or "Total" in col_name:
                    if cell.value and cell.value > 0:
                        cell.font = bold_data_font
                        
        # Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws.freeze_panes = freeze_pane

    actual_output_path = OUTPUT_PATH
    try:
        writer = pd.ExcelWriter(actual_output_path, engine="openpyxl")
    except PermissionError:
        actual_output_path = OUTPUT_PATH.replace(".xlsx", "_New.xlsx")
        try:
            writer = pd.ExcelWriter(actual_output_path, engine="openpyxl")
            print(f"\n[WARNING] Permission denied on '{OUTPUT_PATH}' (is the file open in Excel?). Saving to '{actual_output_path}' instead.")
        except PermissionError:
            import time
            actual_output_path = OUTPUT_PATH.replace(".xlsx", f"_{time.strftime('%H%M%S')}.xlsx")
            print(f"\n[WARNING] Permission denied on both fallback paths. Saving to unique timestamp path '{actual_output_path}'.")
            writer = pd.ExcelWriter(actual_output_path, engine="openpyxl")
        
    with writer:
        # Map store names to user-friendly clean names in shortfall summary and detailed sheets
        if not df_summary_short.empty:
            df_summary_short["Store Name"] = df_summary_short["Store Name"].map(lambda s: REVERSE_MAP.get(s, s))
        if not df_detail_short.empty:
            df_detail_short["Store Name"] = df_detail_short["Store Name"].map(lambda s: REVERSE_MAP.get(s, s))

        df_summary_short.to_excel(writer, sheet_name="Shortfall Summary", index=False, startrow=1)
        df_detail_short.to_excel(writer, sheet_name="Detailed Shortfall", index=False, startrow=1)
        
        # For SKU Pivot, friendly names for stock, transit, allocation
        display_cols = []
        for c in df_pivot_out.columns:
            if c.startswith("Stock_"):
                orig_name = c.replace("Stock_", "")
                display_cols.append(REVERSE_MAP.get(orig_name, orig_name))
            elif c.startswith("Transit_"):
                orig_name = c.replace("Transit_", "")
                display_cols.append(f"🚚 {REVERSE_MAP.get(orig_name, orig_name)}")
            elif c.startswith("Alloc_"):
                orig_name = c.replace("Alloc_", "")
                display_cols.append(f"📦 {REVERSE_MAP.get(orig_name, orig_name)}")
            else:
                display_cols.append(c)
                
        df_pivot_out.to_excel(writer, sheet_name="Footwear SKU Report", index=False, startrow=1)
        
        # Style sheets
        ws_sum = writer.sheets["Shortfall Summary"]
        style_sheet(ws_sum, df_summary_short, "Footwear Store-Wise Shortfall Summary Sheet", freeze_pane="D3")
        
        ws_det = writer.sheets["Detailed Shortfall"]
        style_sheet(ws_det, df_detail_short, "Footwear Detailed Option-Wise Shortfall Sheet", freeze_pane="C3")
        
        ws_piv = writer.sheets["Footwear SKU Report"]
        
        # Apply special header and widths for pivoted sheet
        ws_piv.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_pivot_out.columns))
        title_cell = ws_piv.cell(row=1, column=1)
        title_cell.value = "Footwear Store-Wise Stock, Transit, and Allocation Quantities"
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="5B3F11")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_piv.row_dimensions[1].height = 30
        
        ws_piv.row_dimensions[2].height = 35
        # Set custom column titles with emojis for Stock/Transit/Alloc
        for idx, col_name in enumerate(df_pivot_out.columns, start=1):
            cell = ws_piv.cell(row=2, column=idx)
            cell.value = display_cols[idx - 1]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            
            # Colour bands for Pivoted headers
            if col_name in ["Style", "Color", "Size", "SKU"]:
                cell.fill = PatternFill("solid", fgColor="5B3F11") # dark bronze
            elif col_name.startswith("Stock_"):
                cell.fill = PatternFill("solid", fgColor="8C6239") # medium bronze
            elif col_name.startswith("Transit_"):
                cell.fill = PatternFill("solid", fgColor="7F6000") # light gold
            else:
                cell.fill = PatternFill("solid", fgColor="7030A0") # purple for allocation
                
        # Style pivoted data rows
        thin_side = Side(style="thin", color="E0D0C0")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for row_idx in range(3, len(df_pivot_out) + 3):
            ws_piv.row_dimensions[row_idx].height = 18
            use_alt = (row_idx % 2 == 0)
            for col_idx in range(1, len(df_pivot_out.columns) + 1):
                cell = ws_piv.cell(row=row_idx, column=col_idx)
                cell.font = Font(name="Calibri", size=10)
                cell.border = thin_border
                if use_alt:
                    cell.fill = PatternFill("solid", fgColor="FBF8F2")
                    
                col_name = df_pivot_out.columns[col_idx - 1]
                if col_name in ("Style", "Color", "SKU"):
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                if isinstance(cell.value, (int, float)):
                    cell.value = int(cell.value)
                    
        # Column widths for Pivoted sheet
        for col in ws_piv.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col[1:]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws_piv.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        ws_piv.freeze_panes = "E3"
        
    print("\nReport Saved successfully!")
    print("=" * 65)

if __name__ == "__main__":
    main()
