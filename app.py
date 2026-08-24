"""
Flask UI for Increff Allocation Checker
Run:  python app.py
Then open  http://localhost:5050
"""

import os
import io
import sys
import json
import datetime
import threading
import time

import pandas as pd
from flask import (
    Flask, render_template, request, jsonify,
    send_file, session
)
from werkzeug.utils import secure_filename

# ── bring in the allocation logic ────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from check_allocation import (
    PRIORITY_TO_EBO, PRIORITY_NUM_COL,
    EBO_COL, SKU_COL, ALLOC_COL, AVAIL_COL,
    OUTPUT_DIR, _apply_sheet_style,
)

# ─────────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "increff-alloc-ui-2026"
app.config['JSON_SORT_KEYS'] = False
try:
    app.json.sort_keys = False
except AttributeError:
    pass

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# In-memory store for uploaded data (per process — single user tool)
_state = {
    "df_replen"  : None,
    "df_priority": None,
    "replen_name": None,
    "priority_name": None,
    "replen_path": None,
    "all_skus"   : [],
    "all_stores" : [],
    "df_ebo_curr": None,
    "df_ebo_tran": None,
    "ebo_summary": None,
    "ebo_status" : {"loaded": False, "curr_rows": 0, "tran_rows": 0, "last_loaded": None},
    "ag_store_data": [],
    "ag_wise_data": [],
    "ag_status": {"loaded": False, "error": None, "last_loaded": None}
}

def _read_excel_fast(path, sheet_name=None, header_row=0):
    # Use pandas read_excel with calamine engine if available, fallback to openpyxl
    s_name = sheet_name if sheet_name is not None else 0
    try:
        df = pd.read_excel(path, sheet_name=s_name, header=header_row, engine='calamine')
    except (ImportError, ValueError, Exception):
        df = pd.read_excel(path, sheet_name=s_name, header=header_row, engine='openpyxl')

    
    # Strip whitespace from column names and clean named/unnamed columns
    new_cols = []
    for i, col in enumerate(df.columns):
        if pd.isna(col) or str(col).strip() == "":
            new_cols.append(f"Unnamed_{i}")
        else:
            new_cols.append(str(col).strip())
    df.columns = new_cols
    
    # Strip whitespace from string values while preserving NaN/None
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].apply(lambda x: str(x).strip() if pd.notna(x) else x)
            
    return df


EBO_STOCK_DIR = os.path.join(os.path.dirname(__file__), "ebo stock track data")

def _get_excel_path(subfolder):
    import glob
    folder = os.path.join(EBO_STOCK_DIR, subfolder)
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith('~')]
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def clean_code(val):
    if pd.isna(val):
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


PRIORITY_TO_CODE = {
    "SARATH CITY MALL": "778",
    "TSPL-Vizag": "939",
    "Kuvempunagar Mysore": "815",
    "SALEM": "744",
    "KATRAJ EBO": "817",
    "TSPL-TUP": "743",
    "BASANT NAGAR": "767",
    "ERODE": "779",
    "PUNE KHARADI": "763",
    "CHIKKAJALA": "765",
    "Belgum": None,
    "ROURKELA": "887",
    "HADAPSAR": "885",
    "ATTIBELLE": "886",
    "PUNE PIMPLE": "764",
    "HSR-EBO": "757",
    "WAGHOLI": "930",
    "HYD-EBO": "756",
    "RS PURAM": "739",
    "AS RAO NAGAR": "883",
    "VIZIANAGARAM": "931",
    "KOLHAPUR EBO": "769",
    "PONDICHERRY": "762",
    "BILASPUR": "889",
    "KHARAGPUR": "932",
    "VIJAYAWADA": "766",
    "HASSAN EBO": "781",
    "ANANTAPUR": "927",
    "JABALPUR": "891",
    "MYSORE": "753",
    "DODDABALLAPUR": "780",
    "VELLORE": "888",
    "Hubli (Shirur Park)": "890",
    "KUKATPALLY HYD": "816",
    "Govind Nagar": "884",
    "NAMAKKAL": "850",
    "METTUPALAYAM": "768",
    "BHAGALPUR": "892",
    "BHOPAL EBO": "776",
    "BALLARI": "771",
    "DURGAPUR": "782",
    "MANI SQUARE MALL": "775",
    "HUBBALI EBO": "770",
    "DIVINITY-MALL": "741",
    "RAIPUR EBO": "773",
    "TEX VALLEY": "929",
    "SENTRUM MALL": "774",
    "MOSHI": "928",
    "SHAHEEN BAGH": "894",
    "INDORE-EBO": "758",
    "UDUPI": "893",
    "ELAN EPIC MALL": "777",
    "SALEM-2 EBO": "934"
}


def _load_ebo_data_internal():
    errors = []
    curr_loaded = False
    tran_loaded = False
    
    df_curr = pd.DataFrame()
    df_tran = pd.DataFrame()
    
    ebo_curr_path = _get_excel_path("current stock")
    ebo_tran_path = _get_excel_path("intransit")
    ebo_grc_path  = _get_excel_path("grc date")
    
    if ebo_curr_path and os.path.exists(ebo_curr_path):
        try:
            df_curr = _read_excel_fast(ebo_curr_path)
            curr_loaded = True
        except Exception as e:
            errors.append(f"Error loading Current Stock: {e}")
    else:
        errors.append("Current Stock Excel file not found.")
        
    if ebo_tran_path and os.path.exists(ebo_tran_path):
        try:
            df_tran = _read_excel_fast(ebo_tran_path)
            tran_loaded = True
        except Exception as e:
            errors.append(f"Error loading Transit Data: {e}")
    else:
        errors.append("Transit Data Excel file not found.")
        
    if not curr_loaded and not tran_loaded:
        _state["ebo_status"] = {
            "loaded": False,
            "curr_rows": 0,
            "tran_rows": 0,
            "last_loaded": None,
            "errors": errors
        }
        return False, errors
        
    # Clean data
    if curr_loaded and not df_curr.empty:
        c_map = {}
        sc_col = next((c for c in df_curr.columns if str(c).strip().lower() in ['site_code', 'site code', 'sitecode', 'store code', 'store_code', 'storecode']), None)
        sn_col = next((c for c in df_curr.columns if str(c).strip().lower() in ['store name', 'store_name', 'owner site', 'owner_site']), None)
        qty_col = next((c for c in df_curr.columns if str(c).strip().lower() in ['quantity', 'qty', 'stock quantity', 'stock_quantity']), None)
        style_col = next((c for c in df_curr.columns if str(c).strip().lower() == 'style'), None)
        color_col = next((c for c in df_curr.columns if str(c).strip().lower() in ['color', 'colour']), None)
        size_col = next((c for c in df_curr.columns if str(c).strip().lower() == 'size'), None)
        
        if sc_col: c_map[sc_col] = 'store_code'
        if sn_col: c_map[sn_col] = 'Store Name'
        if qty_col: c_map[qty_col] = 'quantity'
        if style_col: c_map[style_col] = 'Style'
        if color_col: c_map[color_col] = 'Colour'
        if size_col: c_map[size_col] = 'Size'
        df_curr.rename(columns=c_map, inplace=True)
        
        for col in ['store_code', 'quantity', 'Style', 'Colour', 'Size', 'Store Name']:
            if col not in df_curr.columns:
                df_curr[col] = 0 if col == 'quantity' else ""
                
        df_curr['store_code'] = df_curr['store_code'].apply(clean_code)
        df_curr['quantity'] = pd.to_numeric(df_curr['quantity'], errors='coerce').fillna(0).astype(int)
        df_curr['Style'] = df_curr['Style'].astype(str).str.strip()
        df_curr['Colour'] = df_curr['Colour'].astype(str).str.strip()
        df_curr['Size'] = df_curr['Size'].astype(str).str.strip()
        
    if tran_loaded and not df_tran.empty:
        t_map = {}
        t_sc_col = next((c for c in df_tran.columns if str(c).strip().lower() in ['store code', 'store_code', 'storecode', 'site code', 'site_code', 'sitecode']), None)
        t_sn_col = next((c for c in df_tran.columns if str(c).strip().lower() in ['ebo name', 'ebo_name', 'store name', 'store_name', 'owner site', 'owner_site']), None)
        t_qty_col = next((c for c in df_tran.columns if str(c).strip().lower() in ['transit qty', 'transit_qty', 'quantity', 'qty']), None)
        t_style_col = next((c for c in df_tran.columns if str(c).strip().lower() == 'style'), None)
        t_color_col = next((c for c in df_tran.columns if str(c).strip().lower() in ['color', 'colour']), None)
        t_size_col = next((c for c in df_tran.columns if str(c).strip().lower() == 'size'), None)
        
        if t_sc_col: t_map[t_sc_col] = 'Store Code'
        if t_sn_col: t_map[t_sn_col] = 'EBO Name'
        if t_qty_col: t_map[t_qty_col] = 'Transit Qty'
        if t_style_col: t_map[t_style_col] = 'STYLE'
        if t_color_col: t_map[t_color_col] = 'COLOR'
        if t_size_col: t_map[t_size_col] = 'SIZE'
        df_tran.rename(columns=t_map, inplace=True)
        
        for col in ['Store Code', 'Transit Qty', 'STYLE', 'COLOR', 'SIZE', 'EBO Name']:
            if col not in df_tran.columns:
                df_tran[col] = 0 if col == 'Transit Qty' else ""
                
        df_tran['Store Code'] = df_tran['Store Code'].apply(clean_code)
        df_tran['Transit Qty'] = pd.to_numeric(df_tran['Transit Qty'], errors='coerce').fillna(0).astype(int)
        df_tran['STYLE'] = df_tran['STYLE'].astype(str).str.strip()
        df_tran['COLOR'] = df_tran['COLOR'].astype(str).str.strip()
        df_tran['SIZE'] = df_tran['SIZE'].astype(str).str.strip()
        
    # Aggregate summary
    curr_agg = pd.DataFrame(columns=['store_code', 'store_name', 'curr_styles', 'curr_qty'])
    tran_agg = pd.DataFrame(columns=['Store Code', 'tran_name', 'tran_styles', 'tran_qty'])
    
    if curr_loaded and not df_curr.empty:
        curr_agg = df_curr.groupby('store_code').agg(
            store_name=('Store Name', 'first'),
            curr_styles=('Style', 'nunique'),
            curr_qty=('quantity', 'sum')
        ).reset_index()
        
    if tran_loaded and not df_tran.empty:
        tran_agg = df_tran.groupby('Store Code').agg(
            tran_name=('EBO Name', 'first'),
            tran_styles=('STYLE', 'nunique'),
            tran_qty=('Transit Qty', 'sum')
        ).reset_index()
        
    merged = pd.merge(
        curr_agg,
        tran_agg,
        left_on='store_code',
        right_on='Store Code',
        how='outer'
    )
    
    merged['store_code'] = merged['store_code'].fillna(merged['Store Code'])
    merged['Store Name'] = merged['store_name'].fillna(merged['tran_name'])
    
    merged['curr_styles'] = merged['curr_styles'].fillna(0).astype(int)
    merged['curr_qty'] = merged['curr_qty'].fillna(0).astype(int)
    merged['tran_styles'] = merged['tran_styles'].fillna(0).astype(int)
    merged['tran_qty'] = merged['tran_qty'].fillna(0).astype(int)
    merged['total_qty'] = merged['curr_qty'] + merged['tran_qty']

    # ── Load Allocation data per store ─────────────────────────────────────────
    alloc_agg = pd.DataFrame(columns=['alloc_store_code', 'alloc_styles', 'alloc_qty'])
    try:
        import glob as _glob
        alloc_dir = os.path.join(EBO_STOCK_DIR, "ALLOCATION")
        alloc_files = _glob.glob(os.path.join(alloc_dir, "*.xlsx"))
        alloc_files = [f for f in alloc_files if not os.path.basename(f).startswith('~')]
        if alloc_files:
            latest_alloc = max(alloc_files, key=os.path.getmtime)
            df_alloc_store = _read_excel_fast(latest_alloc)
            df_alloc_store.columns = [str(c).strip() for c in df_alloc_store.columns]
            sc_col = next((c for c in df_alloc_store.columns if 'store' in c.lower() and 'code' in c.lower()), None)
            style_col_a = next((c for c in df_alloc_store.columns if c.strip().lower() == 'style'), None)
            qty_col_a = next((c for c in df_alloc_store.columns if 'allocated' in c.lower() and 'qty' in c.lower()), None)
            if sc_col and style_col_a and qty_col_a:
                df_alloc_store[sc_col] = df_alloc_store[sc_col].apply(clean_code)
                df_alloc_store[qty_col_a] = pd.to_numeric(df_alloc_store[qty_col_a], errors='coerce').fillna(0)
                
                # Build style stores allocation map for top 100 style enrichment
                style_stores_alloc = {}
                df_alloc_active = df_alloc_store[df_alloc_store[qty_col_a] > 0]
                for style_code, group in df_alloc_active.groupby(style_col_a):
                    style_stores_alloc[str(style_code).strip().upper()] = set(group[sc_col].dropna().unique())
                _state['style_stores_alloc'] = style_stores_alloc
                
                alloc_agg = df_alloc_store.groupby(sc_col).agg(
                    alloc_styles=(style_col_a, 'nunique'),
                    alloc_qty=(qty_col_a, 'sum')
                ).reset_index().rename(columns={sc_col: 'alloc_store_code'})
                alloc_agg['alloc_qty'] = alloc_agg['alloc_qty'].astype(int)
    except Exception as e:
        errors.append(f"Error loading Allocation per store: {e}")

    if not alloc_agg.empty and 'alloc_store_code' in alloc_agg.columns:
        merged = pd.merge(merged, alloc_agg, left_on='store_code', right_on='alloc_store_code', how='left')
    merged['alloc_styles'] = merged['alloc_styles'].fillna(0).astype(int) if 'alloc_styles' in merged.columns else 0
    merged['alloc_qty'] = merged['alloc_qty'].fillna(0).astype(int) if 'alloc_qty' in merged.columns else 0

    # ── Load GRC dates and get the latest GRC date per store ──────────────────
    grc_map     = {}  # store_code (str) -> latest GRC date display string
    grc_raw_map = {}  # store_code (str) -> latest GRC date as datetime (for age calc)
    if ebo_grc_path and os.path.exists(ebo_grc_path):
        try:
            df_grc = _read_excel_fast(ebo_grc_path)
            df_grc['Store_Code'] = df_grc['Store_Code'].apply(clean_code)
            df_grc['GRC Date'] = pd.to_datetime(df_grc['GRC Date'], errors='coerce')
            latest_grc = df_grc.groupby('Store_Code')['GRC Date'].max().reset_index()
            for _, r in latest_grc.iterrows():
                if pd.notna(r['GRC Date']):
                    grc_map[str(r['Store_Code'])]     = r['GRC Date'].strftime('%d-%b-%Y')
                    grc_raw_map[str(r['Store_Code'])] = r['GRC Date'].to_pydatetime()
        except Exception as e:
            errors.append(f"Error loading GRC Date file: {e}")
    
    summary_list = merged[['store_code', 'Store Name', 'curr_styles', 'curr_qty', 'tran_styles', 'tran_qty', 'total_qty', 'alloc_styles', 'alloc_qty']].to_dict(orient='records')
    
    # ── Build a name->store_code lookup from actual loaded data ───────────────
    # This is more reliable than PRIORITY_TO_CODE which uses different name formats
    name_to_code = {}  # normalised store name -> store_code
    for item in summary_list:
        raw = str(item.get('Store Name', '')).strip()
        name_to_code[raw.lower()] = item['store_code']
    
    def _normalise(s):
        """Strip common prefixes/suffixes and lower for fuzzy matching."""
        s = str(s).lower().strip()
        for pfx in ('tspl-', 'tspl '):
            if s.startswith(pfx):
                s = s[len(pfx):]
        for sfx in (' ebo', '-ebo', ' store', ' mall'):
            if s.endswith(sfx):
                s = s[:-len(sfx)]
        # Remove dots and spaces for even fuzzier matching (e.g., 'R.S PURAM' vs 'RS PURAM')
        s = s.replace('.', '').replace(' ', '')
        return s.strip()
 
    # Build normalised lookup as well
    norm_to_code = {_normalise(k): v for k, v in name_to_code.items()}
    
    # Load priorities from state or scan priority list directory for any xlsx
    pri_dict = {}
    df_pri = _state.get("df_priority")
    if df_pri is None or df_pri.empty:
        try:
            pri_dir = os.path.join(os.path.dirname(__file__), "priority list")
            if os.path.isdir(pri_dir):
                xlsx_files = [f for f in os.listdir(pri_dir) if f.lower().endswith('.xlsx')]
                if xlsx_files:
                    pri_path = os.path.join(pri_dir, xlsx_files[0])  # use first found
                    try:
                        df_pri = _read_excel_fast(pri_path, sheet_name="Export")
                    except (ValueError, Exception):
                        df_pri = _read_excel_fast(pri_path, sheet_name=0)
        except Exception as e:
            errors.append(f"Error loading fallback Priority list: {e}")
            
    if df_pri is not None and not df_pri.empty:
        # Normalize column names
        priority_col_mapping = {
            "Store Name": ["Store Name", "store name", "STORE NAME", "StoreName", "store_name"],
            PRIORITY_NUM_COL: ["priority number", "priority", "Priority", "Priority No",
                               "Priority Number", "priority_number", "priority list"]
        }
        rename_priority = {}
        for standard_col, options in priority_col_mapping.items():
            for col in df_pri.columns:
                if col.strip().lower() == standard_col.lower() or col.strip() in options:
                    rename_priority[col] = standard_col
                    break
        df_pri = df_pri.rename(columns=rename_priority)

        pri_name_col  = 'Store Name'   if 'Store Name'   in df_pri.columns else None
        pri_num_col   = PRIORITY_NUM_COL if PRIORITY_NUM_COL in df_pri.columns else None
        # Fallback: if PRIORITY_NUM_COL not found, try 'Priority'
        if pri_num_col is None and 'Priority' in df_pri.columns:
            pri_num_col = 'Priority'

        if pri_name_col and pri_num_col:
            for _, row in df_pri.iterrows():
                p_name = str(row[pri_name_col]).strip()
                try:
                    p_num = int(row[pri_num_col])
                except Exception:
                    p_num = 999

                code = None

                # 1. Exact match on full name
                code = name_to_code.get(p_name.lower())

                # 2. Normalised match (strip TSPL prefix and EBO/Store/Mall suffix)
                if not code:
                    code = norm_to_code.get(_normalise(p_name))

                # 3. Legacy PRIORITY_TO_CODE lookup (short names)
                if not code:
                    code = PRIORITY_TO_CODE.get(p_name)

                # 4. Substring fallback across norm_to_code
                if not code:
                    p_norm = _normalise(p_name)
                    for stored_norm, stored_code in norm_to_code.items():
                        if p_norm in stored_norm or stored_norm in p_norm:
                            code = stored_code
                            break

                if code:
                    pri_dict[code] = p_num

    # ── Load Increff Qty from Google Sheet ────────────────────────────────────
    increff_dict = {}
    try:
        sheet_url = 'https://docs.google.com/spreadsheets/d/1aP3Yv0p51acvtPpPA9KABJSvhj2nQi5Y3WHIJLHSrAY/export?format=csv&gid=0'
        import urllib.request
        with urllib.request.urlopen(sheet_url, timeout=2.0) as response:
            csv_content = response.read()
        df_sheet = pd.read_csv(io.BytesIO(csv_content))
        df_sheet['INCREFF QTY'] = pd.to_numeric(df_sheet['INCREFF QTY'], errors='coerce').fillna(0)
        # Filters: Garments only, NOT dispached
        if 'MATERIAL TYPE' in df_sheet.columns:
            df_sheet = df_sheet[df_sheet['MATERIAL TYPE'].astype(str).str.lower().str.strip() == 'garments']
        if 'WH PROCESS STATUS' in df_sheet.columns:
            df_sheet = df_sheet[df_sheet['WH PROCESS STATUS'].astype(str).str.lower().str.strip() != 'dispached']
            
        if 'Store code' in df_sheet.columns:
            df_sheet['Store code'] = df_sheet['Store code'].apply(clean_code)
            grouped_sheet = df_sheet.groupby('Store code')['INCREFF QTY'].sum().reset_index()
            for _, s_row in grouped_sheet.iterrows():
                code = str(s_row['Store code']).strip()
                qty = s_row['INCREFF QTY']
                
                if code and code != '#REF!':
                    increff_dict[code] = increff_dict.get(code, 0) + qty
    except Exception as e:
        errors.append(f"Error loading Increff Qty from Google Sheet: {e}")

    for row in summary_list:
        row['priority'] = pri_dict.get(row['store_code'], 999)
        row['grc_date'] = grc_map.get(str(row['store_code']), '-')
        row['increff_qty'] = increff_dict.get(row['store_code'], 0)

        # ── Order status based on days since last GRC ──────────────────────────
        import datetime as _dt
        grc_dt = grc_raw_map.get(str(row['store_code']))
        if grc_dt is None:
            row['order_status'] = 'Needed'          # no GRC record at all
        else:
            today = _dt.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            days_since = (today - grc_dt.replace(tzinfo=None)).days
            if days_since <= 6:
                row['order_status'] = 'Satisfied'
            elif row.get('tran_qty', 0) > 0:
                row['order_status'] = 'In Transit'
            else:
                row['order_status'] = 'Needed'
        
    # Sort by priority number (1, 2, 3, ...) ascending
    summary_list.sort(key=lambda x: x['priority'])
    
    # Cache in global state
    _state['df_ebo_curr'] = df_curr
    _state['df_ebo_tran'] = df_tran
    _state['ebo_summary'] = summary_list
    
    import datetime
    _state['ebo_status'] = {
        "loaded": True,
        "curr_rows": len(df_curr) if curr_loaded else 0,
        "tran_rows": len(df_tran) if tran_loaded else 0,
        "last_loaded": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    return True, errors


def _load_new_style_data_internal():
    import glob
    import os
    import pandas as pd
    folder = os.path.join(os.path.dirname(__file__), "new style performance")
    if not os.path.exists(folder):
        return False, "Folder 'new style performance' does not exist."
    
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    files = [f for f in files if not os.path.basename(f).startswith('~') and 'wariq' not in os.path.basename(f).lower() and 'hosur' not in os.path.basename(f).lower() and 'pipeline' not in os.path.basename(f).lower() and 'live status' not in os.path.basename(f).lower() and 'live_status' not in os.path.basename(f).lower()]
    if not files:
        return False, "No Excel file found in 'new style performance' folder."
        
    latest_file = max(files, key=os.path.getmtime)
    try:
        df = _read_excel_fast(latest_file)
        
        # Make sure STYLE column exists
        style_col = next((c for c in df.columns if str(c).strip().upper() == 'STYLE'), 'STYLE')
        if style_col != 'STYLE' and style_col in df.columns:
            df.rename(columns={style_col: 'STYLE'}, inplace=True)
            
        launch_col = next((c for c in df.columns if 'launch' in str(c).lower()), 'First Launch Month')
        if launch_col != 'First Launch Month' and launch_col in df.columns:
            df.rename(columns={launch_col: 'First Launch Month'}, inplace=True)
            
        if 'First Launch Month' in df.columns:
            df['First Launch Month'] = df['First Launch Month'].astype(str).str.strip()
            df['First Launch Month'] = df['First Launch Month'].replace('No Sales', 'GRN Pending')
        else:
            df['First Launch Month'] = 'GRN Pending'
            
        # Load pipeline data for JFM/AMJ sales
        pipeline_files = glob.glob(os.path.join(folder, "*pipeline*.xlsx"))
        pipeline_files = [f for f in pipeline_files if not os.path.basename(f).startswith('~')]
        pipeline_file = max(pipeline_files, key=os.path.getmtime) if pipeline_files else None
        
        pipeline_map = {}
        if pipeline_file and os.path.exists(pipeline_file):
            try:
                df_p = _read_excel_fast(pipeline_file, header_row=2)
                grn_col = next((c for c in df_p.columns if "Pending GRN Units" in str(c)), None)
                cols_to_sum = [
                    'Style ID',
                    'D2C JAN', 'D2C FEB', 'D2C MAR', 'D2C APR.1', 'D2C MAY.1', 'D2C JUN.1',
                    'EBO JAN', 'EBO FEB', 'EBO MAR', 'EBO APR.1', 'EBO MAY.1', 'EBO JUN.1'
                ]
                if grn_col:
                    cols_to_sum.append(grn_col)
                present_cols = [c for c in cols_to_sum if c in df_p.columns]
                for col in present_cols[1:]:
                    df_p[col] = pd.to_numeric(df_p[col], errors='coerce').fillna(0)
                
                df_p_grouped = df_p.groupby('Style ID')[present_cols[1:]].sum().reset_index()
                for _, row in df_p_grouped.iterrows():
                    style_p = str(row['Style ID']).strip().upper()
                    d2c_jan = row.get('D2C JAN', 0)
                    d2c_feb = row.get('D2C FEB', 0)
                    d2c_mar = row.get('D2C MAR', 0)
                    d2c_apr = row.get('D2C APR.1', 0)
                    d2c_may = row.get('D2C MAY.1', 0)
                    d2c_jun = row.get('D2C JUN.1', 0)
                    
                    ebo_jan = row.get('EBO JAN', 0)
                    ebo_feb = row.get('EBO FEB', 0)
                    ebo_mar = row.get('EBO MAR', 0)
                    ebo_apr = row.get('EBO APR.1', 0)
                    ebo_may = row.get('EBO MAY.1', 0)
                    ebo_jun = row.get('EBO JUN.1', 0)
                    
                    grn_pending_qty = int(row.get(grn_col, 0)) if grn_col else 0
                    
                    d2c_jfm = int(d2c_jan + d2c_feb + d2c_mar)
                    ebo_jfm = int(ebo_jan + ebo_feb + ebo_mar)
                    d2c_amj = int(d2c_apr + d2c_may + d2c_jun)
                    ebo_amj = int(ebo_apr + ebo_may + ebo_jun)
                    
                    total_d2c = d2c_jfm + d2c_amj
                    total_ebo = ebo_jfm + ebo_amj
                    
                    pipeline_map[style_p] = {
                        "d2c_jfm": d2c_jfm,
                        "ebo_jfm": ebo_jfm,
                        "d2c_amj": d2c_amj,
                        "ebo_amj": ebo_amj,
                        "total_d2c": total_d2c,
                        "total_ebo": total_ebo,
                        "grn_pending_qty": grn_pending_qty
                    }
            except Exception as ex:
                print("Error loading/parsing PIPELINE VIEW:", ex)

        # Ensure EBO data is loaded in _state
        df_curr = _state.get('df_ebo_curr')
        df_tran = _state.get('df_ebo_tran')

        curr_stock_map = {}
        if df_curr is not None and not df_curr.empty:
            curr_grouped = df_curr.groupby('Style')['quantity'].sum().to_dict()
            curr_stock_map = {str(k).strip().upper(): int(v) for k, v in curr_grouped.items()}

        tran_stock_map = {}
        if df_tran is not None and not df_tran.empty:
            tran_grouped = df_tran.groupby('STYLE')['Transit Qty'].sum().to_dict()
            tran_stock_map = {str(k).strip().upper(): int(v) for k, v in tran_grouped.items()}

        # Load OMS stock data from CSV
        import glob
        csv_files = glob.glob(os.path.join(folder, "Inventory Available for Sales - OMS*.csv"))
        latest_csv = max(csv_files, key=os.path.getmtime) if csv_files else None
        
        d2c_map = {}
        cred_map = {}
        ebo_map = {}
        common_pool_map = {}
        
        if latest_csv and os.path.exists(latest_csv):
            try:
                df_inv = pd.read_csv(latest_csv, usecols=['Style', 'Reservation Pool', 'Total Available Quantity'])
                df_inv['Style'] = df_inv['Style'].astype(str).str.strip().str.upper()
                df_inv['Reservation Pool'] = df_inv['Reservation Pool'].astype(str).str.strip()
                df_inv['Total Available Quantity'] = pd.to_numeric(df_inv['Total Available Quantity'], errors='coerce').fillna(0).astype(int)
                
                grouped_inv = df_inv.groupby(['Style', 'Reservation Pool'])['Total Available Quantity'].sum().reset_index()
                
                for _, row in grouped_inv.iterrows():
                    style_p = row['Style']
                    pool = row['Reservation Pool']
                    qty = int(row['Total Available Quantity'])
                    
                    if pool == 'D2C-Marketplaces':
                        d2c_map[style_p] = d2c_map.get(style_p, 0) + qty
                    elif pool == 'CRED':
                        cred_map[style_p] = cred_map.get(style_p, 0) + qty
                    elif pool == 'EBO':
                        ebo_map[style_p] = ebo_map.get(style_p, 0) + qty
                    elif pool == 'Common_Pool-TECHNO SPORTSWEAR PRIVATE LIMITED-wms_tiruppur':
                        common_pool_map[style_p] = common_pool_map.get(style_p, 0) + qty
            except Exception as ex:
                print("Error loading/parsing OMS inventory CSV:", ex)

        # Load GT inventory pool
        gt_pool_map = {}
        gt_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\gt inventory pool"
        if os.path.exists(gt_dir):
            try:
                gt_files = glob.glob(os.path.join(gt_dir, "*.xlsx"))
                gt_files = [f for f in gt_files if not os.path.basename(f).startswith('~')]
                if gt_files:
                    latest_gt = max(gt_files, key=os.path.getmtime)
                    df_gt = pd.read_excel(latest_gt)
                    df_gt.columns = [str(c).strip() for c in df_gt.columns]
                    style_col_gt = next((c for c in df_gt.columns if c.strip().lower() == 'style'), None)
                    qty_col_gt   = next((c for c in df_gt.columns if c.strip().lower() == 'grand total'), None)
                    if style_col_gt and qty_col_gt:
                        df_gt[style_col_gt] = df_gt[style_col_gt].astype(str).str.strip().str.upper()
                        df_gt[qty_col_gt]   = pd.to_numeric(df_gt[qty_col_gt], errors='coerce').fillna(0)
                        gt_grouped = df_gt.groupby(style_col_gt)[qty_col_gt].sum().to_dict()
                        gt_pool_map = {str(k).strip().upper(): int(round(v)) for k, v in gt_grouped.items()}
            except Exception as ex:
                print("Error loading/parsing GT inventory pool:", ex)
        _state['gt_pool_map'] = gt_pool_map

        # Load live status
        import glob
        live_files = glob.glob(os.path.join(folder, "*live*status*.xlsx")) + glob.glob(os.path.join(folder, "*live_status*.xlsx"))
        live_files = [f for f in live_files if not os.path.basename(f).startswith('~')]
        live_file = max(live_files, key=os.path.getmtime) if live_files else r"D:\INCREFF ORDER PUNCH\new style performance\LIVE STATUS.xlsx"
        
        live_styles = set()
        if os.path.exists(live_file):
            try:
                # Peek at columns to handle dynamic formats
                peek_live = pd.read_excel(live_file, nrows=0)
                available_cols = [str(c).strip() for c in peek_live.columns]
                
                # Check for Model No. vs SKU column
                model_col = next((c for c in available_cols if c.lower() in ['model no.', 'model_no', 'model no', 'style']), None)
                sku_col = next((c for c in available_cols if c.lower() in ['sku with color', 'sku_with_color', 'client sku id / ean', 'client_sku_id_/_ean', 'old sku', 'old_sku', 'sku']), None)
                
                cols_to_use = []
                if model_col:
                    cols_to_use.append(model_col)
                elif sku_col:
                    cols_to_use.append(sku_col)
                
                # Identify present G-L columns
                gl_cols = ['Nykaa', 'Ajio', 'Myntra', 'Amazon', 'Flipkart FSN', 'Website']
                present_gl_cols = [c for c in gl_cols if c in available_cols]
                cols_to_use.extend(present_gl_cols)
                
                df_live = pd.read_excel(live_file, usecols=cols_to_use)
                
                # Define inactive check
                def is_inactive(val):
                    if pd.isna(val):
                        return True
                    s = str(val).strip().upper()
                    return s in ('', '0', '0.0', 'NA', 'N/A', '-')
                
                # A row is live if NOT all present G-L columns are inactive (meaning at least one is active)
                if present_gl_cols:
                    try:
                        inactive_mask = df_live[present_gl_cols].map(is_inactive)
                    except AttributeError:
                        inactive_mask = df_live[present_gl_cols].applymap(is_inactive)
                    df_live['sku_live'] = ~inactive_mask.all(axis=1)
                else:
                    df_live['sku_live'] = True
                    
                df_live_active = df_live[df_live['sku_live']]
                
                if model_col:
                    live_styles = set(df_live_active[model_col].dropna().astype(str).str.strip().str.upper().unique())
                elif sku_col:
                    live_skus = df_live_active[sku_col].dropna().astype(str).str.strip().str.upper().unique()
                    
                    # Match main styles against SKU strings
                    main_styles = set(df['STYLE'].dropna().astype(str).str.strip().str.upper().unique())
                    for sku in live_skus:
                        for s in main_styles:
                            if s in sku:
                                live_styles.add(s)
            except Exception as ex:
                print(f"Error loading/parsing live status file {live_file}:", ex)

        # Load Allocation data from ALLOCATION folder
        alloc_map = {}
        alloc_dir = os.path.join(EBO_STOCK_DIR, "ALLOCATION")
        try:
            alloc_files = glob.glob(os.path.join(alloc_dir, "*.xlsx"))
            alloc_files = [f for f in alloc_files if not os.path.basename(f).startswith('~')]
            if alloc_files:
                latest_alloc = max(alloc_files, key=os.path.getmtime)
                df_alloc = _read_excel_fast(latest_alloc)
                # Normalize column names to strip whitespace
                df_alloc.columns = [str(c).strip() for c in df_alloc.columns]
                style_col_a = next((c for c in df_alloc.columns if c.strip().lower() == 'style'), None)
                qty_col_a   = next((c for c in df_alloc.columns if 'allocated' in c.lower() and 'qty' in c.lower()), None)
                if style_col_a and qty_col_a:
                    df_alloc[style_col_a] = df_alloc[style_col_a].astype(str).str.strip().str.upper()
                    df_alloc[qty_col_a]   = pd.to_numeric(df_alloc[qty_col_a], errors='coerce').fillna(0)
                    alloc_map = df_alloc.groupby(style_col_a)[qty_col_a].sum().apply(int).to_dict()
        except Exception as ex:
            print("Error loading/parsing ALLOCATION file:", ex)

        # Load LRN Date from Google Sheets (cached in _state)
        lrn_date_map = _state.get('lrn_date_map', {})
        if not lrn_date_map:  # only fetch if not already cached
            try:
                urls = [
                    ('https://docs.google.com/spreadsheets/d/1w13x8Zs4jkCKtRA908wZCiEMYy_mKOm8XgQykikcetc/export?format=csv&gid=0', 1),
                    ('https://docs.google.com/spreadsheets/d/1w13x8Zs4jkCKtRA908wZCiEMYy_mKOm8XgQykikcetc/export?format=csv&gid=604317387', 0),
                    ('https://docs.google.com/spreadsheets/d/1w13x8Zs4jkCKtRA908wZCiEMYy_mKOm8XgQykikcetc/export?format=csv&gid=774433982', 1)
                ]
                import concurrent.futures

                def parse_date_safe(d_str):
                    try:
                        return pd.to_datetime(d_str, dayfirst=True)
                    except:
                        return pd.Timestamp.min

                def fetch_lrn(url_info):
                    url, hdr = url_info
                    local_map = {}
                    try:
                        df_lrn = pd.read_csv(url, header=hdr)
                        date_col = next((c for c in df_lrn.columns if 'date' in str(c).lower()), df_lrn.columns[0])
                        style_col = next((c for c in df_lrn.columns if 'style' in str(c).lower()), None)
                        if date_col and style_col:
                            df_lrn = df_lrn[[style_col, date_col]].dropna()
                            df_lrn = df_lrn[df_lrn[style_col].astype(str).str.lower() != 'nan']
                            df_lrn = df_lrn[df_lrn[date_col].astype(str).str.lower() != 'nan']
                            df_lrn['_base'] = df_lrn[style_col].astype(str).str.strip().str.split('/').str[0].str.strip().str.upper()
                            df_lrn['_date_parsed'] = pd.to_datetime(df_lrn[date_col].astype(str), dayfirst=True, errors='coerce')
                            df_lrn = df_lrn.dropna(subset=['_date_parsed'])
                            # Keep earliest date per style
                            earliest = df_lrn.groupby('_base')['_date_parsed'].min()
                            date_str_map = df_lrn.set_index('_base')[date_col].to_dict()
                            for style, ts in earliest.items():
                                matching = df_lrn[df_lrn['_base'] == style]
                                first_row = matching.loc[matching['_date_parsed'].idxmin()]
                                local_map[style] = str(first_row[date_col]).strip()
                    except Exception as inner_ex:
                        print(f"Error loading LRN Date map from {url}:", inner_ex)
                    return local_map

                with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                    results = list(executor.map(fetch_lrn, urls))
                    for res in results:
                        for k, v in res.items():
                            if k not in lrn_date_map:
                                lrn_date_map[k] = v
                            else:
                                try:
                                    if pd.to_datetime(v, dayfirst=True) < pd.to_datetime(lrn_date_map[k], dayfirst=True):
                                        lrn_date_map[k] = v
                                except:
                                    pass
                _state['lrn_date_map'] = lrn_date_map
            except Exception as ex:
                print("Error loading LRN Date map:", ex)

        # Load GRN Date from EBO STOCK DIR (cached in _state)
        grn_date_map = _state.get('grn_date_map', {})
        if not grn_date_map:  # only load if not already cached
            try:
                grn_folder = os.path.join(EBO_STOCK_DIR, "data of grn")
                if os.path.exists(grn_folder):
                    grn_files = glob.glob(os.path.join(grn_folder, "*.csv"))
                    all_grn_frames = []
                    for gf in grn_files:
                        try:
                            df_g = pd.read_csv(gf, usecols=['Style', 'GRN Date'])
                            all_grn_frames.append(df_g)
                        except Exception:
                            pass
                    if all_grn_frames:
                        df_all_grn = pd.concat(all_grn_frames, ignore_index=True)
                        df_all_grn['Style'] = df_all_grn['Style'].astype(str).str.strip().str.upper().str.split('/').str[0].str.strip()
                        df_all_grn = df_all_grn.dropna(subset=['Style', 'GRN Date'])
                        df_all_grn = df_all_grn[df_all_grn['Style'].str.lower() != 'nan']
                        # Keep earliest GRN Date per style (string min works for YYYY-MM-DD format)
                        grn_date_map = df_all_grn.groupby('Style')['GRN Date'].min().to_dict()
                _state['grn_date_map'] = grn_date_map
            except Exception as e:
                print("Error loading GRN dates:", e)

        def get_sort_date(val):
            val_str = str(val).strip().lower()
            if val_str == 'grn pending':
                return pd.Timestamp('2099-12-31')
            if 'existing' in val_str or 'pre' in val_str:
                return pd.Timestamp('2025-01-01')
            try:
                return pd.to_datetime(val_str, format='%b-%y')
            except:
                try:
                    return pd.to_datetime(val_str)
                except:
                    return pd.Timestamp('2099-12-30')

        df['_sort_date'] = df['First Launch Month'].apply(get_sort_date)
        df.sort_values(by=['_sort_date', 'STYLE'], ascending=[True, True], inplace=True)
        df.drop(columns=['_sort_date'], inplace=True)

        # Identify monthly sales columns
        month_cols = []
        for col in df.columns:
            if isinstance(col, (pd.Timestamp, datetime.date, datetime.datetime)):
                month_cols.append(col)
            elif str(col).strip() not in ('STYLE', 'First Launch Month', 'New Style? (Apr25-Jun26)', 'STYLE_LOWER'):
                try:
                    pd.to_datetime(str(col))
                    month_cols.append(col)
                except:
                    pass
        month_cols.sort(key=lambda x: pd.to_datetime(x))

        records = []
        for _, r in df.iterrows():
            style_val = str(r['STYLE']).strip()
            style_upper = style_val.upper()
            
            store_stock = curr_stock_map.get(style_upper, 0)
            transit_stock = tran_stock_map.get(style_upper, 0)
            allocation_stock = alloc_map.get(style_upper, 0)
            lrn_date = lrn_date_map.get(style_upper, "-")
            grn_date = grn_date_map.get(style_upper, "-")
            
            nod = "-"
            if lrn_date != "-" and grn_date != "-":
                try:
                    l = pd.to_datetime(lrn_date, dayfirst=True)
                    g = pd.to_datetime(grn_date)
                    nod = str(abs((l - g).days))
                except:
                    pass
            
            d2c_stock = d2c_map.get(style_upper, 0)
            cred_stock = cred_map.get(style_upper, 0)
            ebo_stock = ebo_map.get(style_upper, 0)
            common_pool_stock = common_pool_map.get(style_upper, 0)
            
            p_data = pipeline_map.get(style_upper, {
                "d2c_jfm": 0, "ebo_jfm": 0,
                "d2c_amj": 0, "ebo_amj": 0,
                "total_d2c": 0, "total_ebo": 0,
                "grn_pending_qty": 0
            })
            
            d2c_jfm = p_data["d2c_jfm"]
            ebo_jfm = p_data["ebo_jfm"]
            d2c_amj = p_data["d2c_amj"]
            ebo_amj = p_data["ebo_amj"]
            d2c_tot = p_data["total_d2c"]
            ebo_tot = p_data["total_ebo"]
            tot_tot = d2c_tot + ebo_tot
            
            monthly_sales = {}
            for col in month_cols:
                try:
                    dt = pd.to_datetime(col)
                    m_str = dt.strftime('%b-%y')
                except:
                    m_str = str(col)
                val = pd.to_numeric(r[col], errors='coerce')
                monthly_sales[m_str] = int(val) if not pd.isna(val) else 0
            
            records.append({
                'STYLE': style_val,
                'JFM_D2C': d2c_jfm,
                'JFM_EBO': ebo_jfm,
                'AMJ_D2C': d2c_amj,
                'AMJ_EBO': ebo_amj,
                'Total_D2C': d2c_tot,
                'Total_EBO': ebo_tot,
                'Total_Sale_Qty': tot_tot,
                'First Launch Month': str(r['First Launch Month']),
                'Stock_Transit': f"{store_stock}/{transit_stock}",
                'Store_Stock': store_stock,
                'Transit_Stock': transit_stock,
                'Allocation_Stock': allocation_stock,
                'LRN_Date': lrn_date,
                'GRN_Date': grn_date,
                'NOD': nod,
                'D2C_Stock': d2c_stock,
                'CRED_Stock': cred_stock,
                'EBO_Stock': ebo_stock,
                'Common_Pool_Stock': common_pool_stock,
                'GRN_Pending_Qty': p_data.get("grn_pending_qty", 0),
                'Is_Live': style_upper in live_styles,
                'Monthly_Sales': monthly_sales
            })
        
        _state['new_style_data'] = records
        _state['new_style_filename'] = os.path.basename(latest_file)
        _state['curr_stock_map'] = curr_stock_map
        _state['tran_stock_map'] = tran_stock_map
        _state['d2c_map'] = d2c_map
        _state['cred_map'] = cred_map
        _state['ebo_map'] = ebo_map
        _state['common_pool_map'] = common_pool_map
        _state['gt_pool_map'] = gt_pool_map
        _state['pipeline_map'] = pipeline_map
        _state['alloc_map'] = alloc_map
        _state['lrn_date_map'] = lrn_date_map
        _state['grn_date_map'] = grn_date_map
        return True, None
    except Exception as e:
        return False, str(e)


def _load_top100_data_internal(force=False):
    if not force and _state.get('top100_data'):
        return True, None

    import glob
    output_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\output of top 100"
    latest_file = None
    if os.path.exists(output_dir):
        files = glob.glob(os.path.join(output_dir, "Top_*_Sales_Report_*.xlsx"))
        files = [f for f in files if not os.path.basename(f).startswith('~') and 'LATEST' not in os.path.basename(f)]
        if files:
            latest_file = max(files)
            
    if not latest_file:
        latest_file = os.path.join(output_dir, "Top_100_Sales_Report_LATEST.xlsx")
        
    if not os.path.exists(latest_file):
        return False, "No pre-generated Top 100 Sales Report found in 'output of top 100'. Please run the top 100 sales script first."

    try:
        df_summary = pd.read_excel(latest_file, sheet_name='Summary')
        df_summary.columns = [str(c).strip() for c in df_summary.columns]
        
        final_top100 = []
        for _, row in df_summary.iterrows():
            sales_10_val = row.get('Last 10 Days Sales', 0)
            qty_10_val = row.get('Last 10 Days Qty', 0)
            sales_30_val = row.get('Last 30 Days Sales', 0)
            qty_30_val = row.get('Last 30 Days Qty', 0)
            
            final_top100.append({
                'style_no': str(row.get('Style', '')).strip().upper(),
                'sales_10': int(sales_10_val) if pd.notna(sales_10_val) else 0,
                'qty_10': int(qty_10_val) if pd.notna(qty_10_val) else 0,
                'sales_30': int(sales_30_val) if pd.notna(sales_30_val) else 0,
                'qty_30': int(qty_30_val) if pd.notna(qty_30_val) else 0,
                'drr': float(row.get('DRR', 0.0)) if pd.notna(row.get('DRR')) else 0.0,
                'doh': float(row.get('DOH', 0.0)) if pd.notna(row.get('DOH')) else 0.0,
                'valid_options_avg': float(row.get('Valid Options', 0.0)),
                'stores_live': int(row.get('No. of Stores Live', 0)),
                'd2c_pool': int(row.get('D2C Pool', 0)),
                'cred_pool': int(row.get('CRED Pool', 0)),
                'ebo_pool': int(row.get('EBO Pool', 0)),
                'common_pool': int(row.get('Common Pool', 0)),
                'gt_pool': int(row.get('GT Pool', 0))
            })
            
        _state['top100_data'] = final_top100
        _state['top100_filename'] = os.path.basename(latest_file)
        return True, None
    except Exception as e:
        print("Error loading/parsing top 100 sales data from generated report:", e)
        return False, str(e)


def _load_campaign_data_internal(force=False):
    campaign_file = r"D:\INCREFF ORDER PUNCH\campaign\STYLE MASTER.xlsx"
    if not os.path.exists(campaign_file):
        return False, "Campaign master file 'STYLE MASTER.xlsx' does not exist."
    
    try:
        df_campaign = _read_excel_fast(campaign_file)
        if df_campaign.empty:
            return False, "Campaign master file is empty."
        
        # Standardise column names
        df_campaign.columns = [str(c).strip().upper() for c in df_campaign.columns]
        if 'STYLE' not in df_campaign.columns or 'CAMPAIGN' not in df_campaign.columns:
            return False, f"STYLE MASTER.xlsx must contain STYLE and Campaign columns. Found: {list(df_campaign.columns)}"
        
        df_campaign['STYLE'] = df_campaign['STYLE'].astype(str).str.strip().str.upper()
        df_campaign['CAMPAIGN'] = df_campaign['CAMPAIGN'].astype(str).str.strip()
        
        # Ensure new style performance data maps are loaded in _state
        df_curr = _state.get('df_ebo_curr')
        if force or not _state.get('new_style_data') or df_curr is None:
            _load_ebo_data_internal()
            _load_new_style_data_internal()
            
        curr_stock_map = _state.get('curr_stock_map', {})
        tran_stock_map = _state.get('tran_stock_map', {})
        d2c_map = _state.get('d2c_map', {})
        cred_map = _state.get('cred_map', {})
        ebo_map = _state.get('ebo_map', {})
        common_pool_map = _state.get('common_pool_map', {})
        pipeline_map = _state.get('pipeline_map', {})
        alloc_map = _state.get('alloc_map', {})
        lrn_date_map = _state.get('lrn_date_map', {})
        grn_date_map = _state.get('grn_date_map', {})
        
        campaign_styles = {}
        campaign_name_map = {} # lowercase -> preferred case
        
        for _, row in df_campaign.iterrows():
            style_code = row['STYLE']
            camp = row['CAMPAIGN']
            if style_code and camp and camp.lower() != 'nan' and camp != '':
                camp_lower = camp.lower()
                if camp_lower not in campaign_name_map:
                    campaign_name_map[camp_lower] = camp
                preferred_camp = campaign_name_map[camp_lower]
                
                style_upper = style_code.upper()
                
                # Fetch metrics exactly like New Style Performance does
                store_stock = curr_stock_map.get(style_upper, 0)
                transit_stock = tran_stock_map.get(style_upper, 0)
                allocation_stock = alloc_map.get(style_upper, 0)
                lrn_date = lrn_date_map.get(style_upper, "-")
                grn_date = grn_date_map.get(style_upper, "-")
                
                nod = "-"
                if lrn_date != "-" and grn_date != "-":
                    try:
                        l = pd.to_datetime(lrn_date, dayfirst=True)
                        g = pd.to_datetime(grn_date)
                        nod = str(abs((l - g).days))
                    except:
                        pass

                d2c_stock = d2c_map.get(style_upper, 0)
                cred_stock = cred_map.get(style_upper, 0)
                ebo_stock = ebo_map.get(style_upper, 0)
                common_pool_stock = common_pool_map.get(style_upper, 0)
                
                p_data = pipeline_map.get(style_upper, {
                    "d2c_jfm": 0, "ebo_jfm": 0,
                    "d2c_amj": 0, "ebo_amj": 0,
                    "total_d2c": 0, "total_ebo": 0
                })
                
                d2c_jfm = p_data["d2c_jfm"]
                ebo_jfm = p_data["ebo_jfm"]
                d2c_amj = p_data["d2c_amj"]
                ebo_amj = p_data["ebo_amj"]
                d2c_tot = p_data["total_d2c"]
                ebo_tot = p_data["total_ebo"]
                tot_tot = d2c_tot + ebo_tot
                
                if preferred_camp not in campaign_styles:
                    campaign_styles[preferred_camp] = []
                    
                campaign_styles[preferred_camp].append({
                    'STYLE': style_code,
                    'JFM_D2C': d2c_jfm,
                    'JFM_EBO': ebo_jfm,
                    'AMJ_D2C': d2c_amj,
                    'AMJ_EBO': ebo_amj,
                    'Total_D2C': d2c_tot,
                    'Total_EBO': ebo_tot,
                    'Total_Sale_Qty': tot_tot,
                    'Stock_Transit': f"{store_stock}/{transit_stock}",
                    'Store_Stock': store_stock,
                    'Transit_Stock': transit_stock,
                    'Allocation_Stock': allocation_stock,
                    'LRN_Date': lrn_date,
                    'GRN_Date': grn_date,
                    'NOD': nod,
                    'D2C_Stock': d2c_stock,
                    'CRED_Stock': cred_stock,
                    'EBO_Stock': ebo_stock,
                    'Common_Pool_Stock': common_pool_stock
                })
                
        campaigns = []
        for camp, styles in campaign_styles.items():
            campaigns.append({
                "campaign_name": camp,
                "style_count": len(styles)
            })
            
        campaigns.sort(key=lambda x: x['campaign_name'].lower())
        
        _state['campaign_data'] = {
            "campaigns": campaigns,
            "campaign_styles": campaign_styles
        }
        return True, None
    except Exception as e:
        return False, str(e)


def _load_style_images():
    import urllib.request
    import io
    import pandas as pd
    url = 'https://docs.google.com/spreadsheets/d/1hULkjjrvDEEbp7sX1FR4502QQv2bHZ72U3G2LhQ91dc/export?format=csv&gid=0'
    try:
        with urllib.request.urlopen(url, timeout=10.0) as response:
            content = response.read()
        df = pd.read_csv(io.BytesIO(content))
        df['STYLE'] = df['STYLE'].astype(str).str.strip().str.upper()
        df['IMAGE URL'] = df['IMAGE URL'].astype(str).str.strip()
        
        # Build dictionary
        mapping = {}
        for _, row in df.iterrows():
            style = row['STYLE']
            img_url = row['IMAGE URL']
            if style and img_url and img_url.lower() != 'nan' and img_url != '':
                mapping[style] = img_url
        
        _state['style_images'] = mapping
        return True, None
    except Exception as e:
        return False, str(e)


def _start_background_loading():
    print("Starting background load of EBO Stock Track Data...")
    try:
        ok, errors = _load_ebo_data_internal()
        if ok:
            print("Background load of EBO Stock Track Data complete!")
        else:
            print("Background load of EBO Stock Track Data failed:", errors)
    except Exception as e:
        print("Background load of EBO Stock Track Data crashed:", e)

    print("Starting background load of New Style Performance Data...")
    try:
        ok, error = _load_new_style_data_internal()
        if ok:
            print("Background load of New Style Performance Data complete!")
        else:
            print("Background load of New Style Performance Data failed:", error)
    except Exception as e:
        print("Background load of New Style Performance Data crashed:", e)

    print("Starting background load of Style Images...")
    try:
        ok, error = _load_style_images()
        if ok:
            print("Background load of Style Images complete!")
        else:
            print("Background load of Style Images failed:", error)
    except Exception as e:
        print("Background load of Style Images crashed:", e)

    print("Starting background load of Campaign Data...")
    try:
        ok, error = _load_campaign_data_internal()
        if ok:
            print("Background load of Campaign Data complete!")
        else:
            print("Background load of Campaign Data failed:", error)
    except Exception as e:
        print("Background load of Campaign Data crashed:", e)

    print("Starting background load of Top 100 Styles Data...")
    try:
        ok, error = _load_top100_data_internal()
        if ok:
            print("Background load of Top 100 Styles Data complete!")
        else:
            print("Background load of Top 100 Styles Data failed:", error)
    except Exception as e:
        print("Background load of Top 100 Styles Data crashed:", e)

# All data is loaded on-demand to prevent OOM crash on Render 512MB free plan.
# Click the "Reload Data" button in each tab to load data after app starts.
# threading.Thread(target=_start_background_loading, daemon=True).start()

# ── AG Validation Loading ────────────────────────────────────────────────────
AG_VALIDATION_DIR = os.path.join(os.path.dirname(__file__), "AG VALIDATION DATA", "AG WORKING")

def _load_ag_data_internal():
    import glob
    
    # Get allocation excel for "Sum of Final Options"
    alloc_dir = os.path.join(os.path.dirname(__file__), "AG VALIDATION DATA", "AG WORKING")
    path_alloc = None
    if os.path.exists(alloc_dir):
        alloc_files = glob.glob(os.path.join(alloc_dir, "*.xlsx"))
        alloc_files = [f for f in alloc_files if not os.path.basename(f).startswith('~')]
        if alloc_files:
            path_alloc = max(alloc_files, key=os.path.getmtime)
    
    valid_style_dir = os.path.join(os.path.dirname(__file__), "VALID STYLE OUTPUT")
    latest_file_path = os.path.join(valid_style_dir, "AG_Validation_Output_v2_LATEST.xlsx")
    if os.path.exists(latest_file_path):
        path_new = latest_file_path
    else:
        files_new = glob.glob(os.path.join(valid_style_dir, "*.xlsx"))
        files_new = [f for f in files_new if not os.path.basename(f).startswith('~')]
        path_new = max(files_new, key=os.path.getmtime) if files_new else None
    
    if not path_alloc or not path_new:
        err_msg = f"Missing required files. path_alloc={path_alloc}, path_new={path_new}"
        print(err_msg)
        _state["ag_status"] = {"loaded": False, "error": err_msg, "last_loaded": None}
        return False, err_msg
        
    try:
        # Load CM mapping
        cm_file = os.path.join(os.path.dirname(__file__), "AG VALIDATION DATA", "CM", "Name_Zones (2).xlsx")
        cm_dict = {}
        loaded_cm = False
        if os.path.exists(cm_file):
            try:
                df_cm = _read_excel_fast(cm_file, sheet_name="Sheet1")
                df_cm['Store Code'] = df_cm['Store Code'].apply(clean_code)
                df_cm['CM'] = df_cm['CM'].astype(str).str.strip()
                for _, r in df_cm.iterrows():
                    sc = r['Store Code']
                    cm_val = r['CM']
                    if pd.notna(cm_val) and str(cm_val).strip().lower() != 'nan':
                        cm_dict[sc] = str(cm_val).strip()
                loaded_cm = True
            except Exception as e:
                print(f"Error loading CM mapping file in app: {e}")
                
        if not loaded_cm and path_new and os.path.exists(path_new):
            try:
                df_cm = _read_excel_fast(path_new, sheet_name="CM")
                df_cm['Store Code'] = df_cm['Store Code'].apply(clean_code)
                df_cm['CM'] = df_cm['CM'].astype(str).str.strip()
                for _, r in df_cm.iterrows():
                    sc = r['Store Code']
                    cm_val = r['CM']
                    if pd.notna(cm_val) and str(cm_val).strip().lower() != 'nan':
                        cm_dict[sc] = str(cm_val).strip()
                print("Loaded CM mapping from validation output Excel sheet 'CM'.")
            except Exception as e:
                print(f"Could not load CM sheet from validation output: {e}")

        # ----- LOAD ALLOCATION FILE (AG Working) -----
        df_ag_working = _read_excel_fast(path_alloc, sheet_name="AG Working")
        
        # Ensure correct types
        df_ag_working['Store Code'] = df_ag_working['Store Code'].apply(clean_code)
        # Support both 'AG' and 'AG Name' column headers
        ag_col = 'AG' if 'AG' in df_ag_working.columns else 'AG Name'
        df_ag_working[ag_col] = df_ag_working[ag_col].astype(str).str.strip()
        df_ag_working['Final Options'] = pd.to_numeric(df_ag_working['Final Options'], errors='coerce').fillna(0)
        
        # Classify categories
        def get_ag_category(ag_name):
            name = str(ag_name).strip().lower()
            if name.startswith('b'):
                return 'K'
            elif name.startswith('w'):
                return 'W'
            elif name.startswith('m'):
                return 'M'
            return 'M'
            
        df_ag_working['Category'] = df_ag_working[ag_col].apply(get_ag_category)

        # Group by Store Code
        store_name_src = 'EBO NAME' if 'EBO NAME' in df_ag_working.columns else ('Store Name' if 'Store Name' in df_ag_working.columns else df_ag_working.columns[1])
        df_store_alloc = df_ag_working.groupby('Store Code', as_index=False).agg(
            Sum_of_Final_Options=('Final Options', 'sum'),
            Store_Name=(store_name_src, 'first')
        ).rename(columns={'Store Code': 'store_code', 'Store_Name': 'Store Name', 'Sum_of_Final_Options': 'Sum of Final Options'})
        
        # Add M, W, K breakdowns for Sum of Final Options
        for cat in ['M', 'W', 'K']:
            df_cat = df_ag_working[df_ag_working['Category'] == cat]
            if not df_cat.empty:
                df_cat_sum = df_cat.groupby('Store Code', as_index=False).agg(
                    Sum_of_Final_Options_Cat=('Final Options', 'sum')
                ).rename(columns={'Store Code': 'store_code', 'Sum_of_Final_Options_Cat': f'Sum of Final Options - {cat}'})
                df_store_alloc = pd.merge(df_store_alloc, df_cat_sum, on='store_code', how='left').fillna(0)
            else:
                df_store_alloc[f'Sum of Final Options - {cat}'] = 0

        # Group by AG
        df_ag_alloc = df_ag_working.groupby(ag_col, as_index=False).agg(
            Sum_of_Final_Options=('Final Options', 'sum')
        ).rename(columns={ag_col: 'AG Name', 'Sum_of_Final_Options': 'Sum of Final Options'})

        # ----- LOAD NEW FILE -----
        df_store_new = _read_excel_fast(path_new, sheet_name="STORE WISE")
        
        # Remove TOTAL row added by validate_styles formatting
        df_store_new = df_store_new[df_store_new['store_code'] != 'TOTAL']
        
        # Reverse map formatted headers back to internal names
        header_map = {
            'Final Options\nOverall': 'Sum of Final Options',
            'Final Options\nM': 'Sum of Final Options - M',
            'Final Options\nW': 'Sum of Final Options - W',
            'Final Options\nK': 'Sum of Final Options - K',
            'Only Stock\nOverall': 'Only Stock Valid Options',
            'Only Stock\nM': 'Only Stock Valid Options - M',
            'Only Stock\nW': 'Only Stock Valid Options - W',
            'Only Stock\nK': 'Only Stock Valid Options - K',
            'Stock & Transit\nOverall': 'Stock and Transit Valid Options',
            'Stock & Transit\nM': 'Stock and Transit Valid Options - M',
            'Stock & Transit\nW': 'Stock and Transit Valid Options - W',
            'Stock & Transit\nK': 'Stock and Transit Valid Options - K',
            'All Valid\nOverall': 'All Valid Options',
            'All Valid\nM': 'All Valid Options - M',
            'All Valid\nW': 'All Valid Options - W',
            'All Valid\nK': 'All Valid Options - K',
        }
        df_store_new = df_store_new.rename(columns=header_map)
        
        # Strip percentages and convert valid options back to int
        for col in header_map.values():
            if col in df_store_new.columns and 'Sum of Final' not in col:
                df_store_new[col] = df_store_new[col].astype(str).str.split('\n').str[0].str.strip()
                df_store_new[col] = pd.to_numeric(df_store_new[col], errors='coerce').fillna(0).astype(int)
        
        df_store_new['store_code'] = df_store_new['store_code'].apply(clean_code)
        
        df_ag_new = _read_excel_fast(path_new, sheet_name="AG WISE")
        if "Row Labels" in df_ag_new.columns:
            df_ag_new = df_ag_new.rename(columns={"Row Labels": "AG Name"})
        df_ag_new['AG Name'] = df_ag_new['AG Name'].astype(str).str.strip()
        
        try:
            df_style_new = _read_excel_fast(path_new, sheet_name="STYLE WISE")
            # Always apply CM mapping from cm_dict (authoritative source),
            # overwriting any stale CM column that may be in the Excel file.
            df_style_new['CM'] = df_style_new['store_code'].apply(clean_code).map(cm_dict).fillna('Unmapped')
            # Reorder so CM appears right after Store Name
            style_data_cols = df_style_new.columns.tolist()
            non_cm_cols = [c for c in style_data_cols if c not in ['CM']]
            # Insert CM after 'Store Name' if present, else after 'store_code'
            insert_after = 'Store Name' if 'Store Name' in non_cm_cols else 'store_code'
            insert_pos = non_cm_cols.index(insert_after) + 1
            ordered_cols = non_cm_cols[:insert_pos] + ['CM'] + non_cm_cols[insert_pos:]
            df_style_new = df_style_new[ordered_cols]
            _state["ag_style_data"] = df_style_new.to_dict(orient="records")
            _state["ag_style_cols"] = df_style_new.columns.tolist()
        except Exception as e:
            _state["ag_style_data"] = []
            _state["ag_style_cols"] = []
            print("Could not load STYLE WISE:", e)

        # ----- MERGE -----
        # Deduplicate columns to prevent pandas suffixing (_x, _y)
        dup_store_cols = [c for c in df_store_alloc.columns if c in df_store_new.columns and c not in ['store_code', 'Store Name']]
        if dup_store_cols:
            df_store_new = df_store_new.drop(columns=dup_store_cols)
            
        dup_ag_cols = [c for c in df_ag_alloc.columns if c in df_ag_new.columns and c != 'AG Name']
        if dup_ag_cols:
            df_ag_new = df_ag_new.drop(columns=dup_ag_cols)

        df_store_merged = pd.merge(df_store_new, df_store_alloc.drop(columns=['Store Name'], errors='ignore'), on='store_code', how='outer').fillna(0)
        df_ag_merged = pd.merge(df_ag_new, df_ag_alloc, on='AG Name', how='outer').fillna(0)
        
        # We need Store Name for outer joined rows that might only be in new file
        if 'Store Name' in df_store_new.columns:
            store_name_map = df_store_new.set_index('store_code')['Store Name'].to_dict()
            df_store_merged['Store Name'] = df_store_merged.apply(
                lambda r: r['Store Name'] if r['Store Name'] != 0 and pd.notnull(r['Store Name']) else store_name_map.get(r['store_code'], r['store_code']), axis=1
            )

        # Add CM to merged store data
        df_store_merged['CM'] = df_store_merged['store_code'].map(cm_dict).fillna('Unmapped')

        # Sort by store wise priority
        priority_dir = os.path.join(os.path.dirname(__file__), "ag validation priority")
        store_pri_file = os.path.join(priority_dir, "Store wise May sales qty.xlsx")
        store_pri_dict = {}
        if os.path.exists(store_pri_file):
            try:
                df_store_pri = _read_excel_fast(store_pri_file)
                df_store_pri['Store Code'] = df_store_pri['Store Code'].apply(clean_code)
                store_pri_dict = dict(zip(df_store_pri['Store Code'], df_store_pri['priority']))
            except Exception as e:
                print(f"Error loading store priority in app: {e}")
                
        def sort_df_by_store_priority(df, store_code_col='store_code'):
            if df.empty:
                return df
            df = df.copy()
            clean_codes = df[store_code_col].apply(clean_code)
            df['_store_pri'] = clean_codes.map(store_pri_dict).fillna(999999)
            df_sorted = df.sort_values(by=['_store_pri', store_code_col]).drop(columns=['_store_pri'])
            return df_sorted
            
        df_store_merged = sort_df_by_store_priority(df_store_merged, 'store_code')
            
        # Reorder columns explicitly to fix UI order
        store_cols = [
            'store_code', 'Store Name', 'Sum of Final Options',
            'Sum of Final Options - M', 'Sum of Final Options - W', 'Sum of Final Options - K',
            'Only Stock Valid Options', 'Only Stock Valid Options - M', 'Only Stock Valid Options - W', 'Only Stock Valid Options - K',
            'Stock and Transit Valid Options', 'Stock and Transit Valid Options - M', 'Stock and Transit Valid Options - W', 'Stock and Transit Valid Options - K',
            'All Valid Options', 'All Valid Options - M', 'All Valid Options - W', 'All Valid Options - K'
        ]
        ag_cols = ['AG Name', 'Sum of Final Options', 'Only Stock Valid Options', 'Stock and Transit Valid Options', 'All Valid Options']
        
        # Ensure all columns exist before selecting
        for col in store_cols:
            if col not in df_store_merged.columns: df_store_merged[col] = 0
        for col in ag_cols:
            if col not in df_ag_merged.columns: df_ag_merged[col] = 0
            
        cols_to_select = store_cols + (['CM'] if 'CM' in df_store_merged.columns else [])
        df_store_merged = df_store_merged[cols_to_select]
        df_ag_merged = df_ag_merged[ag_cols]
        
        # Load raw intersection data for filtering
        try:
            df_raw = _read_excel_fast(path_new, sheet_name="STORE AG RAW")
            df_raw['store_code'] = df_raw['store_code'].apply(clean_code)
            df_raw['AG Name'] = df_raw['AG Name'].astype(str).str.strip()
            # Merge with alloc raw for final options
            df_raw = pd.merge(df_raw, df_ag_working.groupby(['Store Code', ag_col], as_index=False).agg(
                Sum_of_Final_Options=('Final Options', 'sum')
            ).rename(columns={'Store Code': 'store_code', ag_col: 'AG Name', 'Sum_of_Final_Options': 'Sum of Final Options'}), on=['store_code', 'AG Name'], how='outer').fillna(0)
            
            # Use original store name mapping if missing
            if 'Store Name' in df_store_new.columns:
                store_name_map = df_store_new.set_index('store_code')['Store Name'].to_dict()
                df_raw['Store Name'] = df_raw.apply(
                    lambda r: r['Store Name'] if r['Store Name'] != 0 and pd.notnull(r['Store Name']) else store_name_map.get(r['store_code'], r['store_code']), axis=1
                )
                
            df_raw['CM'] = df_raw['store_code'].map(cm_dict).fillna('Unmapped')
            df_raw = sort_df_by_store_priority(df_raw, 'store_code')
            _state["ag_raw_data"] = df_raw.to_dict(orient="records")
        except Exception as e:
            _state["ag_raw_data"] = []
            print("Could not load STORE AG RAW:", e)
            
        _state["ag_store_data"] = df_store_merged.to_dict(orient="records")
        _state["ag_wise_data"] = df_ag_merged.to_dict(orient="records")
        
        import datetime
        _state["ag_status"] = {
            "loaded": True, 
            "error": None, 
            "last_loaded": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        return True, None
    except Exception as e:
        _state["ag_status"] = {"loaded": False, "error": str(e), "last_loaded": None}
        return False, str(e)

def _start_ag_loading():
    import time
    time.sleep(5)
    print("Starting background load of AG Validation Data...")
    try:
        ok, err = _load_ag_data_internal()
        if ok:
            print("Background load of AG Validation Data complete!")
        else:
            print("AG Validation load failed:", err)
    except Exception as e:
        print("AG Validation load crashed:", e)

# AG Validation data is loaded on-demand via the Reload Data button
# (disabled auto-start to prevent OOM on Render 512MB free plan)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _load_priority(path):
    try:
        df = pd.read_excel(path, sheet_name="Export")
    except Exception:
        df = pd.read_excel(path, sheet_name=0)

    # Normalize priority columns
    priority_col_mapping = {
        "Store Name": ["Store Name", "store name", "STORE NAME", "StoreName", "store_name"],
        PRIORITY_NUM_COL: ["priority number", "priority", "Priority", "Priority No", "Priority Number", "priority_number", "priority list"]
    }
    rename_priority = {}
    for standard_col, options in priority_col_mapping.items():
        for col in df.columns:
            if col.strip().lower() == standard_col.lower() or col.strip() in options:
                rename_priority[col] = standard_col
                break
    df = df.rename(columns=rename_priority)

    if PRIORITY_NUM_COL in df.columns:
        df = df.sort_values(PRIORITY_NUM_COL).reset_index(drop=True)
    return df


def _simulate_full(sku, df_replen, df_priority):
    """
    Returns list of dicts — full allocation trace (every EBO, not just shortfall).
    """
    sku_rows = df_replen[df_replen[SKU_COL] == sku].copy()
    if sku_rows.empty:
        return [], 0, 0

    ebo_need = (
        sku_rows.groupby(EBO_COL, as_index=False)
        .agg(needed=(ALLOC_COL, "sum"), avail=(AVAIL_COL, "first"))
    )

    total_stock = int(ebo_need["avail"].iloc[0])
    total_need  = int(ebo_need["needed"].sum())
    remaining   = total_stock
    need_lookup = dict(zip(ebo_need[EBO_COL], ebo_need["needed"]))

    rows      = []
    seen_ebos = set()

    for _, row in df_priority.iterrows():
        priority_name = row["Store Name"]
        pri_num       = int(row[PRIORITY_NUM_COL]) if PRIORITY_NUM_COL in row.index else 0
        ebo_name      = PRIORITY_TO_EBO.get(priority_name)
        if not ebo_name:
            if priority_name in PRIORITY_TO_EBO.values():
                ebo_name = priority_name
            else:
                for idx in ebo_need[EBO_COL].unique():
                    if str(idx).strip().lower() == str(priority_name).strip().lower():
                        ebo_name = idx
                        break
        if not ebo_name:
            ebo_name = priority_name  # fallback

        if ebo_name is None or ebo_name not in need_lookup:
            continue
        if ebo_name in seen_ebos:
            continue
        seen_ebos.add(ebo_name)

        needed = int(need_lookup[ebo_name])
        if needed <= 0:
            continue

        if remaining <= 0:
            given, shortfall, status = 0, needed, "No Stock"
        elif remaining >= needed:
            given = needed; remaining -= needed; shortfall = 0; status = "Fully Met"
        else:
            given = remaining; shortfall = needed - given; remaining = 0; status = "Partial"

        rows.append({
            "priority"  : pri_num,
            "ebo"       : ebo_name,
            "needed"    : needed,
            "given"     : given,
            "shortfall" : shortfall,
            "status"    : status,
        })

    # Catch any EBOs that need stock but are missing from the priority list
    for ebo_name, needed in need_lookup.items():
        if ebo_name not in seen_ebos:
            needed = int(needed)
            if needed <= 0:
                continue
            given, shortfall, status = 0, needed, "Unmapped"
            rows.append({
                "priority"  : "-",
                "ebo"       : ebo_name,
                "needed"    : needed,
                "given"     : given,
                "shortfall" : shortfall,
                "status"    : status,
            })

    return rows, total_stock, total_need


def _run_all_allocation(df_replen, df_priority):
    """Batch run — returns master DataFrame."""
    pivot_need = (
        df_replen.groupby([EBO_COL, SKU_COL], as_index=False)
        .agg(needed=(ALLOC_COL, "sum"), avail=(AVAIL_COL, "first"))
        .pivot(index=EBO_COL, columns=SKU_COL, values="needed")
        .fillna(0)
    )
    avail_per_sku = df_replen.groupby(SKU_COL)[AVAIL_COL].first()

    seen_ebos    = set()
    ordered_ebos = []
    for _, row in df_priority.iterrows():
        priority_name = row["Store Name"]
        ebo_name = PRIORITY_TO_EBO.get(priority_name)
        if not ebo_name:
            if priority_name in PRIORITY_TO_EBO.values():
                ebo_name = priority_name
            else:
                for idx in pivot_need.index:
                    if str(idx).strip().lower() == str(priority_name).strip().lower():
                        ebo_name = idx
                        break
        if not ebo_name:
            ebo_name = priority_name  # fallback

        if ebo_name is None or ebo_name not in pivot_need.index:
            continue
        if ebo_name in seen_ebos:
            continue
        seen_ebos.add(ebo_name)
        
        pri_val = row[PRIORITY_NUM_COL]
        try:
            pri_num = int(pri_val)
        except:
            pri_num = 999
        ordered_ebos.append((pri_num, ebo_name))

    all_skus = sorted(pivot_need.columns)
    all_rows = []

    for sku in all_skus:
        total_stock = int(avail_per_sku.get(sku, 0))
        remaining   = total_stock

        for pri_num, ebo_name in ordered_ebos:
            needed = int(pivot_need.at[ebo_name, sku]) if ebo_name in pivot_need.index else 0
            if needed == 0:
                continue

            if remaining <= 0:
                given, shortfall = 0, needed
            elif remaining >= needed:
                given = needed; remaining -= needed; shortfall = 0
            else:
                given = remaining; shortfall = needed - given; remaining = 0

            if shortfall > 0:
                all_rows.append({
                    "SKU"          : sku,
                    "Priority No"  : pri_num,
                    "EBO Name"     : ebo_name,
                    "Needed Qty"   : needed,
                    "Allocated Qty": given,
                    "Shortfall Qty": shortfall,
                })

        # Second pass: EBOs missing from priority mapping
        for ebo_name in pivot_need.index:
            if ebo_name not in seen_ebos:
                needed = int(pivot_need.at[ebo_name, sku])
                if needed > 0:
                    given, shortfall = 0, needed
                    all_rows.append({
                        "SKU"          : sku,
                        "Priority No"  : "-",
                        "EBO Name"     : ebo_name,
                        "Needed Qty"   : needed,
                        "Allocated Qty": given,
                        "Shortfall Qty": shortfall,
                    })

    if not all_rows:
        return pd.DataFrame(columns=["SKU", "Priority No", "EBO Name", "Needed Qty", "Allocated Qty", "Shortfall Qty"])
    return pd.DataFrame(all_rows)


def _save_all_excel(master_df):
    from openpyxl.utils import get_column_letter
    import time
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"ALL_SKU_shortfall_report_{timestamp}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        master_df.to_excel(writer, index=False, sheet_name="All Shortfalls", startrow=2)
        ws_all = writer.sheets["All Shortfalls"]
        for i, w in enumerate([18, 34, 14, 16, 16, 16], start=1):
            ws_all.column_dimensions[get_column_letter(i)].width = w
        _apply_sheet_style(ws_all, master_df,
            "All SKU Allocation Shortfall Report  |  All EBOs  |  All SKUs",
            n_extra_cols=1)

        for ebo in master_df["EBO Name"].unique():
            ebo_df     = master_df[master_df["EBO Name"] == ebo].copy()
            safe_name  = ebo[:31].replace("/", "-").replace("\\", "-").replace("?","").replace("*","").replace("[","").replace("]","")
            ebo_df_out = ebo_df.drop(columns=["EBO Name"]).reset_index(drop=True)
            ebo_df_out.to_excel(writer, index=False, sheet_name=safe_name, startrow=2)
            ws_ebo = writer.sheets[safe_name]
            for i, w in enumerate([18, 14, 16, 16, 16], start=1):
                ws_ebo.column_dimensions[get_column_letter(i)].width = w
            _apply_sheet_style(ws_ebo, ebo_df_out,
                f"Shortfall Report  |  EBO: {ebo}  |  Total Shortfall SKUs: {len(ebo_df_out)}")

    return output_path


def _save_sku_excel(rows, sku, total_stock):
    from openpyxl.utils import get_column_letter
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    shortfall_rows = [r for r in rows if r["shortfall"] > 0]
    if not shortfall_rows:
        df = pd.DataFrame(columns=["Priority No","EBO Name","Needed Qty","Allocated Qty","Shortfall Qty"])
    else:
        df = pd.DataFrame(shortfall_rows)[["priority","ebo","needed","given","shortfall"]]
        df.columns = ["Priority No","EBO Name","Needed Qty","Allocated Qty","Shortfall Qty"]

    safe_sku    = "".join(c if c.isalnum() or c in ("-","_") else "_" for c in sku)
    output_path = os.path.join(OUTPUT_DIR, f"check_allocation_{safe_sku}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Shortfall Report", startrow=2)
        ws = writer.sheets["Shortfall Report"]
        for i, w in enumerate([14, 34, 14, 16, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        _apply_sheet_style(ws, df,
            f"Allocation Shortfall Report  |  SKU: {sku}  |  Available Stock: {total_stock}")
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("alloc_index.html")


@app.route("/api/upload", methods=["POST"])
def upload():
    errors = []
    if "replen_file" in request.files and request.files["replen_file"].filename:
        f = request.files["replen_file"]
        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)
        try:
            try:
                _state["df_replen"]   = _read_excel_fast(path, sheet_name="Sheet1")
            except ValueError:
                _state["df_replen"]   = _read_excel_fast(path, sheet_name=0)

            # Normalize replenishment columns
            replen_col_mapping = {
                EBO_COL: ["EBO NAME", "ebo name", "Ebo Name", "EBO Name"],
                SKU_COL: ["Client SKU Id / EAN", "client sku id / ean", "Client Sku Id / Ean", "Client SKU ID / EAN"],
                ALLOC_COL: ["Allocated Qty", "allocated qty", "Allocated QTY", "Allocated quantity", "allocated quantity", "Allocated  Qty"],
                AVAIL_COL: ["available qty", "Available Qty", "available Qty", "Available QTY", "available quantity", "available stock"]
            }
            rename_replen = {}
            for standard_col, options in replen_col_mapping.items():
                for col in _state["df_replen"].columns:
                    normalized_col = " ".join(col.split()).lower()
                    normalized_std = " ".join(standard_col.split()).lower()
                    normalized_opts = [" ".join(opt.split()).lower() for opt in options]
                    if normalized_col == normalized_std or normalized_col in normalized_opts:
                        rename_replen[col] = standard_col
                        break
            _state["df_replen"] = _state["df_replen"].rename(columns=rename_replen)

            _state["replen_name"] = f.filename
            _state["replen_path"] = path
            _state["all_skus"]    = sorted(_state["df_replen"][SKU_COL].dropna().unique().tolist())
            _state["all_stores"]  = sorted(_state["df_replen"][EBO_COL].dropna().unique().tolist())
        except Exception as e:
            errors.append(f"Replenishment file error: {e}")

    if "priority_file" in request.files and request.files["priority_file"].filename:
        f = request.files["priority_file"]
        path = os.path.join(UPLOAD_FOLDER, secure_filename(f.filename))
        f.save(path)
        try:
            _state["df_priority"]    = _load_priority(path)
            _state["priority_name"]  = f.filename
        except Exception as e:
            errors.append(f"Priority file error: {e}")

    return jsonify({
        "ok"           : len(errors) == 0,
        "errors"       : errors,
        "replen_loaded": _state["replen_name"],
        "priority_loaded": _state["priority_name"],
        "sku_count"    : len(_state["all_skus"]),
        "store_count"  : len(_state["all_stores"]),
    })


@app.route("/api/skus")
def get_skus():
    q = request.args.get("q", "").upper()
    skus = [s for s in _state["all_skus"] if q in str(s).upper()][:50]
    return jsonify(skus)


@app.route("/api/stores")
def get_stores():
    q = request.args.get("q", "").upper()
    stores = [s for s in _state["all_stores"] if q in str(s).upper()][:50]
    return jsonify(stores)


@app.route("/api/run_sku", methods=["POST"])
def run_sku():
    data = request.get_json()
    sku  = (data or {}).get("sku", "").strip()

    if not sku:
        return jsonify({"ok": False, "error": "No SKU provided."})
    if _state["df_replen"] is None or _state["df_priority"] is None:
        return jsonify({"ok": False, "error": "Please upload both files first."})

    rows, total_stock, total_need = _simulate_full(
        sku, _state["df_replen"], _state["df_priority"]
    )

    if not rows:
        return jsonify({"ok": False, "error": f"SKU '{sku}' not found in replenishment data."})

    total_short  = sum(r["shortfall"] for r in rows)
    total_given  = sum(r["given"]     for r in rows)
    short_ebos   = sum(1 for r in rows if r["shortfall"] > 0)

    return jsonify({
        "ok"         : True,
        "sku"        : sku,
        "total_stock": total_stock,
        "total_need" : total_need,
        "total_given": total_given,
        "total_short": total_short,
        "short_ebos" : short_ebos,
        "rows"       : rows,
    })


@app.route("/api/download_sku", methods=["POST"])
def download_sku():
    data = request.get_json()
    sku  = (data or {}).get("sku", "").strip()
    if not sku or _state["df_replen"] is None:
        return jsonify({"ok": False, "error": "Bad request."})

    rows, total_stock, _ = _simulate_full(sku, _state["df_replen"], _state["df_priority"])
    path = _save_sku_excel(rows, sku, total_stock)
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path))


def _save_store_excel(rows, store):
    from openpyxl.utils import get_column_letter
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    if not rows:
        df = pd.DataFrame(columns=["SKU","Needed Qty","Allocated Qty","Shortfall Qty","Status"])
    else:
        df = pd.DataFrame(rows)[["sku","needed","given","shortfall","status"]]
        df.columns = ["SKU","Needed Qty","Allocated Qty","Shortfall Qty","Status"]

    safe_store = "".join(c if c.isalnum() or c in ("-","_") else "_" for c in store)
    output_path = os.path.join(OUTPUT_DIR, f"check_allocation_store_{safe_store}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Store Report", startrow=2)
        ws = writer.sheets["Store Report"]
        for i, w in enumerate([25, 14, 16, 16, 16], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        _apply_sheet_style(ws, df, f"Store Allocation Report  |  EBO: {store}")
    return output_path


@app.route("/api/run_store", methods=["POST"])
def run_store():
    data = request.get_json()
    store = (data or {}).get("store", "").strip()

    if not store: return jsonify({"ok": False, "error": "No Store provided."})
    if _state["df_replen"] is None or _state["df_priority"] is None:
        return jsonify({"ok": False, "error": "Please upload both files first."})

    if store.upper() == "ALL":
        # Run full summary
        summary_rows = _run_store_summary(_state["df_replen"], _state["df_priority"])
        total_need = sum(r["needed"] for r in summary_rows)
        total_given = sum(r["given"] for r in summary_rows)
        total_short = sum(r["shortfall"] for r in summary_rows)

        return jsonify({
            "ok": True,
            "store": "ALL",
            "sku_count": len(summary_rows), # this represents store_count in this context
            "total_need": total_need,
            "total_given": total_given,
            "total_short": total_short,
            "rows": summary_rows
        })


    store_rows = _state["df_replen"][_state["df_replen"][EBO_COL] == store]
    skus_needed = store_rows[SKU_COL].unique()
    store_results = []
    total_need = 0
    total_given = 0
    total_short = 0

    df_priority_sorted = _state["df_priority"].sort_values(by=PRIORITY_NUM_COL)
    store_pri = "-"
    for _, row in df_priority_sorted.iterrows():
        pname = row["Store Name"]
        ename = PRIORITY_TO_EBO.get(pname)
        if not ename:
            if pname == store or pname.strip().lower() == store.strip().lower():
                ename = store
            elif pname in PRIORITY_TO_EBO.values() and pname == store:
                ename = store
        if ename == store:
            store_pri = row[PRIORITY_NUM_COL]
            break

    for s in skus_needed:
        rows, _, _ = _simulate_full(s, _state["df_replen"], _state["df_priority"])
        
        # find the store in rows
        row = next((r for r in rows if r["ebo"] == store), None)
        if row:
            store_results.append({
                "sku": s,
                "needed": row["needed"],
                "given": row["given"],
                "shortfall": row["shortfall"],
                "status": row["status"],
            })
            total_need += row["needed"]
            total_given += row["given"]
            total_short += row["shortfall"]

    # Sort so that items with shortfalls appear at the top
    store_results.sort(key=lambda x: x["shortfall"], reverse=True)

    return jsonify({
        "ok": True,
        "store": store,
        "priority": store_pri,
        "sku_count": len(store_results),
        "total_need": total_need,
        "total_given": total_given,
        "total_short": total_short,
        "rows": store_results
    })


@app.route("/api/download_store", methods=["POST"])
def download_store():
    data = request.get_json()
    store = (data or {}).get("store", "").strip()
    if not store or _state["df_replen"] is None: return jsonify({"ok": False, "error": "Bad request."})

    if store.upper() == "ALL":
        summary_rows = _run_store_summary(_state["df_replen"], _state["df_priority"])
        path = _save_store_summary_excel(summary_rows)
        return send_file(path, as_attachment=True, download_name=os.path.basename(path))

    store_rows = _state["df_replen"][_state["df_replen"][EBO_COL] == store]
    skus_needed = store_rows[SKU_COL].unique()
    store_results = []
    for s in skus_needed:
        rows, _, _ = _simulate_full(s, _state["df_replen"], _state["df_priority"])
        row = next((r for r in rows if r["ebo"] == store), None)
        if row: store_results.append({"sku":s, "needed":row["needed"], "given":row["given"], "shortfall":row["shortfall"], "status":row["status"]})

    # Sort shortfalls to the top in the Excel report too
    store_results.sort(key=lambda x: x["shortfall"], reverse=True)

    path = _save_store_excel(store_results, store)
    return send_file(path, as_attachment=True, download_name=os.path.basename(path))

def _run_store_summary(df_replen, df_priority):
    stock_agg = df_replen.groupby(SKU_COL, as_index=False).agg({AVAIL_COL: "first"})
    avail_per_sku = dict(zip(stock_agg[SKU_COL], stock_agg[AVAIL_COL]))

    pivot_need = df_replen.pivot_table(index=EBO_COL, columns=SKU_COL, values=ALLOC_COL, aggfunc="sum").fillna(0)


    df_priority_sorted = df_priority.sort_values(by=PRIORITY_NUM_COL)
    ordered_ebos = []
    ebo_to_pri = {}
    for _, row in df_priority_sorted.iterrows():
        pname = row["Store Name"]
        ename = PRIORITY_TO_EBO.get(pname)
        if not ename:
            if pname in pivot_need.index:
                ename = pname
            elif pname in PRIORITY_TO_EBO.values():
                ename = pname
            else:
                for idx in pivot_need.index:
                    if str(idx).strip().lower() == str(pname).strip().lower():
                        ename = idx
                        break
        if ename:
            ordered_ebos.append(ename)
            ebo_to_pri[ename] = row[PRIORITY_NUM_COL]

    store_stats = {ebo: {"sku_count": 0, "needed": 0, "given": 0, "shortfall": 0} for ebo in pivot_need.index}

    for sku in pivot_need.columns:
        remaining = int(avail_per_sku.get(sku, 0))
        seen_ebos = set()

        for ebo_name in ordered_ebos:
            if ebo_name not in pivot_need.index: continue
            needed = int(pivot_need.at[ebo_name, sku])
            if needed <= 0: continue
            seen_ebos.add(ebo_name)
            
            if remaining >= needed:
                given = needed; remaining -= needed
            else:
                given = remaining; remaining = 0
            
            shortfall = needed - given
            st = store_stats[ebo_name]
            st["sku_count"] += 1
            st["needed"] += needed
            st["given"] += given
            st["shortfall"] += shortfall

        for ebo_name in pivot_need.index:
            if ebo_name not in seen_ebos:
                needed = int(pivot_need.at[ebo_name, sku])
                if needed > 0:
                    st = store_stats[ebo_name]
                    st["sku_count"] += 1
                    st["needed"] += needed
                    st["given"] += 0
                    st["shortfall"] += needed
            

    summary_rows = []
    for ebo, st in store_stats.items():
        if st["needed"] > 0:
            summary_rows.append({
                "store": ebo, "priority": ebo_to_pri.get(ebo, "-"), "sku_count": st["sku_count"],
                "needed": st["needed"], "given": st["given"], "shortfall": st["shortfall"]
            })

    def _sort_key(x):
        try:
            return float(x["priority"])
        except (ValueError, TypeError):
            return 999999
            
    summary_rows.sort(key=lambda x: (_sort_key(x), -x["shortfall"]))
    return summary_rows

def _save_store_summary_excel(rows):
    from openpyxl.utils import get_column_letter
    import time

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    df = pd.DataFrame(rows)[["priority", "store", "sku_count", "needed", "given", "shortfall"]]
    df.columns = ["Priority", "Store Name", "Total SKUs", "Needed Qty", "Allocated Qty", "Shortfall Qty"]
    
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"ALL_STORES_summary_{timestamp}.xlsx")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All Stores", startrow=2)
        ws = writer.sheets["All Stores"]
        for i, w in enumerate([35, 12, 14, 14, 14], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        _apply_sheet_style(ws, df, "All Stores Summary Report")
    return output_path

# ── All-SKU batch job (SSE-style streaming via polling) ───────────────────────
_batch_job = {"running": False, "progress": 0, "total": 0, "done": False,
              "result": None, "error": None, "output_path": None}

def _batch_worker():
    try:
        master_df = _run_all_allocation(_state["df_replen"], _state["df_priority"])
        path      = _save_all_excel(master_df)
        
        # Run process_replenishment script
        import subprocess
        try:
            print("Running process_replenishment.py...")
            cmd = ["python", "process_replenishment.py"]
            if _state["replen_path"]:
                cmd.append(_state["replen_path"])
            subprocess.run(cmd, check=True)
        except Exception as e:
            print("Error running process_replenishment.py:", e)

        _batch_job.update({
            "done"       : True,
            "running"    : False,
            "result"     : {
                "total_skus"     : len(_state["all_skus"]),
                "skus_with_short": int(master_df["SKU"].nunique()) if not master_df.empty else 0,
                "total_rows"     : len(master_df),
                "ebos_affected"  : int(master_df["EBO Name"].nunique()) if not master_df.empty else 0,
            },
            "output_path": path,
        })
    except Exception as e:
        _batch_job.update({"running": False, "done": True, "error": str(e)})


@app.route("/api/run_all", methods=["POST"])
def run_all():
    if _state["df_replen"] is None or _state["df_priority"] is None:
        return jsonify({"ok": False, "error": "Please upload both files first."})
    if _batch_job["running"]:
        return jsonify({"ok": False, "error": "Batch job already running."})

    _batch_job.update({"running": True, "done": False, "error": None,
                       "result": None, "output_path": None,
                       "total": len(_state["all_skus"]), "progress": 0})
    threading.Thread(target=_batch_worker, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/batch_status")
def batch_status():
    return jsonify({k: v for k, v in _batch_job.items() if k != "output_path"})


@app.route("/api/download_all")
def download_all():
    path = _batch_job.get("output_path")
    if not path or not os.path.exists(path):
        return "No report available", 404
    return send_file(path, as_attachment=True,
                     download_name=os.path.basename(path))


@app.route("/api/status")
def status():
    return jsonify({
        "replen_loaded"  : _state["replen_name"],
        "priority_loaded": _state["priority_name"],
        "sku_count"      : len(_state["all_skus"]),
        "store_count"    : len(_state["all_stores"]),
    })


@app.route("/api/ebo_status")
def ebo_status():
    return jsonify(_state.get("ebo_status", {"loaded": False}))


@app.route("/api/ebo_reload", methods=["POST"])
def ebo_reload():
    ok, errors = _load_ebo_data_internal()
    status_info = _state.get("ebo_status", {"loaded": False})
    return jsonify({
        "ok": ok,
        "errors": errors,
        "status": status_info
    })


@app.route("/api/ebo_summary")
def ebo_summary():
    summary = _state.get("ebo_summary")
    if summary is None:
        return jsonify([])
    return jsonify(summary)

@app.route("/api/ag_data")
def ag_data():
    cache_path = os.path.join(os.path.dirname(__file__), "ag_data_cache.json")
    # Serve from pre-processed JSON cache (near-zero RAM usage)
    if _state.get("ag_store_data") is not None and len(_state.get("ag_store_data", [])) > 0:
        # Already loaded in memory - serve directly
        return jsonify({
            "status": _state.get("ag_status"),
            "store_data": _state.get("ag_store_data", []),
            "ag_data": _state.get("ag_wise_data", []),
            "raw_data": _state.get("ag_raw_data", []),
            "style_data": _state.get("ag_style_data", []),
            "style_cols": _state.get("ag_style_cols", []),
            "store_cols": [
                'store_code', 'Store Name', 'Sum of Final Options',
                'Sum of Final Options - M', 'Sum of Final Options - W', 'Sum of Final Options - K',
                'Only Stock Valid Options', 'Only Stock Valid Options - M', 'Only Stock Valid Options - W', 'Only Stock Valid Options - K',
                'Stock and Transit Valid Options', 'Stock and Transit Valid Options - M', 'Stock and Transit Valid Options - W', 'Stock and Transit Valid Options - K',
                'All Valid Options', 'All Valid Options - M', 'All Valid Options - W', 'All Valid Options - K'
            ],
            "ag_cols": ['AG Name', 'Sum of Final Options', 'Only Stock Valid Options', 'Stock and Transit Valid Options', 'All Valid Options']
        })
    elif os.path.exists(cache_path):
        # Load from pre-processed JSON cache (lightweight, no Excel processing)
        import json
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
        # Store in _state for subsequent requests
        _state["ag_store_data"] = cache.get("store_data", [])
        _state["ag_wise_data"] = cache.get("ag_data", [])
        _state["ag_raw_data"] = cache.get("raw_data", [])
        _state["ag_style_data"] = cache.get("style_data", [])
        _state["ag_style_cols"] = cache.get("style_cols", [])
        _state["ag_status"] = cache.get("status", {"loaded": True, "error": None, "last_loaded": None})
        _state["ag_status"]["loaded"] = True
        return jsonify(cache)
    else:
        return jsonify({
            "status": {"loaded": False, "error": "No cache file found. Run export_ag_cache.py locally and push to GitHub.", "last_loaded": None},
            "store_data": [], "ag_data": [], "raw_data": [], "style_data": [],
            "style_cols": [],
            "store_cols": [
                'store_code', 'Store Name', 'Sum of Final Options',
                'Sum of Final Options - M', 'Sum of Final Options - W', 'Sum of Final Options - K',
                'Only Stock Valid Options', 'Only Stock Valid Options - M', 'Only Stock Valid Options - W', 'Only Stock Valid Options - K',
                'Stock and Transit Valid Options', 'Stock and Transit Valid Options - M', 'Stock and Transit Valid Options - W', 'Stock and Transit Valid Options - K',
                'All Valid Options', 'All Valid Options - M', 'All Valid Options - W', 'All Valid Options - K'
            ],
            "ag_cols": ['AG Name', 'Sum of Final Options', 'Only Stock Valid Options', 'Stock and Transit Valid Options', 'All Valid Options']
        })

@app.route("/api/ag_reload", methods=["POST"])
def ag_reload():
    """Load AG data from the pre-processed JSON cache (lightweight, no Excel processing on server)."""
    cache_path = os.path.join(os.path.dirname(__file__), "ag_data_cache.json")
    if os.path.exists(cache_path):
        import json
        with open(cache_path, "r", encoding="utf-8") as f:
            cache = json.load(f)
        _state["ag_store_data"] = cache.get("store_data", [])
        _state["ag_wise_data"] = cache.get("ag_data", [])
        _state["ag_raw_data"] = cache.get("raw_data", [])
        _state["ag_style_data"] = cache.get("style_data", [])
        _state["ag_style_cols"] = cache.get("style_cols", [])
        _state["ag_status"] = cache.get("status", {"loaded": True, "error": None, "last_loaded": None})
        _state["ag_status"]["loaded"] = True
        return jsonify({"ok": True, "message": f"Loaded {len(_state['ag_store_data'])} stores from cache.", "status": _state["ag_status"]})
    else:
        return jsonify({"ok": False, "error": "No cache file found. Run export_ag_cache.py locally and push to GitHub."})




@app.route("/api/ag_download")
def ag_download():
    import glob
    valid_style_dir = os.path.join(os.path.dirname(__file__), "VALID STYLE OUTPUT")
    files_new = glob.glob(os.path.join(valid_style_dir, "*.xlsx"))
    files_new = [f for f in files_new if not os.path.basename(f).startswith('~')]
    if not files_new:
        return "No AG Validation report available", 404
    path_new = max(files_new, key=os.path.getmtime)
    return send_file(path_new, as_attachment=True, download_name=os.path.basename(path_new))


@app.route("/api/ag_filtered_extract", methods=["POST"])
def ag_filtered_extract():
    """
    Export filtered columns from STORE WISE + STYLE WISE based on selected types.
    POST body: { "types": ["stock", "trans", "allocation"] }  (any subset)
    """
    import glob
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    try:
        payload   = request.get_json(silent=True) or {}
        types_req = [t.lower().strip() for t in payload.get("types", [])]
        if not types_req:
            return jsonify({"error": "No column types selected"}), 400

        want_stock = "stock"  in types_req
        want_trans = "trans"  in types_req
        want_alloc = "allocation" in types_req

        valid_style_dir = os.path.join(os.path.dirname(__file__), "VALID STYLE OUTPUT")
        files_new = glob.glob(os.path.join(valid_style_dir, "*.xlsx"))
        files_new = [f for f in files_new if not os.path.basename(f).startswith('~')]
        if not files_new:
            return "No AG Validation report available", 404
        path_new = max(files_new, key=os.path.getmtime)

        xl = pd.ExcelFile(path_new)
        id_cols = ['store_code', 'Store Name', 'CM']

        # ── STYLE WISE ────────────────────────────────────────────────────────
        # Columns are like OR10_Stock, OR10_Trans, OR10_Allocation
        df_style = xl.parse("STYLE WISE") if "STYLE WISE" in xl.sheet_names else pd.DataFrame()
        if not df_style.empty:
            sel = []
            for c in df_style.columns:
                if c in id_cols:
                    continue
                if want_stock and c.endswith("_Stock"):
                    sel.append(c)
                elif want_trans and c.endswith("_Trans"):
                    sel.append(c)
                elif want_alloc and c.endswith("_Allocation"):
                    sel.append(c)
            keep_style = [c for c in id_cols if c in df_style.columns] + sel
            df_style = df_style[keep_style]

        # ── STORE WISE ────────────────────────────────────────────────────────
        # Columns: "Only Stock\nOverall/M/W/K", "Stock & Transit\nOverall/M/W/K", "All Valid\nOverall/M/W/K"
        # We also keep "Final Options\nOverall/M/W/K" always as context
        df_store = xl.parse("STORE WISE") if "STORE WISE" in xl.sheet_names else pd.DataFrame()
        store_id_cols = ['store_code', 'Store Name']
        if not df_store.empty:
            sel_s = []
            for c in df_store.columns:
                if c in store_id_cols:
                    continue
                cl = str(c).lower()
                # Always include "Final Options" as context column
                if "final options" in cl:
                    sel_s.append(c)
                    continue
                if want_stock and "only stock" in cl:
                    sel_s.append(c)
                if want_trans and "stock & transit" in cl:
                    sel_s.append(c)
                if want_alloc and "all valid" in cl:
                    sel_s.append(c)
            keep_store = [c for c in store_id_cols if c in df_store.columns] + sel_s
            df_store = df_store[keep_store]

        # ── Build Excel output ─────────────────────────────────────────────────
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            sheet_label = " + ".join(t.capitalize() for t in types_req)
            if not df_store.empty:
                df_store.to_excel(writer, sheet_name=f"STORE WISE ({sheet_label})", index=False)
            if not df_style.empty:
                df_style.to_excel(writer, sheet_name=f"STYLE WISE ({sheet_label})", index=False)

            header_fill = PatternFill(start_color="1A4A7A", end_color="1A4A7A", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            center_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for sname in writer.sheets:
                ws = writer.sheets[sname]
                for ci in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=ci)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_aln
                    ws.column_dimensions[get_column_letter(ci)].width = 20
                ws.freeze_panes = "C2"

        output.seek(0)
        ts       = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        label_fs = "_".join(t.capitalize() for t in types_req)
        filename = f"AG_Extract_{label_fs}_{ts}.xlsx"
        return send_file(
            output,
            download_name=filename,
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("ag_filtered_extract error:", e)
        return f"Export failed: {e}", 500


@app.route("/api/ebo_store_detail/<store_code>")
def ebo_store_detail(store_code):
    df_curr = _state.get('df_ebo_curr')
    df_tran = _state.get('df_ebo_tran')
    
    if df_curr is None or df_tran is None:
        return jsonify({"ok": False, "error": "EBO Stock Data not loaded."})
        
    store_curr = df_curr[df_curr['store_code'] == store_code] if not df_curr.empty else pd.DataFrame()
    store_tran = df_tran[df_tran['Store Code'] == store_code] if not df_tran.empty else pd.DataFrame()
    
    if store_curr.empty and store_tran.empty:
        return jsonify({"ok": False, "error": f"No data found for store {store_code}."})
        
    store_name = "Unknown Store"
    if not store_curr.empty:
        store_name = store_curr['Store Name'].iloc[0]
    elif not store_tran.empty:
        store_name = store_tran['EBO Name'].iloc[0]
        
    curr_details = []
    if not store_curr.empty:
        curr_grouped = store_curr.groupby(['Style', 'Colour', 'Size'], as_index=False).agg(qty=('quantity', 'sum'))
        curr_grouped = curr_grouped.sort_values(by=['Style', 'Colour', 'Size'])
        curr_details = curr_grouped.to_dict(orient='records')
        
    tran_details = []
    if not store_tran.empty:
        tran_grouped = store_tran.groupby(['STYLE', 'COLOR', 'SIZE'], as_index=False).agg(qty=('Transit Qty', 'sum'))
        tran_grouped = tran_grouped.sort_values(by=['STYLE', 'COLOR', 'SIZE'])
        tran_details = [{'Style': r['STYLE'], 'Colour': r['COLOR'], 'Size': r['SIZE'], 'qty': r['qty']} for r in tran_grouped.to_dict(orient='records')]
        
    return jsonify({
        "ok": True,
        "store_code": store_code,
        "store_name": store_name,
        "current_stock": curr_details,
        "transit": tran_details
    })


@app.route("/api/ebo_export_store/<store_code>")
def ebo_export_store(store_code):
    df_curr = _state.get('df_ebo_curr')
    df_tran = _state.get('df_ebo_tran')
    
    if df_curr is None or df_tran is None:
        return "EBO Stock Data not loaded", 404
        
    store_curr = df_curr[df_curr['store_code'] == store_code] if not df_curr.empty else pd.DataFrame()
    store_tran = df_tran[df_tran['Store Code'] == store_code] if not df_tran.empty else pd.DataFrame()
    
    if store_curr.empty and store_tran.empty:
        return f"No stock or transit data found for store code {store_code}", 404
        
    store_name = "Store"
    if not store_curr.empty:
        store_name = store_curr['Store Name'].iloc[0]
    elif not store_tran.empty:
        store_name = store_tran['EBO Name'].iloc[0]
        
    safe_store = "".join(c if c.isalnum() or c in ("-","_") else "_" for c in store_name)
    output_filename = f"EBO_Stock_Transit_{safe_store}_{store_code}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if not store_curr.empty:
            cols_to_export = [c for c in store_curr.columns if c != 'store_code']
            store_curr[cols_to_export].to_excel(writer, index=False, sheet_name="Current Stock")
        else:
            pd.DataFrame(columns=["Store Name","store_code","Department","SKU","Style","Colour","Size","quantity","day"]).to_excel(writer, index=False, sheet_name="Current Stock")
            
        if not store_tran.empty:
            cols_to_export = [c for c in store_tran.columns if c != 'Store Code']
            store_tran[cols_to_export].to_excel(writer, index=False, sheet_name="In Transit")
        else:
            pd.DataFrame(columns=["Store Code","EBO Name","ITEM_CODE","DIVISION","SECTION","DEPARTMENT","BARCODE","SKU","STOCK STATUS","STYLE","COLOR","SIZE","CCODE","ITEM NAME","Transit Qty"]).to_excel(writer, index=False, sheet_name="In Transit")
            
    return send_file(output_path, as_attachment=True, download_name=output_filename)


@app.route("/api/ebo_export_all")
def ebo_export_all():
    df_curr = _state.get('df_ebo_curr')
    df_tran = _state.get('df_ebo_tran')
    
    if df_curr is None or df_tran is None:
        return "EBO Stock Data not loaded", 404
        
    output_filename = "EBO_Stock_and_Transit_Master_Report.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)
    
    # Build GRC map for export
    grc_map_export = {}
    if os.path.exists(EBO_GRC_PATH):
        try:
            df_grc_exp = pd.read_excel(EBO_GRC_PATH)
            df_grc_exp['Store_Code'] = df_grc_exp['Store_Code'].apply(clean_code)
            df_grc_exp['GRC Date'] = pd.to_datetime(df_grc_exp['GRC Date'], errors='coerce')
            latest_grc_exp = df_grc_exp.groupby('Store_Code')['GRC Date'].max().reset_index()
            for _, r in latest_grc_exp.iterrows():
                if pd.notna(r['GRC Date']):
                    grc_map_export[str(r['Store_Code'])] = r['GRC Date'].strftime('%d-%b-%Y')
        except Exception:
            pass

    # Add GRC Date column to current stock sheet
    df_curr_export = df_curr.copy()
    df_curr_export['GRC Date'] = df_curr_export['store_code'].apply(lambda x: grc_map_export.get(str(x), '-'))
    cols_curr = [c for c in df_curr_export.columns if c != 'store_code']
    # Move GRC Date right after Store Name
    if 'GRC Date' in cols_curr:
        cols_curr.remove('GRC Date')
        insert_pos = cols_curr.index('Store Name') + 1 if 'Store Name' in cols_curr else 1
        cols_curr.insert(insert_pos, 'GRC Date')

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_curr_export[cols_curr].to_excel(writer, index=False, sheet_name="Current Stock")
        
        cols_tran = [c for c in df_tran.columns if c != 'Store Code']
        df_tran[cols_tran].to_excel(writer, index=False, sheet_name="In Transit")
        
    return send_file(output_path, as_attachment=True, download_name=output_filename)


@app.route("/api/new_style_performance")
def new_style_performance():
    force_refresh = request.args.get('refresh', 'false').lower() == 'true'
    
    records = _state.get('new_style_data')
    filename = _state.get('new_style_filename')
    
    if force_refresh or not records:
        ok, error = _load_new_style_data_internal()
        if not ok:
            return jsonify({"ok": False, "error": error})
        records = _state.get('new_style_data')
        filename = _state.get('new_style_filename')
        
    return jsonify({
        "ok": True,
        "filename": filename,
        "data": records
    })


@app.route("/api/style_images")
def style_images():
    force_refresh = request.args.get('refresh', 'false').lower() == 'true'
    mapping = _state.get('style_images')
    
    if force_refresh or not mapping:
        ok, error = _load_style_images()
        if not ok:
            return jsonify({"ok": False, "error": error})
        mapping = _state.get('style_images', {})
        
    return jsonify({
        "ok": True,
        "images": mapping
    })


@app.route("/api/campaign_performance")
def campaign_performance():
    force_refresh = request.args.get('refresh', 'false').lower() == 'true'
    
    records = _state.get('campaign_data')
    if force_refresh or not records:
        ok, error = _load_campaign_data_internal(force=force_refresh)
        if not ok:
            return jsonify({"ok": False, "error": error})
        records = _state.get('campaign_data')
        
    return jsonify({
        "ok": True,
        "campaigns": records.get("campaigns", []),
        "campaign_styles": records.get("campaign_styles", {})
    })


@app.route("/api/campaign_reload", methods=["POST"])
def campaign_reload():
    ok, error = _load_campaign_data_internal(force=True)
    return jsonify({
        "ok": ok,
        "error": error
    })


@app.route("/api/style_stock_details")
def style_stock_details():
    style = request.args.get("style", "").strip().upper()
    if not style:
        return jsonify({"ok": False, "error": "No style parameter provided."})
        
    df_curr = _state.get('df_ebo_curr')
    df_tran = _state.get('df_ebo_tran')
    if df_curr is None or df_tran is None or df_curr.empty or df_tran.empty:
        _load_ebo_data_internal()
        df_curr = _state.get('df_ebo_curr')
        df_tran = _state.get('df_ebo_tran')

    # Get all EBO stores from ebo_summary
    ebo_summary = _state.get('ebo_summary')
    if not ebo_summary:
        _load_ebo_data_internal()
        ebo_summary = _state.get('ebo_summary')
        
    all_stores = {}
    if ebo_summary:
        for s in ebo_summary:
            code = str(s.get('store_code', '')).strip()
            name = str(s.get('Store Name', '')).strip()
            if code and code.lower() != 'nan':
                all_stores[code] = name

    curr_list = []
    curr_store_codes = set()
    if df_curr is not None and not df_curr.empty:
        df_style_curr = df_curr[df_curr['Style'].astype(str).str.strip().str.upper() == style]
        if not df_style_curr.empty:
            grouped = df_style_curr.groupby('store_code').agg(
                store_name=('Store Name', 'first'),
                qty=('quantity', 'sum')
            ).reset_index()
            grouped = grouped[grouped['qty'] > 0].sort_values(by='qty', ascending=False)
            for _, r in grouped.iterrows():
                code = str(r['store_code']).strip()
                curr_list.append({
                    "store_code": code,
                    "store_name": str(r['store_name']).strip(),
                    "qty": int(r['qty'])
                })
                curr_store_codes.add(code)

    tran_list = []
    tran_store_codes = set()
    if df_tran is not None and not df_tran.empty:
        df_style_tran = df_tran[df_tran['STYLE'].astype(str).str.strip().str.upper() == style]
        if not df_style_tran.empty:
            grouped = df_style_tran.groupby('Store Code').agg(
                store_name=('EBO Name', 'first'),
                qty=('Transit Qty', 'sum')
            ).reset_index()
            grouped = grouped[grouped['qty'] > 0].sort_values(by='qty', ascending=False)
            for _, r in grouped.iterrows():
                code = str(r['Store Code']).strip()
                tran_list.append({
                    "store_code": code,
                    "store_name": str(r['store_name']).strip(),
                    "qty": int(r['qty'])
                })
                tran_store_codes.add(code)

    # Load store-wise allocation for this style
    alloc_list = []
    alloc_store_codes = set()
    try:
        import glob
        alloc_dir = os.path.join(EBO_STOCK_DIR, "ALLOCATION")
        alloc_files = glob.glob(os.path.join(alloc_dir, "*.xlsx"))
        alloc_files = [f for f in alloc_files if not os.path.basename(f).startswith('~')]
        if alloc_files:
            latest_alloc = max(alloc_files, key=os.path.getmtime)
            df_alloc = _read_excel_fast(latest_alloc)
            df_alloc.columns = [str(c).strip() for c in df_alloc.columns]
            sc_col = next((c for c in df_alloc.columns if 'store' in c.lower() and 'code' in c.lower()), None)
            style_col_a = next((c for c in df_alloc.columns if c.strip().lower() == 'style'), None)
            qty_col_a = next((c for c in df_alloc.columns if 'allocated' in c.lower() and 'qty' in c.lower()), None)
            sn_col = next((c for c in df_alloc.columns if 'store' in c.lower() and 'name' in c.lower()), None)
            if sc_col and style_col_a and qty_col_a:
                df_alloc[sc_col] = df_alloc[sc_col].apply(clean_code)
                df_alloc[style_col_a] = df_alloc[style_col_a].astype(str).str.strip().str.upper()
                df_alloc[qty_col_a] = pd.to_numeric(df_alloc[qty_col_a], errors='coerce').fillna(0)
                
                df_style_alloc = df_alloc[df_alloc[style_col_a] == style]
                if not df_style_alloc.empty:
                    grouped_alloc = df_style_alloc.groupby(sc_col)[qty_col_a].sum().reset_index()
                    grouped_alloc = grouped_alloc[grouped_alloc[qty_col_a] > 0].sort_values(by=qty_col_a, ascending=False)
                    for _, r in grouped_alloc.iterrows():
                        code = str(r[sc_col]).strip()
                        name = all_stores.get(code)
                        if not name and sn_col:
                            match_rows = df_style_alloc[df_style_alloc[sc_col] == code]
                            if not match_rows.empty:
                                name = str(match_rows.iloc[0][sn_col]).strip()
                        if not name:
                            name = f"Store {code}"
                        alloc_list.append({
                            "store_code": code,
                            "store_name": name,
                            "qty": int(r[qty_col_a])
                        })
                        alloc_store_codes.add(code)
    except Exception as ex:
        print("Error loading store-wise allocation details:", ex)

    # Find stores that have NO stock, NO transit, and NO allocation
    no_stock_list = []
    for code, name in all_stores.items():
        if code not in curr_store_codes and code not in tran_store_codes and code not in alloc_store_codes:
            no_stock_list.append({
                "store_code": code,
                "store_name": name
            })
    
    no_stock_list.sort(key=lambda x: x['store_name'].lower())

    return jsonify({
        "ok": True,
        "style": style,
        "current_stock": curr_list,
        "transit_stock": tran_list,
        "alloc_stock": alloc_list,
        "no_stock_stores": no_stock_list,
        "total_current": sum(item['qty'] for item in curr_list),
        "total_transit": sum(item['qty'] for item in tran_list),
        "total_alloc": sum(item['qty'] for item in alloc_list),
        "total_no_stock": len(no_stock_list)
    })


@app.route("/api/campaign_export")
def campaign_export():
    campaign = request.args.get("campaign", "").strip()
    if not campaign:
        return "Missing campaign name", 400
        
    records = _state.get('campaign_data')
    if not records:
        ok, error = _load_campaign_data_internal(force=False)
        if not ok:
            return f"Error loading campaign data: {error}", 500
        records = _state.get('campaign_data')
        
    campaign_styles = records.get("campaign_styles", {})
    styles = campaign_styles.get(campaign, [])
    
    if not styles:
        return f"No styles found for campaign: {campaign}", 404
        
    # Build dataframe
    data = []
    for row in styles:
        data.append({
            "Style Number": row.get("STYLE"),
            "JFM D2C Sales": row.get("JFM_D2C"),
            "JFM EBO Sales": row.get("JFM_EBO"),
            "AMJ D2C Sales": row.get("AMJ_D2C"),
            "AMJ EBO Sales": row.get("AMJ_EBO"),
            "Total D2C Sales": row.get("Total_D2C"),
            "Total EBO Sales": row.get("Total_EBO"),
            "Total Sales Qty": row.get("Total_Sale_Qty"),
            "Store Stock": row.get("Store_Stock"),
            "Transit Stock": row.get("Transit_Stock"),
            "Allocation Stock": row.get("Allocation_Stock"),
            "D2C Stock (OMS)": row.get("D2C_Stock"),
            "CRED Stock (OMS)": row.get("CRED_Stock"),
            "EBO Stock (OMS)": row.get("EBO_Stock"),
            "Common Pool Stock (OMS)": row.get("Common_Pool_Stock"),
            "LRN Date": row.get("LRN_Date"),
            "GRN Date": row.get("GRN_Date"),
            "NOD": row.get("NOD")
        })
        
    df_export = pd.DataFrame(data)
    
    # Sort by Total Sales Qty descending
    df_export.sort_values(by="Total Sales Qty", ascending=False, inplace=True)
    
    # Save to a BytesIO object using openpyxl
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name=campaign[:31].replace("/", "-"), index=False)
        
        # Apply standard auto-fit columns and nice formatting
        workbook = writer.book
        worksheet = writer.sheets[campaign[:31].replace("/", "-")]
        
        # Add basic header formatting using openpyxl directly
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # Style headers
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Style data rows and auto-fit columns
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            max_len = len(col_name)
            for row_idx in range(2, len(df_export) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                val = cell.value
                max_len = max(max_len, len(str(val or "")))
                
                # Alignments based on type/column
                if isinstance(val, (int, float)):
                    cell.alignment = right_align
                elif col_name in ("Style Number", "LRN Date", "GRN Date"):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            
            # Set column width
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
    output.seek(0)
    
    filename = f"{secure_filename(campaign)}_Campaign_Export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/upload_top100_sales", methods=["POST"])
def upload_top100_sales():
    try:
        if 'sales_file' not in request.files:
            return jsonify({"ok": False, "error": "No file uploaded."})
        f = request.files['sales_file']
        if f.filename == '':
            return jsonify({"ok": False, "error": "No file selected."})
            
        sales_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\input of top 100 style"
        if not os.path.exists(sales_dir):
            os.makedirs(sales_dir, exist_ok=True)
            
        # Clean old files in that folder
        import glob
        for old_file in glob.glob(os.path.join(sales_dir, "*.csv")) + glob.glob(os.path.join(sales_dir, "*.xlsx")):
            try:
                os.remove(old_file)
            except Exception:
                pass
                
        # Save new file
        from werkzeug.utils import secure_filename
        dest_path = os.path.join(sales_dir, secure_filename(f.filename))
        f.save(dest_path)
        
        # Run top_100_sales.py script to process the uploaded file and write output to the output folder
        import subprocess
        script_path = r"D:\INCREFF ORDER PUNCH\top_100_sales.py"
        print(f"Running top_100_sales.py with input: {dest_path}")
        result = subprocess.run(
            [sys.executable, script_path, dest_path],
            capture_output=True,
            text=True,
            check=False
        )
        
        if result.returncode != 0:
            print("Error running top_100_sales.py:")
            print("Stdout:", result.stdout)
            print("Stderr:", result.stderr)
            return jsonify({"ok": False, "error": f"Error running top_100_sales script: {result.stderr or result.stdout}"})
            
        # Load the data using our internal function
        ok, error = _load_top100_data_internal(force=True)
        if not ok:
            return jsonify({"ok": False, "error": error})
            
        data = _state.get('top100_data', [])
        filename = _state.get('top100_filename', f.filename)
        return jsonify({"ok": True, "data": data, "filename": filename})
        
    except Exception as e:
        print("Error uploading/processing top 100 sales data:", e)
        return jsonify({"ok": False, "error": f"Failed to process sales data: {str(e)}"})


@app.route("/api/top100_performance")
def top100_performance():
    _load_top100_data_internal(force=True)
    data = _state.get('top100_data')
    filename = _state.get('top100_filename')
    return jsonify({
        "ok": True,
        "data": data if data else [],
        "filename": filename if filename else ""
    })


@app.route("/api/top100_export", methods=["GET", "POST"])
def top100_export():
    search_query = request.args.get('search', '').strip().lower()
    payload = request.get_json(silent=True) if request.method == 'POST' else {}
    if not search_query and payload:
        search_query = str(payload.get('search', '')).strip().lower()

    data = _state.get('top100_data') or []
    
    if search_query and data:
        data = [r for r in data if search_query in str(r.get('style_no', '')).lower()]
    elif not search_query:
        import glob
        output_dir = r"D:\INCREFF ORDER PUNCH\ebo stock track data\output of top 100"
        latest_file = None
        if os.path.exists(output_dir):
            files = glob.glob(os.path.join(output_dir, "Top_*_Sales_Report_*.xlsx"))
            files = [f for f in files if not os.path.basename(f).startswith('~') and 'LATEST' not in os.path.basename(f)]
            if files:
                latest_file = max(files)
                
        if not latest_file:
            latest_file = os.path.join(output_dir, "Top_100_Sales_Report_LATEST.xlsx")
            
        if os.path.exists(latest_file):
            filename = f"Top_100_Sales_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(latest_file, as_attachment=True, download_name=filename)
        
    if not data:
        return "No Top 100 Styles data available to export.", 400
        
    # Build dataframe fallback
    export_rows = []
    for row in data:
        export_rows.append({
            "Style Number": row.get("style_no"),
            "Last 10 Days Sales": row.get("sales_10"),
            "Last 10 Days Qty": row.get("qty_10"),
            "Last 30 Days Sales": row.get("sales_30"),
            "Last 30 Days Qty": row.get("qty_30"),
            "DRR": row.get("drr"),
            "DOH": row.get("doh"),
            "Valid Options": row.get("valid_options_avg"),
            "No. of Stores Live": row.get("stores_live"),
            "D2C Pool Stock (OMS)": row.get("d2c_pool"),
            "CRED Pool Stock (OMS)": row.get("cred_pool"),
            "EBO Pool Stock (OMS)": row.get("ebo_pool"),
            "Common Pool Stock (OMS)": row.get("common_pool"),
            "GT Pool Stock": row.get("gt_pool")
        })
        
    df_export = pd.DataFrame(export_rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name="Top 100 Styles", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Top 100 Styles"]
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        for col_idx in range(1, len(df_export.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            max_len = len(col_name)
            for row_idx in range(2, len(df_export) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                val = cell.value
                max_len = max(max_len, len(str(val or "")))
                if isinstance(val, (int, float)):
                    cell.alignment = right_align
                elif col_name == "Style Number":
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    output.seek(0)
    filename = f"Top_100_Styles_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ─── INWARDS DETAILS (reads from pre-processed output file) ─────────────────

INWARDS_OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "inward details output")
INWARDS_OUTPUT_FILE = os.path.join(INWARDS_OUTPUT_DIR, "Inwards_Processed.xlsx")
INWARDS_SCRIPT      = os.path.join(os.path.dirname(__file__), "process_inwards.py")


def _load_inwards_processed():
    """
    Read the pre-processed Summary sheet from Inwards_Processed.xlsx.
    Caches result in _state.  Returns (df, file_mtime_str).
    """
    if not os.path.exists(INWARDS_OUTPUT_FILE):
        return pd.DataFrame(), "File not found"

    mtime = os.path.getmtime(INWARDS_OUTPUT_FILE)

    # Use cache if file hasn't changed
    if (_state.get('inwards_df') is not None and
            _state.get('inwards_file_mtime') == mtime):
        return _state['inwards_df'], _state.get('inwards_file_time_str', '')

    df = pd.read_excel(INWARDS_OUTPUT_FILE, sheet_name='Summary', engine='openpyxl')
    df.columns = [str(c).strip() for c in df.columns]

    # Expected cols: GRN Date, GRN Date ISO, Category, Style, Size, Color, MRP, Received Qty
    for col in ['GRN Date', 'GRN Date ISO', 'Category', 'Style', 'Size', 'Color']:
        if col not in df.columns:
            df[col] = ''
        df[col] = df[col].fillna('').astype(str).str.strip().replace({'nan': '', 'NaN': '', '<NA>': ''})

    for col in ['MRP', 'Received Qty']:
        if col not in df.columns:
            df[col] = 0
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    df = df.fillna('')

    file_time_str = datetime.datetime.fromtimestamp(mtime).strftime('%d-%b-%Y %H:%M')
    _state['inwards_df']          = df
    _state['inwards_file_mtime']  = mtime
    _state['inwards_file_time_str'] = file_time_str
    return df, file_time_str


@app.route('/api/inwards_data')
def api_inwards_data():
    try:
        force_reload = request.args.get('reload', '0') == '1'
        if force_reload:
            _state.pop('inwards_df', None)
            _state.pop('inwards_file_mtime', None)

        df, file_time = _load_inwards_processed()

        if df is None or df.empty:
            no_file = not os.path.exists(INWARDS_OUTPUT_FILE)
            return jsonify({
                'ok': True, 'records': [], 'total_raw': 0, 'files_loaded': 0,
                'file_time': '',
                'no_output_file': no_file,
                'filter_meta': {'categories': [], 'styles': [], 'colors': [], 'sizes': []},
                'kpi': {'total_qty': 0, 'unique_styles': 0, 'unique_skus': 0,
                        'unique_cats': 0, 'date_from': '', 'date_to': ''}
            })

        # Apply server-side filters
        date_from = request.args.get('date_from', '').strip()
        date_to   = request.args.get('date_to',   '').strip()
        category  = request.args.get('category',  '').strip()
        style     = request.args.get('style',     '').strip()
        color     = request.args.get('color',     '').strip()
        size      = request.args.get('size',      '').strip()
        search    = request.args.get('search',    '').strip().lower()

        dff = df.copy()
        if date_from:
            dff = dff[dff['GRN Date ISO'] >= date_from]
        if date_to:
            dff = dff[dff['GRN Date ISO'] <= date_to]
        if category:
            cats = [c.strip() for c in category.split(',') if c.strip()]
            if cats: dff = dff[dff['Category'].isin(cats)]
        if style:
            styles = [s.strip() for s in style.split(',') if s.strip()]
            if styles: dff = dff[dff['Style'].isin(styles)]
        if color:
            colors = [c.strip() for c in color.split(',') if c.strip()]
            if colors: dff = dff[dff['Color'].isin(colors)]
        if size:
            sizes = [s.strip() for s in size.split(',') if s.strip()]
            if sizes: dff = dff[dff['Size'].isin(sizes)]
        if search:
            mask = (
                dff['Style'].str.lower().str.contains(search, na=False) |
                dff['Color'].str.lower().str.contains(search, na=False) |
                dff['Category'].str.lower().str.contains(search, na=False)
            )
            dff = dff[mask]

        # Filter dropdowns from FULL dataset
        def clean_list(series):
            return sorted([x for x in series.dropna().unique().tolist() if x and x not in ('nan', 'NaN')])

        filter_meta = {
            'categories': clean_list(df['Category']),
            'styles':     clean_list(df['Style']),
            'colors':     clean_list(df['Color']),
            'sizes':      clean_list(df['Size']),
        }

        # KPIs from filtered data
        total_qty     = int(dff['Received Qty'].sum())
        unique_styles = int(dff['Style'].nunique())
        unique_cats   = int(dff['Category'].nunique())
        dates_valid   = dff['GRN Date ISO'].replace('', None).dropna()
        date_range_from = dates_valid.min() if not dates_valid.empty else ''
        date_range_to   = dates_valid.max() if not dates_valid.empty else ''

        # Fast vectorized output
        dff = dff.fillna('')
        import numpy as np
        dff = dff.replace({np.nan: ''})
        
        # Dynamic Groupby based on active filters
        groupby_cols = ['GRN Date', 'GRN Date ISO', 'Category', 'Style']
        rename_map = {
            'GRN Date':     'grn_date',
            'GRN Date ISO': 'grn_date_iso',
            'Category':     'category',
            'Style':        'style',
            'Received Qty': 'received_qty',
        }

        if color:
            groupby_cols.append('Color')
            rename_map['Color'] = 'color'
        if size:
            groupby_cols.append('Size')
            rename_map['Size'] = 'size'

        agg_df = dff.groupby(groupby_cols, as_index=False)['Received Qty'].sum()
        records = agg_df.rename(columns=rename_map).to_dict(orient='records')

        import pandas as pd
        for r in records:
            for k, v in r.items():
                if pd.isna(v):
                    r[k] = ''

        return jsonify({
            'ok':           True,
            'debug_version': 'V3',
            'records':      records,
            'total_raw':    len(df),
            'files_loaded': 1,
            'file_time':    file_time,
            'filter_meta':  filter_meta,
            'kpi': {
                'total_qty':     total_qty,
                'unique_styles': unique_styles,
                'unique_skus':   0,
                'unique_cats':   unique_cats,
                'date_from':     date_range_from if pd.notna(date_range_from) else '',
                'date_to':       date_range_to if pd.notna(date_range_to) else '',
            }
        })
    except Exception as e:
        import traceback
        return jsonify({'ok': False, 'error': str(e), 'trace': traceback.format_exc()})


@app.route('/api/run_inwards_script', methods=['POST'])
def api_run_inwards_script():
    """Re-run process_inwards.py to refresh the output Excel."""
    import subprocess
    try:
        if not os.path.exists(INWARDS_SCRIPT):
            return jsonify({'ok': False, 'error': f'Script not found: {INWARDS_SCRIPT}'})
        result = subprocess.run(
            ['python', INWARDS_SCRIPT],
            capture_output=True, text=True, timeout=120,
            cwd=os.path.dirname(__file__)
        )
        # Clear cache so next data fetch re-reads the new file
        _state.pop('inwards_df', None)
        _state.pop('inwards_file_mtime', None)
        success = result.returncode == 0
        return jsonify({
            'ok':     success,
            'stdout': result.stdout[-3000:] if result.stdout else '',
            'stderr': result.stderr[-1000:] if result.stderr else '',
        })
    except subprocess.TimeoutExpired:
        return jsonify({'ok': False, 'error': 'Script timed out (>120s).'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/inwards_export')
def api_inwards_export():
    """Stream the pre-processed Inwards_Processed.xlsx directly."""
    try:
        if not os.path.exists(INWARDS_OUTPUT_FILE):
            return jsonify({'ok': False, 'error': 'Output file not found. Please run process_inwards.py first.'})

        # Check if any filters are active — if yes, build filtered in-memory file
        date_from = request.args.get('date_from', '').strip()
        date_to   = request.args.get('date_to',   '').strip()
        category  = request.args.get('category',  '').strip()
        style     = request.args.get('style',     '').strip()
        color     = request.args.get('color',     '').strip()
        size      = request.args.get('size',      '').strip()
        search    = request.args.get('search',    '').strip().lower()

        no_filters = not any([date_from, date_to, category, style, color, size, search])
        if no_filters:
            # Serve the full pre-processed file directly — fastest path
            fname = f"Inwards_Processed_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(INWARDS_OUTPUT_FILE, download_name=fname, as_attachment=True,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # Filtered export
        df, _ = _load_inwards_processed()
        if df is None or df.empty:
            return jsonify({'ok': False, 'error': 'No data.'})

        dff = df.copy()
        if date_from: dff = dff[dff['GRN Date ISO'] >= date_from]
        if date_to:   dff = dff[dff['GRN Date ISO'] <= date_to]
        if category:  dff = dff[dff['Category'] == category]
        if style:     dff = dff[dff['Style'] == style]
        if color:     dff = dff[dff['Color'] == color]
        if size:      dff = dff[dff['Size'] == size]
        if search:
            mask = (
                dff['Style'].str.lower().str.contains(search, na=False) |
                dff['Color'].str.lower().str.contains(search, na=False) |
                dff['Category'].str.lower().str.contains(search, na=False)
            )
            dff = dff[mask]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dff.to_excel(writer, sheet_name='Inwards Filtered', index=False)
            ws = writer.sheets['Inwards Filtered']
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            hdr_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
            hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            for col_idx in range(1, len(dff.columns)+1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = hdr_fill; cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_idx, _ in enumerate(dff.columns, 1):
                max_len = max(
                    len(str(dff.columns[col_idx-1])),
                    *(len(str(ws.cell(r, col_idx).value or '')) for r in range(2, min(len(dff)+2, 202)))
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len+3, 40)
        output.seek(0)
        fname = f"Inwards_Filtered_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, download_name=fname, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})



if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print("=" * 55)
    print(f"  Increff Allocation UI  ->  http://0.0.0.0:{port}")
    print("=" * 55)
    app.run(host="0.0.0.0", port=port, debug=False)

