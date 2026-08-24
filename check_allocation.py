import pandas as pd
import os
import sys

# ──────────────────────────────────────────────────────────
#  FILE PATHS  (edit if files move)
# ──────────────────────────────────────────────────────────
REPLEN_FILE    = r"DATA\Requirement as of 06.07.2026.xlsx"
PRIORITY_FILE  = r"priority list\Priority list.xlsx"
OUTPUT_DIR     = r"allocation problem out"

REPLEN_SHEET   = "Sheet1"
PRIORITY_SHEET = "Export"

# Sheet1 column names
EBO_COL   = "EBO NAME"
SKU_COL   = "Client SKU Id / EAN"
ALLOC_COL = "Allocated Qty"    # units the EBO needs
AVAIL_COL = "available qty"    # warehouse stock

# ──────────────────────────────────────────────────────────
#  MANUAL NAME MAPPING
#  Priority-list name  -->  Sheet1 EBO NAME
#  Add / fix rows here whenever new stores are added.
# ──────────────────────────────────────────────────────────
PRIORITY_TO_EBO = {
    "SARATH CITY MALL"   : "TSPL SARATH CITY MALL",
    "TSPL-Vizag"         : "TSPL VIZAG EBO",           # actual name in replen file
    "Kuvempunagar Mysore": "TSPL MYSORE 2 EBO",
    "SALEM"              : "SALEM",
    "KATRAJ EBO"         : "TSPL KATRAJ EBO",
    "TSPL-TUP"           : "TSPL-TUP",
    "BASANT NAGAR"       : "TSPL-BESANTNGR-EBO",
    "ERODE"              : "TSPL ERODE STORE",
    "PUNE KHARADI"       : "PUNE-KH-EBO",
    "CHIKKAJALA"         : "TSPL-CHIKKA-EBO",
    "Belgum"             : None,                       # not in current replen plan
    "ROURKELA"           : "TSPL ROURKELA EBO",
    "HADAPSAR"           : "TSPL HADAPSAR EBO",
    "ATTIBELLE"          : "TSPL ATTIBELE EBO",
    "PUNE PIMPLE"        : "PUNE-PIM-EBO",
    "HSR-EBO"            : "HSR-EBO",
    "WAGHOLI"            : "TSPL WAGHOLI EBO",
    "HYD-EBO"            : "HYD-EBO",
    "RS PURAM"           : "TSPL-RS PURAM-EBO",
    "AS RAO NAGAR"       : "TSPL AS NAGAR EBO",
    "VIZIANAGARAM"       : "TSPL VIZIANAGARAM EBO",
    "KOLHAPUR EBO"       : "TSPL-KOLHAPUR-EBO",
    "PONDICHERRY"        : "PONDY-EBO",
    "BILASPUR"           : "TSPL BILASPUR EBO",
    "KHARAGPUR"          : None,                       # not in current replen plan
    "VIJAYAWADA"         : "TSPL-VIJAYAWADA-EBO",
    "HASSAN EBO"         : "TSPL HASSAN EBO",
    "ANANTAPUR"          : "TSPL ANANTAPUR EBO",
    "JABALPUR"           : "TSPL JABALPUR EBO",
    "MYSORE"             : "MYSORE",
    "DODDABALLAPUR"      : "TSPL DODDABALLA EBO",
    "VELLORE"            : "TSPL VELLORE EBO",
    "Hubli (Shirur Park)": "TSPL SHIRUR_PARK EBO",
    "KUKATPALLY HYD"     : "TSPL KUKATPALLY EBO",
    "Govind Nagar"       : "TSPL GOVINDNAGAR EBO",
    "NAMAKKAL"           : "TSPL NAMAKKAL EBO",
    "METTUPALAYAM"       : "TSPL-METTUPALAYAM-EBO",
    "BHAGALPUR"          : "TSPL BHAGALPUR EBO",
    "BHOPAL EBO"         : "TSPL BHOPAL STORE",
    "BALLARI"            : "TSPL-BALLARI-EBO",
    "DURGAPUR"           : "TSPL DURGAPUR STORE",
    "MANI SQUARE MALL"   : "TSPL MANI SQUARE MALL",
    "HUBBALI EBO"        : "TSPL-HUBBALI-EBO",
    "DIVINITY-MALL"      : "TSPL-DIVINITY-MALL",
    "RAIPUR EBO"         : "TSPL RAIPUR EBO",
    "TEX VALLEY"         : None,                       # not in current replen plan
    "SENTRUM MALL"       : "TSPL SENTRUM MALL",
    "MOSHI"              : "TSPL MOUSHI EBO",
    "SHAHEEN BAGH"       : None,                       # not in current replen plan
    "INDORE-EBO"         : "INDORE-EBO",
    "UDUPI"              : "TSPL UDUPI EBO",
    "ELAN EPIC MALL"     : None,                       # not in current replen plan
    "SALEM-2 EBO"        : None,                       # not in current replen plan
}


# ──────────────────────────────────────────────────────────
PRIORITY_NUM_COL = "priority number"   # column name in the priority Excel

def load_data():
    """Load replenishment Sheet1 and the priority list (sorted by priority number)."""
    for path in (REPLEN_FILE, PRIORITY_FILE):
        if not os.path.exists(path):
            print(f"[ERROR] File not found: {path}")
            return None, None

    try:
        df_replen = pd.read_excel(REPLEN_FILE, sheet_name=REPLEN_SHEET)
    except Exception:
        df_replen = pd.read_excel(REPLEN_FILE, sheet_name=0)

    try:
        df_priority = pd.read_excel(PRIORITY_FILE, sheet_name=PRIORITY_SHEET)
    except Exception:
        df_priority = pd.read_excel(PRIORITY_FILE, sheet_name=0)

    # Normalize replenishment columns to match standard ones case-insensitively
    replen_col_mapping = {
        EBO_COL: ["EBO NAME", "ebo name", "Ebo Name", "EBO Name"],
        SKU_COL: ["Client SKU Id / EAN", "client sku id / ean", "Client Sku Id / Ean", "Client SKU ID / EAN"],
        ALLOC_COL: ["Allocated Qty", "allocated qty", "Allocated QTY", "Allocated quantity", "allocated quantity", "Allocated  Qty"],
        AVAIL_COL: ["available qty", "Available Qty", "available Qty", "Available QTY", "available quantity", "available stock"]
    }
    
    rename_replen = {}
    for standard_col, options in replen_col_mapping.items():
        for col in df_replen.columns:
            normalized_col = " ".join(col.split()).lower()
            normalized_std = " ".join(standard_col.split()).lower()
            normalized_opts = [" ".join(opt.split()).lower() for opt in options]
            if normalized_col == normalized_std or normalized_col in normalized_opts:
                rename_replen[col] = standard_col
                break
    df_replen = df_replen.rename(columns=rename_replen)

    # Normalize priority columns
    priority_col_mapping = {
        "Store Name": ["Store Name", "store name", "STORE NAME", "StoreName", "store_name"],
        PRIORITY_NUM_COL: ["priority number", "priority", "Priority", "Priority No", "Priority Number", "priority_number", "priority list"]
    }
    
    rename_priority = {}
    for standard_col, options in priority_col_mapping.items():
        for col in df_priority.columns:
            if col.strip().lower() == standard_col.lower() or col.strip() in options:
                rename_priority[col] = standard_col
                break
    df_priority = df_priority.rename(columns=rename_priority)

    # Sort by the priority number column so row order in the file doesn't matter
    if PRIORITY_NUM_COL in df_priority.columns:
        df_priority = df_priority.sort_values(PRIORITY_NUM_COL).reset_index(drop=True)
        print(f"[INFO] Priority list loaded: {len(df_priority)} stores, sorted by '{PRIORITY_NUM_COL}'")
    else:
        print(f"[WARN] Column '{PRIORITY_NUM_COL}' not found in priority file. Using row order.")

    return df_replen, df_priority


# ──────────────────────────────────────────────────────────
def simulate_allocation(sku, df_replen, df_priority):
    """
    Walk the priority list in priority-number order.
    Deduct each EBO's Allocated Qty from the available stock.
    Collect EBOs that end up with zero stock (shortfall).

    Returns:
        shortfall_df  - DataFrame of unfulfilled EBOs
        total_stock   - original warehouse qty
    """
    sku_rows = df_replen[df_replen[SKU_COL] == sku].copy()

    if sku_rows.empty:
        print(f"[WARN] SKU '{sku}' not found in the replenishment file.")
        return pd.DataFrame(), 0

    ebo_need = (
        sku_rows.groupby(EBO_COL, as_index=False)
        .agg(needed=(ALLOC_COL, "sum"), avail=(AVAIL_COL, "first"))
    )

    total_stock = int(ebo_need["avail"].iloc[0])
    remaining   = total_stock

    print()
    print("=" * 60)
    print(f"  SKU            : {sku}")
    print(f"  Available Stock: {total_stock}")
    print(f"  EBOs needing this SKU: {len(ebo_need)}")
    print("=" * 60)
    print(f"  {'Pri':<5} {'EBO Name':<30} {'Need':>5} {'Got':>5} {'Short':>6}  Status")
    print("  " + "-" * 58)

    need_lookup = dict(zip(ebo_need[EBO_COL], ebo_need["needed"]))
    shortfall_rows = []
    seen_ebos = set()   # guard against duplicate mappings (same EBO, two priority names)

    # Iterate rows already sorted by priority number
    for _, row in df_priority.iterrows():
        priority_name = row["Store Name"]
        pri_num       = int(row[PRIORITY_NUM_COL]) if PRIORITY_NUM_COL in row.index else "-"
        ebo_name      = PRIORITY_TO_EBO.get(priority_name)

        if ebo_name is None or ebo_name not in need_lookup:
            continue
        if ebo_name in seen_ebos:
            continue          # duplicate mapping — already allocated with higher priority
        seen_ebos.add(ebo_name)

        needed = int(need_lookup[ebo_name])

        if remaining <= 0:
            given, shortfall, status = 0, needed, "No Stock"
        elif remaining >= needed:
            given = needed; remaining -= needed; shortfall = 0; status = "Fully Met"
        else:
            given = remaining; shortfall = needed - given; remaining = 0; status = "Partial"

        print(f"  {pri_num:<5} {ebo_name:<30} {needed:>5} {given:>5} {shortfall:>6}  {status}")

        if shortfall > 0:
            shortfall_rows.append({
                "Priority No"    : pri_num,
                "EBO Name"       : ebo_name,
                "Needed Qty"     : needed,
                "Allocated Qty"  : given,
                "Shortfall Qty"  : shortfall,
            })

    # Catch any EBOs that need stock but are missing from the priority list
    for ebo_name, needed in need_lookup.items():
        if ebo_name not in seen_ebos:
            needed = int(needed)
            given, shortfall, status = 0, needed, "Unmapped"
            print(f"  {'-':<5} {ebo_name:<30} {needed:>5} {given:>5} {shortfall:>6}  {status}")
            shortfall_rows.append({
                "Priority No"    : "-",
                "EBO Name"       : ebo_name,
                "Needed Qty"     : needed,
                "Allocated Qty"  : given,
                "Shortfall Qty"  : shortfall,
            })

    print("  " + "-" * 58)
    print(f"  Remaining stock after all allocations: {remaining}")
    print()

    return pd.DataFrame(shortfall_rows), total_stock


# ──────────────────────────────────────────────────────────
def _apply_sheet_style(ws, df, title_text, n_extra_cols=0):
    """
    Common styling helper used by both single-SKU and all-SKU reports.
    Expects data written at startrow=2 (header on row 3, data from row 4).
    n_extra_cols: number of columns beyond the base 5 (for the all-SKU report).
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils  import get_column_letter

    hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    red_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    c_align  = Alignment(horizontal="center", vertical="center")
    l_align  = Alignment(horizontal="left",   vertical="center")

    total_cols = 5 + n_extra_cols

    # Title row (no merge to avoid Excel corruption bugs)
    ws["A1"] = title_text
    ws["A1"].font      = Font(bold=True, size=13, color="1F3864")
    ws["A1"].alignment = l_align
    ws.row_dimensions[1].height = 26

    # Header row (row 3)
    header_row = 3
    for c_idx in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=c_idx)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = c_align
        cell.border    = border
    ws.row_dimensions[header_row].height = 24

    # Data rows (row 4 onwards)
    for r_idx in range(header_row + 1, header_row + 1 + len(df)):
        fill = red_fill if (r_idx % 2 == 0) else alt_fill
        for c_idx in range(1, total_cols + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill      = fill
            cell.border    = border
            # Column 2 = EBO Name -> left aligned; all others -> center
            cell.alignment = l_align if c_idx == 2 else c_align

    ws.freeze_panes = f"A{header_row + 1}"


# ──────────────────────────────────────────────────────────
def save_excel(shortfall_df, sku, total_stock):
    """Write a styled Excel report for a single SKU."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    from openpyxl.utils import get_column_letter

    safe_sku    = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in sku)
    output_path = os.path.join(OUTPUT_DIR, f"check_allocation_{safe_sku}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        shortfall_df.to_excel(writer, index=False, sheet_name="Shortfall Report", startrow=2)
        ws = writer.sheets["Shortfall Report"]

        widths = [14, 34, 14, 16, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _apply_sheet_style(
            ws, shortfall_df,
            f"Allocation Shortfall Report  |  SKU: {sku}  |  Available Stock: {total_stock}"
        )

    print(f"[OUTPUT] Report saved: {output_path}")
    return output_path


# ──────────────────────────────────────────────────────────
def save_all_excel(master_df):
    """
    Write the consolidated ALL-SKUs report.
    Sheet 1 : full summary (every SKU x EBO shortfall row)
    Sheet 2+: one sheet per EBO that has any shortfall
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    from openpyxl.utils  import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    import time
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"ALL_SKU_shortfall_report_{timestamp}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet 1: Full Summary ──────────────────────────────────────
        master_df.to_excel(writer, index=False, sheet_name="All Shortfalls", startrow=2)
        ws_all = writer.sheets["All Shortfalls"]

        col_widths_all = [18, 34, 14, 16, 16, 16]
        for i, w in enumerate(col_widths_all, start=1):
            ws_all.column_dimensions[get_column_letter(i)].width = w

        _apply_sheet_style(
            ws_all, master_df,
            "All SKU Allocation Shortfall Report  |  All EBOs  |  All SKUs",
            n_extra_cols=1   # 6 columns total (SKU column added)
        )

        # ── Sheet 2+ : one tab per EBO ────────────────────────────────
        ebo_list = master_df["EBO Name"].unique()

        for ebo in ebo_list:
            ebo_df = master_df[master_df["EBO Name"] == ebo].copy()

            # Safe sheet name (max 31 chars, no special chars)
            safe_name = ebo[:31].replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "")

            # Drop EBO Name column (redundant per-EBO sheet)
            ebo_df_out = ebo_df.drop(columns=["EBO Name"]).reset_index(drop=True)

            ebo_df_out.to_excel(writer, index=False, sheet_name=safe_name, startrow=2)
            ws_ebo = writer.sheets[safe_name]

            col_widths_ebo = [18, 14, 16, 16, 16]
            for i, w in enumerate(col_widths_ebo, start=1):
                ws_ebo.column_dimensions[get_column_letter(i)].width = w

            _apply_sheet_style(
                ws_ebo, ebo_df_out,
                f"Shortfall Report  |  EBO: {ebo}  |  Total Shortfall SKUs: {len(ebo_df_out)}",
                n_extra_cols=0
            )

    print(f"[OUTPUT] Consolidated report saved: {output_path}")
    return output_path


# ──────────────────────────────────────────────────────────
def simulate_allocation_silent(sku, df_replen, df_priority):
    """
    Same logic as simulate_allocation() but no console output.
    Used in batch/run-all mode.
    """
    sku_rows = df_replen[df_replen[SKU_COL] == sku].copy()
    if sku_rows.empty:
        return pd.DataFrame(), 0

    ebo_need = (
        sku_rows.groupby(EBO_COL, as_index=False)
        .agg(needed=(ALLOC_COL, "sum"), avail=(AVAIL_COL, "first"))
    )

    total_stock = int(ebo_need["avail"].iloc[0])
    remaining   = total_stock
    need_lookup = dict(zip(ebo_need[EBO_COL], ebo_need["needed"]))

    shortfall_rows = []

    for _, row in df_priority.iterrows():
        priority_name = row["Store Name"]
        pri_num       = int(row[PRIORITY_NUM_COL]) if PRIORITY_NUM_COL in row.index else 0
        ebo_name      = PRIORITY_TO_EBO.get(priority_name)

        if ebo_name is None or ebo_name not in need_lookup:
            continue

        needed = int(need_lookup[ebo_name])

        if remaining <= 0:
            given, shortfall = 0, needed
        elif remaining >= needed:
            given = needed; remaining -= needed; shortfall = 0
        else:
            given = remaining; shortfall = needed - given; remaining = 0

        if shortfall > 0:
            shortfall_rows.append({
                "SKU"           : sku,
                "Priority No"   : pri_num,
                "EBO Name"      : ebo_name,
                "Needed Qty"    : needed,
                "Allocated Qty" : given,
                "Shortfall Qty" : shortfall,
            })

    return pd.DataFrame(shortfall_rows), total_stock


# ──────────────────────────────────────────────────────────
def run_all(df_replen, df_priority):
    """
    Internally builds an EBO x SKU pivot, runs allocation from it,
    and produces the shortfall Excel — no pivot added to the output file.

    Pivot (internal only):
        rows    = EBO NAME  (will be walked in priority order)
        columns = Client SKU Id / EAN
        values  = sum of Allocated Qty  (how much each EBO needs per SKU)
    """
    print("Building internal EBO x SKU pivot...")

    # ── Build pivot ───────────────────────────────────────────────────
    # rows=EBO, cols=SKU, values=needed qty
    pivot_need = (
        df_replen.groupby([EBO_COL, SKU_COL], as_index=False)
        .agg(needed=(ALLOC_COL, "sum"), avail=(AVAIL_COL, "first"))
        .pivot(index=EBO_COL, columns=SKU_COL, values="needed")
        .fillna(0)
    )

    # Available stock per SKU (one value per SKU, same across all EBOs)
    avail_per_sku = df_replen.groupby(SKU_COL)[AVAIL_COL].first()

    # Build ordered priority list: (priority_num, ebo_sheet1_name)
    # Only keep EBOs in the pivot; DEDUPLICATE — if two priority-list names map
    # to the same Sheet1 EBO, keep only the first (highest priority) occurrence.
    seen_ebos    = set()
    ordered_ebos = []
    for _, row in df_priority.iterrows():
        priority_name = row["Store Name"]
        ebo_name = PRIORITY_TO_EBO.get(priority_name)
        if not ebo_name:
            if priority_name in PRIORITY_TO_EBO.values():
                ebo_name = priority_name
            else:
                for idx in pivot_need.index:
                    if str(idx).strip().lower() == str(priority_name).strip().lower():
                        ebo_name = idx
                        break
        if not ebo_name:
            ebo_name = priority_name  # fallback

        if ebo_name is None or ebo_name not in pivot_need.index:
            continue
        if ebo_name in seen_ebos:
            continue                          # already added with higher priority
        seen_ebos.add(ebo_name)

        pri_val = row[PRIORITY_NUM_COL]
        try:
            pri_num = int(pri_val)
        except:
            pri_num = 999
        ordered_ebos.append((pri_num, ebo_name))

    all_skus = sorted(pivot_need.columns)
    total    = len(all_skus)
    print(f"  Pivot ready: {len(pivot_need)} EBOs x {total} SKUs")
    print(f"  Running allocation from pivot...")
    print("-" * 40)

    all_rows        = []
    skus_with_short = 0
    skus_fully_ok   = 0

    for i, sku in enumerate(all_skus, start=1):
        total_stock = int(avail_per_sku.get(sku, 0))
        remaining   = total_stock
        sku_short   = []

        for pri_num, ebo_name in ordered_ebos:
            # Read directly from pivot cell — O(1) lookup, no DataFrame filter
            needed = int(pivot_need.at[ebo_name, sku]) if ebo_name in pivot_need.index else 0
            if needed == 0:
                continue

            if remaining <= 0:
                given, shortfall = 0, needed
            elif remaining >= needed:
                given = needed; remaining -= needed; shortfall = 0
            else:
                given = remaining; shortfall = needed - given; remaining = 0

            if shortfall > 0:
                sku_short.append({
                    "SKU"           : sku,
                    "Priority No"   : pri_num,
                    "EBO Name"      : ebo_name,
                    "Needed Qty"    : needed,
                    "Allocated Qty" : given,
                    "Shortfall Qty" : shortfall,
                })

        # Second pass: EBOs missing from priority mapping
        for ebo_name in pivot_need.index:
            if ebo_name not in seen_ebos:
                needed = int(pivot_need.at[ebo_name, sku])
                if needed > 0:
                    given, shortfall = 0, needed
                    sku_short.append({
                        "SKU"           : sku,
                        "Priority No"   : "-",
                        "EBO Name"      : ebo_name,
                        "Needed Qty"    : needed,
                        "Allocated Qty" : given,
                        "Shortfall Qty" : shortfall,
                    })

        if sku_short:
            all_rows.extend(sku_short)
            skus_with_short += 1
            print(f"  [{i:>4}/{total}] {sku:<25}  shortfall in {len(sku_short)} EBO(s)")
        else:
            skus_fully_ok += 1
            print(f"  [{i:>4}/{total}] {sku:<25}  fully met")

    print("-" * 40)

    if not all_rows:
        print("All SKUs are fully fulfilled across all EBOs. No report generated.")
        return

    master_df   = pd.DataFrame(all_rows)
    output_path = save_all_excel(master_df)

    print()
    print("=" * 60)
    print("  ALL-SKU RUN COMPLETE")
    print("=" * 60)
    print(f"  Total SKUs processed : {total}")
    print(f"  SKUs with shortfall  : {skus_with_short}")
    print(f"  SKUs fully fulfilled : {skus_fully_ok}")
    print(f"  Total shortfall rows : {len(master_df)}")
    print(f"  EBOs affected        : {master_df['EBO Name'].nunique()}")
    print(f"  Report saved to      : {output_path}")
    print("=" * 60)


# ──────────────────────────────────────────────────────────
def main():
    """
    Usage:
      python check_allocation.py           -> prompts for a single SKU
      python check_allocation.py <SKU>     -> runs for that specific SKU
      python check_allocation.py --all     -> processes every SKU and saves combined report
    """
    df_replen, df_priority = load_data()
    if df_replen is None:
        return

    # Determine mode
    arg = sys.argv[1].strip() if len(sys.argv) > 1 else ""

    if arg == "--all" or arg == "":
        if arg == "":
            choice = input("Enter SKU  OR  type 'all' to run for every SKU: ").strip()
        else:
            choice = "all"

        if choice.lower() == "all":
            run_all(df_replen, df_priority)
            return
        else:
            sku = choice
    else:
        sku = arg

    if not sku:
        print("[ERROR] No SKU entered.")
        return

    # Single SKU mode
    shortfall_df, total_stock = simulate_allocation(sku, df_replen, df_priority)

    if shortfall_df.empty:
        print(f"All EBOs that need SKU '{sku}' are FULLY FULFILLED.")
        print("No Excel report generated.")
        return

    output_path = save_excel(shortfall_df, sku, total_stock)

    print("=" * 60)
    print("  SUMMARY")
    print("=" * 60)
    print(f"  SKU              : {sku}")
    print(f"  Total Stock      : {total_stock}")
    print(f"  EBOs shortlisted : {len(shortfall_df)}")
    print(f"  Report saved to  : {output_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
